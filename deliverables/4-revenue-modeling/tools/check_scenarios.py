"""
Scenario-direction check.

Flips the scenario selector to Base / Aggressive / Conservative, recalculates
each, and asserts the revenue ordering is Conservative < Base < Aggressive.

WHY THIS IS A SEPARATE TOOL. verify_model.py reads ONE already-recalculated
workbook and never invokes LibreOffice; this has to recalculate three times, so
it does not belong there. Run it before any delivery, and after touching a
scenario band.

WHAT IT CATCHES that nothing else can. A scenario parameter whose aggressive
and conservative columns are the wrong way round is INVISIBLE to every other
check: the value is in range, the mirror matches, the parameter is used, the
formula is right. It only shows up when the scenario is actually flipped.
Found exactly that in "Programme manager share of interchange" on 2026-08-21 -
aggressive handed MORE interchange to the programme manager, so the optimistic
case cut Aurumix's own revenue by 46%.

  Usage:  python check_scenarios.py
"""
import json
import os
import shutil
import subprocess
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "Aurumix_Revenue_Model.xlsx")
TMP = os.path.join(HERE, "_scenario_probe.xlsx")
SCENARIOS = [(1, "Base"), (2, "Aggressive"), (3, "Conservative")]

# Rows worth reporting. Total revenue is what the ordering is asserted on; the
# streams are printed so a single misbehaving one is visible even when the
# total happens to come out in the right order.
WATCH = ["TOTAL NET REVENUE",
         "Stream 1a - Entry fee, SIP", "Stream 1b - Entry fee, SPOT",
         "Stream 2 - Card interchange", "Stream 3 - Family plan and Digital Will",
         "Stream 4 - Cardholder fees", "Stream 5 - Lending revenue share",
         "Stream 6 - B2B platform fee",
         "PAYING CUSTOMERS", "GRAMS UNDER CUSTODY"]


def selector_cell(path):
    """Find the scenario index cell rather than hardcoding it, so this survives
    rows being inserted above the selector."""
    ws = load_workbook(path, data_only=False)["Scenario Parameters"]
    for r in range(1, 30):
        if ws["A%d" % r].value == "Select Scenario:":
            return "C%d" % r
    raise AssertionError("scenario selector not found on Scenario Parameters")


def run(idx):
    shutil.copy(SRC, TMP)
    wb = load_workbook(TMP, data_only=False)
    wb["Scenario Parameters"][CELL] = idx
    wb.save(TMP)
    wb.close()
    p = subprocess.run([sys.executable, os.path.join(HERE, "recalc_lo.py"), TMP],
                       capture_output=True, text=True, cwd=HERE)
    # recalc_lo prints a JSON object, but may print other lines around it - take
    # the outermost braces rather than assuming the whole of stdout is JSON.
    out = p.stdout or ""
    assert "{" in out and "}" in out, "recalc produced no JSON:\n%s\n%s" % (out[-800:], (p.stderr or "")[-800:])
    d = json.loads(out[out.index("{"):out.rindex("}") + 1])
    assert d["status"] == "success" and d["total_errors"] == 0, d
    m = load_workbook(os.path.join(HERE, "_recalc", os.path.basename(TMP)),
                      data_only=True)["Model"]
    out = {}
    for lbl in WATCH:
        for i in range(1, m.max_row + 1):
            if m["A%d" % i].value == lbl:
                out[lbl] = m.cell(row=i, column=31).value
                break
    return out


CELL = selector_cell(SRC)
print("scenario selector at Scenario Parameters!%s\n" % CELL)
res = {name: run(i) for i, name in SCENARIOS}

print("{:<42} {:>13} {:>13} {:>13}".format("Y7", "Conservative", "Base", "Aggressive"))
print("-" * 84)
for lbl in WATCH:
    c, b, a = res["Conservative"][lbl], res["Base"][lbl], res["Aggressive"][lbl]
    flag = "" if c <= b <= a else "   <-- OUT OF ORDER"
    print("{:<42} {:>13,.0f} {:>13,.0f} {:>13,.0f}{}".format(lbl[:42], c, b, a, flag))

print()
bad = [l for l in WATCH
       if not (res["Conservative"][l] <= res["Base"][l] <= res["Aggressive"][l])]
if bad:
    print("FAIL - %d row(s) do not increase from Conservative to Aggressive:" % len(bad))
    for l in bad:
        print("   %s" % l)
    print("\nA row out of order means at least one scenario band is inverted, or a")
    print("parameter moves that row against the direction of the scenario as a whole.")
    print("Neither is necessarily a defect - but neither should be a surprise.")
    sys.exit(1)
print("PASS - every watched row rises from Conservative through Base to Aggressive.")
for f in (TMP, os.path.join(HERE, "_recalc", os.path.basename(TMP))):
    if os.path.exists(f):
        os.remove(f)
