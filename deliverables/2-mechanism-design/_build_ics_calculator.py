"""Builds Aurumix_ICS_Score_Calculator.xlsx — one tab, laid out for presenting
to the client rather than for auditing.  Mechanics per _draft_ics-scoring.md
(2026-08-12, post flaw-review)."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY, GOLD, INK, MUTE = "1F3864", "BF8F00", "000000", "595959"
BLUE = Font(color="0000FF")
BLUE_B = Font(color="0000FF", bold=True)
BLK = Font(color=INK)
BLK_B = Font(bold=True)
BIG = Font(bold=True, size=18, color=NAVY)
STRAP = Font(bold=True, size=12, color=GOLD)
SUB = Font(italic=True, size=10, color=MUTE)
LEAD = Font(size=11, color=INK)
HDR = Font(bold=True, color="FFFFFF", size=10)
SECT = Font(bold=True, size=13, color="FFFFFF")
Q = Font(bold=True, size=11, color=NAVY)

F_NAVY = PatternFill("solid", start_color=NAVY)
F_GOLD = PatternFill("solid", start_color="FFF2CC")
F_BAND = PatternFill("solid", start_color="F2F2F2")
F_YEL = PatternFill("solid", start_color="FFFF00")
F_GRN = PatternFill("solid", start_color="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(top=thin, bottom=thin, left=thin, right=thin)
COLS = "BCDEFGHIJKLMN"

wb = Workbook()
ws = wb.active
ws.title = "ICS"
r = 1


def put(cell, v, font=None, fmt=None, fill=None, border=False, align=None, wrap=False):
    ws[cell] = v
    if font:
        ws[cell].font = font
    if fmt:
        ws[cell].number_format = fmt
    if fill:
        ws[cell].fill = fill
    if border:
        ws[cell].border = BOX
    if align or wrap:
        ws[cell].alignment = Alignment(horizontal=align or "general", wrap_text=wrap, vertical="center")


def band(row, fill, cols="ABCDEFGHIJKLMN"):
    for c in cols:
        ws["%s%d" % (c, row)].fill = fill


def section(title):
    global r
    r += 1
    put("A%d" % r, "  " + title, SECT)
    band(r, F_NAVY)
    ws.row_dimensions[r].height = 24
    r += 1
    return r - 1


def header(row, labels, cols="ABCDEFGHIJK"):
    for c, h in zip(cols, labels):
        put("%s%d" % (c, row), h, HDR, fill=F_NAVY, align="center")
    ws["%s%d" % (cols[0], row)].alignment = Alignment(horizontal="left")


# ============================================================ title
put("A1", "AURUMIX  ·  THE INVESTOR CONVICTION SCORE", BIG)
ws.row_dimensions[1].height = 26
put("A2", "How a saver earns their tier — and how they can lose it.", STRAP)
put("A3", "Every number below is live. Change anything blue and the whole model recalculates.", SUB)
r = 4

# ============================================================ the idea
section("THE IDEA   —   we ask three questions, and nothing else")
put("A%d" % r, "The score never looks at how much money you have. It looks at what you did.", LEAD)
r += 2
qs = [
    ("1.  What did you contribute?", "RECORD",
     "How many months you have ever paid. This is a fact about your history — nothing reduces it, ever."),
    ("2.  Are you contributing now?", "STANDING",
     "How many of the last 12 months you paid. This is a fact about your present, and it moves both ways."),
    ("3.  Did you keep your gold?", "RETENTION",
     "How much of your gold you still hold. Saving and then selling is not saving."),
]
for q, name, expl in qs:
    put("A%d" % r, q, Q, border=True)
    put("B%d" % r, name, Font(bold=True, color=GOLD), border=True, align="center")
    put("C%d" % r, expl, BLK, border=True)
    band(r, F_GOLD, "AB")
    r += 1
r += 1
put("A%d" % r, "Your score is the WEAKER of the first two, reduced by the third. "
               "A long record cannot cover a bad year, and a perfect year cannot cover a short record.", LEAD)
r += 1

# ============================================================ record
section("QUESTION 1   —   WHAT DID YOU CONTRIBUTE?      gives you your RECORD, out of 100")
put("A%d" % r, "Your first year of payments takes you to 50.  The next four years take you from 50 to 100.", LEAD)
r += 1
put("A%d" % r, "It stops at five years — but the app keeps showing the real count, which rises forever.", SUB)
r += 2
put("A%d" % r, "Months you have paid", BLK_B)
months = [0, 3, 6, 9, 12, 18, 24, 30, 36, 42, 48, 54, 60]
row_m = r
for i, m in enumerate(months):
    put("%s%d" % (COLS[i], r), m, BLUE, fmt="0", border=True, align="center")
r += 1
put("A%d" % r, "Your RECORD", Font(bold=True, color=NAVY))
row_rec = r
for i in range(len(months)):
    c = COLS[i]
    put("%s%d" % (c, r),
        "=IF({0}{1}<=$B${2},{0}{1}*50/$B${2},IF({0}{1}>=$B${3},100,50+({0}{1}-$B${2})*50/($B${3}-$B${2})))"
        .format(c, row_m, "P_M12", "P_M60"), BLK_B, fmt="0", border=True, align="center", fill=F_BAND)
r += 2

# ============================================================ standing
section("QUESTION 2   —   ARE YOU CONTRIBUTING NOW?      gives you your STANDING, out of 100")
put("A%d" % r, "Each of the last 12 months you paid is worth 8.3 points.  Twelve out of twelve is 100.", LEAD)
r += 1
put("A%d" % r, "A month you missed drops out of view a year later, so an isolated lapse heals itself. "
               "A habit of missing does not.", SUB)
r += 2
put("A%d" % r, "Months paid, of the last 12", BLK_B)
row_rc = r
for i in range(13):
    put("%s%d" % (COLS[i], r), i, BLUE, fmt="0", border=True, align="center")
r += 1
put("A%d" % r, "Your STANDING", Font(bold=True, color=NAVY))
row_std = r
for i in range(13):
    c = COLS[i]
    put("%s%d" % (c, r), "=MIN({0}{1},$B${2})*100/$B${2}".format(c, row_rc, "P_WIN"),
        BLK_B, fmt="0", border=True, align="center", fill=F_BAND)
r += 2

# ============================================================ retention
section("QUESTION 3   —   DID YOU KEEP YOUR GOLD?      gives you your RETENTION, a multiplier")
put("A%d" % r, "Take out up to a THIRD of your gold in a year and nothing happens at all.", LEAD)
r += 1
put("A%d" % r, "Past that, every further 7% you sell costs you 10% of your score — down to zero if you empty the account.", LEAD)
r += 2
put("A%d" % r, "We measure it as:      1  −  ( gold you hold now  ÷  [ gold you held a year ago  +  gold you have bought since ] )", Font(italic=True, color=NAVY))
r += 1
put("A%d" % r, "In words: what share of everything you had did you not keep? Three numbers, read straight off the ledger. "
               "It cannot be made cheaper by choosing when to sell.", SUB)
r += 2
put("A%d" % r, "Share sold in the last 12 months", BLK_B)
row_sold = r
for i in range(11):
    put("%s%d" % (COLS[i], r), i / 10.0, BLUE, fmt="0%", border=True, align="center")
r += 1
put("A%d" % r, "Your RETENTION", Font(bold=True, color=NAVY))
row_ret = r
for i in range(11):
    c = COLS[i]
    put("%s%d" % (c, r),
        "=IF({0}{1}<=$B${2},1,MAX(0,1-({0}{1}-$B${2})/(1-$B${2})))".format(c, row_sold, "P_ALW"),
        BLK_B, fmt="0.00", border=True, align="center", fill=F_BAND)
r += 1
put("A%d" % r, "Transfers to a family account or under the Digital Will are NOT sales — the gold stays in the product, only the name on it changes.", SUB)
r += 2

# ============================================================ formula
section("PUTTING THEM TOGETHER")
put("A%d" % r, "YOUR SCORE   =   the LOWER of ( Record , Standing )   ×   Retention", Font(bold=True, size=13, color=NAVY))
band(r, F_GOLD)
ws.row_dimensions[r].height = 22
r += 2
for line in [
    "Why the LOWER of the two?  Because both have to be true. A twenty-year record does not excuse a year of not paying, "
    "and a perfect year does not make you a twenty-year saver. Taking the lower number is what makes it AND rather than OR.",
    "Why does Retention MULTIPLY?  Because keeping your gold is not one good habit among several — it is a condition on all of them. "
    "Sell everything and the score is zero, however perfectly you paid.",
    "One protection: once you have made six payments, your score never falls below 25. You are never sent back to the beginning.",
]:
    put("A%d" % r, "•   " + line, BLK, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1
r += 1

# ============================================================ ladder
section("THE LADDER")
header(r, ["Score needed", "Tier", "In plain English", "What it buys"], "ABCD")
r += 1
row_tier0 = r
ladder = [
    (0, "No tier", "Account open, gate not yet passed. No score is calculated.", "—"),
    (25, "Silver", "Six months straight. You are established, permanently.", "0.4pp off the entry fee · 10% off the will plan"),
    (50, "Gold", "One year paid, and paying now.", "Credit at 50% LTV · card issued · Gold Rewards begin · 0.8pp off"),
    (75, "Platinum", "Three years paid, and a strong year behind you.", "LTV 65% · card level 2 · Rewards 0.45% · 1.2pp off · will 35%"),
    (100, "Sovereign", "Five years paid, a perfect year, gold intact.", "LTV 80% · top card · Rewards 0.75% · 1.5pp off · will 50%"),
]
for bound, name, plain, buys in ladder:
    put("A%d" % r, bound, Font(bold=True, size=12, color=NAVY), fmt="0", border=True, align="center")
    put("B%d" % r, name, Font(bold=True, size=12, color=GOLD), border=True)
    put("C%d" % r, plain, BLK, border=True)
    put("D%d" % r, buys, SUB, border=True)
    r += 1
row_tier1 = r - 1
TIER_RANGE = "$A$%d:$B$%d" % (row_tier0, row_tier1)
put("A%d" % r, "The gate comes first: six consecutive payments before any score exists. Everyone enters at Silver, 25.", SUB)
r += 1
put("A%d" % r, "Silver is permanent. Every tier above it is rented by conduct — which is the point.", SUB)
r += 2

# ============================================================ story
section("A WORKED STORY   —   Rajesh, USD 75 a month")
put("A%d" % r, "One missed payment and one real withdrawal, across five years. Follow the last two columns.", LEAD)
r += 2
header(r, ["When", "What happened", "Months", "of last 12", "Sold", "Record", "Standing", "Retention", "SCORE", "", "Tier"],
       "ABCDEFGHIJK")
ws["J%d" % r] = ""
r += 1
story = [
    ("Month 1", "Opens an account, first payment clears.", 1, 1, 0.0),
    ("Month 6", "Six payments in. Confirmed — and Silver, permanently.", 6, 6, 0.0),
    ("Month 12", "One clean year. Gold: credit, card and Rewards all switch on.", 12, 12, 0.0),
    ("Month 24", "Still paying. Score climbing, tier unchanged.", 24, 12, 0.0),
    ("Month 30", "Misses a payment — hospital bill. Costs him nothing today.", 29, 11, 0.0),
    ("Month 36", "Three calendar years, but 35 payments. Just short of Platinum.", 35, 12, 0.0),
    ("Month 37", "The 36th payment lands. Platinum.", 36, 12, 0.0),
    ("Month 50", "Sells a quarter of his gold for a family wedding.", 49, 12, 0.25),
    ("Month 51", "Inside the allowance. Nothing happened to his score.", 50, 12, 0.25),
    ("Month 60", "Five calendar years — but that one miss is still owed.", 59, 12, 0.25),
    ("Month 61", "The 60th payment. Sovereign.", 60, 12, 0.25),
]
row_story0 = r
for when, what, m, rc, sold in story:
    put("A%d" % r, when, BLK_B, border=True)
    put("B%d" % r, what, BLK, border=True)
    put("C%d" % r, m, BLUE, fmt="0", border=True, align="center")
    put("D%d" % r, rc, BLUE, fmt="0", border=True, align="center")
    put("E%d" % r, sold, BLUE, fmt="0%", border=True, align="center")
    put("F%d" % r, "=IF(C{0}<=$B${1},C{0}*50/$B${1},IF(C{0}>=$B${2},100,50+(C{0}-$B${1})*50/($B${2}-$B${1})))"
        .format(r, "P_M12", "P_M60"), BLK, fmt="0", border=True, align="center")
    put("G%d" % r, "=MIN(D{0},$B${1})*100/$B${1}".format(r, "P_WIN"), BLK, fmt="0", border=True, align="center")
    put("H%d" % r, "=IF(E{0}<=$B${1},1,MAX(0,1-(E{0}-$B${1})/(1-$B${1})))".format(r, "P_ALW"),
        BLK, fmt="0.00", border=True, align="center")
    put("I%d" % r, "=IF(C{0}>=$B${1},MAX($B${2},MIN(F{0},G{0})*H{0}),MIN(F{0},G{0})*H{0})".format(r, "P_CNF", "P_FLR"),
        Font(bold=True, size=11, color=NAVY), fmt="0", border=True, align="center")
    put("J%d" % r, '=REPT("|",ROUND(I{0}/4,0))'.format(r), Font(color=GOLD), border=True)
    put("K%d" % r, "=VLOOKUP(I{0},{1},2,TRUE)".format(r, TIER_RANGE),
        Font(bold=True, color=GOLD), border=True, align="center")
    r += 1
r += 1
put("A%d" % r, "The two moments worth pausing on:", Font(bold=True, color=NAVY))
r += 1
put("A%d" % r, "•   Month 30 — he misses one payment and his tier does not move. The score has slack in the middle of a climb, by design.", BLK)
r += 1
put("A%d" % r, "•   Month 50 — he sells a quarter of his gold and nothing happens at all. That is what the one-third allowance is for.", BLK)
r += 1
put("A%d" % r, "•   The whole cost of that one missed payment is that Sovereign arrives in month 61 instead of month 60. One month, and nothing else.", BLK)
r += 2

# ============================================================ calculator
section("TRY IT   —   change the blue cells")
put("A%d" % r, "Enter a customer's position and the model gives you their tier.", LEAD)
r += 2
header(r, ["Input", "Value", "What it means"], "ABC")
r += 1
IN_M, IN_RC = r, r + 1
IN_G0, IN_GB, IN_GN = r + 2, r + 3, r + 4
live = [
    ("Months paid — total, lifetime", 42, "0", "Every month a payment cleared. Never falls."),
    ("Months paid — of the last 12", 6, "0", "Between 0 and 12."),
    ("Gold held 12 months ago (g)", 30, "0.0", "From the token ledger."),
    ("Gold bought since (g)", 8, "0.0", "Contributions plus any spot purchases, plus Gold Rewards credited."),
    ("Gold held now (g)", 38, "0.0", "If this equals the two above added together, nothing was sold."),
]
for label, val, fmt, note in live:
    put("A%d" % r, label, BLK, border=True)
    put("B%d" % r, val, BLUE_B, fmt=fmt, border=True, align="center")
    put("C%d" % r, note, SUB, border=True)
    r += 1
r += 1
header(r, ["Working", "Result", "Where it comes from"], "ABC")
r += 1
C_SOLD, C_REC, C_STD, C_RET, C_MIN, C_ICS, C_TIER, C_BIND, C_CONF = (r, r + 1, r + 2, r + 3, r + 4, r + 5, r + 6, r + 7, r + 8)
steps = [
    ("Share of your gold sold", "=IF((B{0}+B{1})<=0,0,MAX(0,1-B{2}/(B{0}+B{1})))".format(IN_G0, IN_GB, IN_GN),
     "0%", "1 − (held now ÷ [held a year ago + bought since])."),
    ("RECORD", "=IF(B{0}<=$B${1},B{0}*50/$B${1},IF(B{0}>=$B${2},100,50+(B{0}-$B${1})*50/($B${2}-$B${1})))".format(IN_M, "P_M12", "P_M60"),
     "0.0", "Question 1. Months paid, on the strip above."),
    ("STANDING", "=MIN(B{0},$B${1})*100/$B${1}".format(IN_RC, "P_WIN"), "0.0", "Question 2. Months paid of the last 12."),
    ("RETENTION", "=IF(B{0}<=$B${1},1,MAX(0,1-(B{0}-$B${1})/(1-$B${1})))".format(C_SOLD, "P_ALW"),
     "0.00", "Question 3. 1.00 means nothing was lost."),
    ("The lower of Record and Standing", "=MIN(B{0},B{1})".format(C_REC, C_STD), "0.0", "Both have to be true, so we take the weaker."),
    ("SCORE", "=IF(B{0}>=$B${1},MAX($B${2},B{3}*B{4}),B{3}*B{4})".format(IN_M, "P_CNF", "P_FLR", C_MIN, C_RET),
     "0.0", "Multiplied by Retention, then the floor of 25 if six payments have been made."),
]
for label, f, fmt, note in steps:
    put("A%d" % r, label, BLK_B if label.isupper() or "SCORE" in label else BLK, border=True)
    put("B%d" % r, f, BLK_B, fmt=fmt, border=True, align="center")
    put("C%d" % r, note, SUB, border=True)
    r += 1
put("A%d" % r, "TIER", Font(bold=True, size=13), border=True)
put("B%d" % r, "=VLOOKUP(B{0},{1},2,TRUE)".format(C_ICS, TIER_RANGE), Font(bold=True, size=13, color=GOLD),
    border=True, align="center")
put("C%d" % r, "The highest tier this score has reached.", SUB, border=True)
band(r, F_GRN, "AB")
r += 1
put("A%d" % r, "What is holding them back", BLK, border=True)
put("B%d" % r, '=IF(B{0}<1,"Selling",IF(B{1}<B{2},"Time — keep paying","Discipline — recent misses"))'
    .format(C_RET, C_REC, C_STD), BLK_B, border=True, align="center")
put("C%d" % r, "The app can tell the customer exactly what to fix. This is the retention hook.", SUB, border=True)
r += 1
put("A%d" % r, "Six payments made? (drives the floor)", BLK, border=True)
put("B%d" % r, '=IF(B{0}>=$B${1},"Yes — Silver secured","Not yet")'.format(IN_M, "P_CNF"), BLK, border=True, align="center")
put("C%d" % r, "Confirmed status is simply the 6th payment. It is permanent.", SUB, border=True)
r += 2

# ============================================================ cases
section("EVERY CASE THAT MATTERS")
put("A%d" % r, "The rows we would expect a sharp reader to test. All of them are live.", LEAD)
r += 2
header(r, ["Case", "Months", "of last 12", "Sold", "Record", "Standing", "Retention", "SCORE", "Tier", "What it shows"],
       "ABCDEFGHIJ")
r += 1
cases = [
    ("FAIRNESS", None, None, None, ""),
    ("   USD 20 a month, perfect, five years", 60, 12, 0.0,
     "The smallest saver in the product reaches the top. This is the test the design exists to pass."),
    ("   USD 2,000 a month, perfect, five years", 60, 12, 0.0,
     "Identical, to the day. A hundred times the money buys zero tiers."),
    ("   Six payments scattered, never six in a row", 0, 0, 0.0,
     "Never passes the gate, so no score is ever calculated. Six real payments, no tier."),
    ("SELLING", None, None, None, ""),
    ("   Sells a quarter of his gold", 36, 12, 0.25, "Inside the allowance. Nothing happens."),
    ("   Sells a third", 36, 12, 0.30, "Exactly on the line. Still nothing."),
    ("   Sells half", 36, 12, 0.50, "One tier. He keeps his gold, his loan and his card."),
    ("   Sells everything", 36, 12, 1.00, "Three tiers, to the floor. Recovers in twelve clean months."),
    ("   The cycler — buys and sells every month", 60, 12, 1.00,
     "A flawless payment record held at Silver. He never recovers, because he never stops."),
    ("MISSING PAYMENTS  (a ten-year saver who stops)", None, None, None, ""),
    ("   one missed month", 120, 11, 0.0, "One tier."),
    ("   three missed months", 120, 9, 0.0, "Still Platinum. It takes four in a row to lose it."),
    ("   four missed months", 120, 8, 0.0, "Platinum lost."),
    ("   seven missed months", 120, 5, 0.0, "Gold lost."),
    ("   stopped for a full year", 120, 0, 0.0,
     "Ten years of history does not hold the top tier. Status is rented, not owned."),
    ("   pays every other month, forever", 60, 6, 0.0,
     "Capped at Gold for life. Nothing in the rules stops him — the arithmetic does."),
]
row_case0 = r
for label, m, rc, sold, note in cases:
    if m is None:
        put("A%d" % r, label, Font(bold=True, italic=True, color=NAVY))
        band(r, F_BAND, "ABCDEFGHIJ")
    else:
        put("A%d" % r, label, BLK, border=True)
        put("B%d" % r, m, BLUE, fmt="0", border=True, align="center")
        put("C%d" % r, rc, BLUE, fmt="0", border=True, align="center")
        put("D%d" % r, sold, BLUE, fmt="0%", border=True, align="center")
        put("E%d" % r, "=IF(B{0}<=$B${1},B{0}*50/$B${1},IF(B{0}>=$B${2},100,50+(B{0}-$B${1})*50/($B${2}-$B${1})))"
            .format(r, "P_M12", "P_M60"), BLK, fmt="0", border=True, align="center")
        put("F%d" % r, "=MIN(C{0},$B${1})*100/$B${1}".format(r, "P_WIN"), BLK, fmt="0", border=True, align="center")
        put("G%d" % r, "=IF(D{0}<=$B${1},1,MAX(0,1-(D{0}-$B${1})/(1-$B${1})))".format(r, "P_ALW"),
            BLK, fmt="0.00", border=True, align="center")
        put("H%d" % r, "=IF(B{0}>=$B${1},MAX($B${2},MIN(E{0},F{0})*G{0}),MIN(E{0},F{0})*G{0})".format(r, "P_CNF", "P_FLR"),
            Font(bold=True, color=NAVY), fmt="0", border=True, align="center")
        put("I%d" % r, "=VLOOKUP(H{0},{1},2,TRUE)".format(r, TIER_RANGE), Font(bold=True, color=GOLD),
            border=True, align="center")
        put("J%d" % r, note, SUB, border=True)
    r += 1
row_case1 = r - 1
r += 1

# ============================================================ settings
section("THE NUMBERS WE CHOSE   —   everything above is driven by these six cells")
header(r, ["Setting", "Value", "Why this number"], "ABC")
r += 1
P0 = r
settings = [
    ("Months to reach a Record of 50", 12, "0", "One year. The first year is the hardest, so it is worth the most."),
    ("Months to reach a Record of 100", 60, "0", "Five years. Long enough to be meaningful, short enough to be reachable."),
    ("Trailing window (months)", 12, "0", "A year. Long enough to see a habit, short enough to forgive an accident."),
    ("Free withdrawal allowance", 0.30, "0%", "A third. Covers a genuine household need without touching the score."),
    ("Payments that secure Silver", 6, "0", "Six. Proves the account is real. Permanent once earned."),
    ("Silver floor score", 25, "0", "Nobody who has made six payments is ever sent back to the beginning."),
]
names = ["P_M12", "P_M60", "P_WIN", "P_ALW", "P_CNF", "P_FLR"]
param_row = {}
for (label, val, fmt, why), nm in zip(settings, names, ):
    put("A%d" % r, label, BLK, border=True)
    put("B%d" % r, val, BLUE_B, fmt=fmt, border=True, align="center")
    put("C%d" % r, why, SUB, border=True)
    param_row[nm] = r
    r += 1
ws["B%d" % param_row["P_ALW"]].fill = F_YEL
r += 1
put("A%d" % r, "The one number worth revisiting once we have real data is the withdrawal allowance. "
               "Everything else is structural.", SUB)

# ------------------------------------------------ resolve placeholder refs
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and "P_" in cell.value:
            v = cell.value
            for nm, rw in param_row.items():
                v = v.replace(nm, str(rw))
            cell.value = v

widths = {"A": 48, "B": 13, "C": 13, "D": 12, "E": 12, "F": 12, "G": 12,
          "H": 12, "I": 14, "J": 62, "K": 14, "L": 12, "M": 12, "N": 12}
for k, v in widths.items():
    ws.column_dimensions[k].width = v
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# Resolve beside this script, so the build runs on any machine.
# Was hardcoded to one workstation's path and failed everywhere else.
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Aurumix_ICS_Score_Calculator.xlsx")
wb.save(out)
print("saved:", out)
print("tier range:", TIER_RANGE, "| story rows:", row_story0, "| case rows:", row_case0, "-", row_case1)
print("calc cells: sold=B%d rec=B%d std=B%d ret=B%d min=B%d ics=B%d tier=B%d" % (C_SOLD, C_REC, C_STD, C_RET, C_MIN, C_ICS, C_TIER))
