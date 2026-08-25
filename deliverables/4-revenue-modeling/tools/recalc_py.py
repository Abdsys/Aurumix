"""
Recalculate an xlsx IN PURE PYTHON, writing real cached values while KEEPING
every formula intact.

WHY THIS EXISTS ALONGSIDE recalc_lo.py. recalc_lo.py drives LibreOffice, which
is the preferred route. But LibreOffice is NOT INSTALLED on every build machine
- it is absent on this one - and without it the workbook openpyxl writes carries
NO cached values at all. Every formula then reads as None, verify_model.py sees
an empty sheet, and the whole suite passes vacuously. That silent pass is the
exact failure recalc_lo.py was written to stop, so the fallback has to be a real
evaluation, not a no-op.

WHY IT WRITES XML DIRECTLY. openpyxl cannot hold a formula AND its cached value
in the same cell - assigning the value destroys the formula. A first cut of this
script did exactly that, and verify_model.py caught it immediately: the five
reference-chain checks (Model reads Assumptions, ceilings are DERIVED not typed,
and so on) all failed, because those checks read the formula STRINGS and the
formulas were gone. Values without formulas is a different workbook, not a
calculated one. So this injects <v> alongside the existing <f> in the sheet XML,
which is what LibreOffice's --convert-to does and what Excel expects.

Usage:  python recalc_py.py <file.xlsx>
Writes: _recalc/<file.xlsx>  and prints a JSON summary.
"""
import json
import shutil
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import formulas

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#N/A", "#SPILL!")
ET.register_namespace("", NS)


def _scalar(v):
    """formulas returns 1x1 arrays and numpy scalars; reduce to a python value."""
    try:
        v = v.value
    except AttributeError:
        pass
    try:
        v = v[0, 0]
    except Exception:
        pass
    try:
        v = v.item()
    except AttributeError:
        pass
    return v


def sheet_paths(zf):
    """sheet NAME -> xl/worksheets/sheetN.xml, resolved through the rels."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    tgt = {r.get("Id"): r.get("Target") for r in rels.findall("{%s}Relationship" % PKG)}
    out = {}
    for s in wb.find("{%s}sheets" % NS):
        t = tgt.get(s.get("{%s}id" % RNS), "")
        out[s.get("name")] = "xl/" + t.lstrip("/").replace("xl/", "", 1)
    return out


def recalc(path):
    src = Path(path).resolve()
    out_dir = src.parent / "_recalc"
    out_dir.mkdir(exist_ok=True)
    dst = out_dir / src.name

    xl = formulas.ExcelModel().loads(str(src)).finish()
    solution = xl.calculate()

    vals = {}
    for k, v in solution.items():
        if "'!" not in k:
            continue
        ref = k.split("'!")[-1].strip("'")
        sheet = k.split("]")[-1].split("'!")[0].strip("'").upper()
        if not ref or not ref[0].isalpha():
            continue
        vals[(sheet, ref)] = _scalar(v)

    shutil.copy(src, dst)
    with zipfile.ZipFile(dst) as zf:
        names = zf.namelist()
        paths = sheet_paths(zf)
        blobs = {n: zf.read(n) for n in names}

    written, errors, missing = 0, [], 0
    for sname, spath in paths.items():
        if spath not in blobs:
            continue
        root = ET.fromstring(blobs[spath])
        key = sname.upper()
        changed = False
        for c in root.iter("{%s}c" % NS):
            f = c.find("{%s}f" % NS)
            if f is None:
                continue
            got = vals.get((key, c.get("r")))
            if got is None:
                missing += 1
                continue
            for old in c.findall("{%s}v" % NS):
                c.remove(old)
            v = ET.SubElement(c, "{%s}v" % NS)
            if isinstance(got, bool):
                c.set("t", "b"); v.text = "1" if got else "0"
            elif isinstance(got, str):
                if got in ERRS:
                    c.set("t", "e"); errors.append("%s!%s -> %s" % (sname, c.get("r"), got))
                else:
                    c.set("t", "str")
                v.text = got
            elif isinstance(got, (int, float)):
                if c.get("t") in ("str", "b", "e"):
                    del c.attrib["t"]
                v.text = repr(float(got))
            else:
                c.remove(v); missing += 1; continue
            written += 1
            changed = True
        if changed:
            blobs[spath] = ET.tostring(root, xml_declaration=True, encoding="UTF-8")

    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:
            zf.writestr(n, blobs[n])

    # ASSERT THE CACHE IS REAL AND THE FORMULAS SURVIVED. Either half missing is
    # a silent failure that would otherwise look like a pass.
    from openpyxl import load_workbook
    vw = load_workbook(dst, data_only=True)
    fw = load_workbook(dst)
    cached = sum(1 for ws in vw.worksheets for r in ws.iter_rows() for c in r if c.value is not None)
    formula_cells = sum(1 for ws in fw.worksheets for r in ws.iter_rows() for c in r
                        if isinstance(c.value, str) and c.value.startswith("="))
    if written == 0 or cached == 0:
        return {"status": "error", "detail": "no values written - evaluation produced nothing"}
    if formula_cells == 0:
        return {"status": "error", "detail": "formulas were destroyed - values-only workbook"}

    return {"status": "success" if not errors else "errors",
            "output": str(dst), "values_written": written, "formulas_preserved": formula_cells,
            "cells_cached": cached, "formulas_unresolved": missing,
            "total_errors": len(errors), "errors": errors[:25]}


if __name__ == "__main__":
    p = sys.argv[1] if len(sys.argv) > 1 else "Aurumix_Revenue_Model.xlsx"
    print(json.dumps(recalc(p), indent=2, default=str))
