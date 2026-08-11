"""Evaluates every formula stored in Aurumix_ICS_Score_Calculator.xlsx and checks
the results against _draft_ics-scoring.md.  Layout-agnostic: it finds the scored
rows by looking for the tier VLOOKUP.  Stands in for recalc.py where LibreOffice
is unavailable."""
import os
import re
from openpyxl import load_workbook

# Resolve beside this script, so the check runs on any machine.
PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    "Aurumix_ICS_Score_Calculator.xlsx")
ws = load_workbook(PATH)["ICS"]

VL = re.compile(r"VLOOKUP\(([A-N]\d+),\$A\$(\d+):\$B\$(\d+),2,TRUE\)")
REPT = re.compile(r'REPT\("\|",ROUND\(([A-N]\d+)/4,0\)\)')


def ev(ref, seen=None):
    seen = seen or set()
    if ref in seen:
        raise RuntimeError("circular reference at " + ref)
    v = ws[ref].value
    if not isinstance(v, str) or not v.startswith("="):
        return v
    f, seen = v[1:], seen | {ref}

    m = VL.fullmatch(f)
    if m:
        val = ev(m.group(1), seen)
        best = None
        for rw in range(int(m.group(2)), int(m.group(3)) + 1):
            if val >= ws["A%d" % rw].value:
                best = ws["B%d" % rw].value
        return best
    m = REPT.fullmatch(f)
    if m:
        return "|" * int(round(ev(m.group(1), seen) / 4))
    if f.startswith("IF(") and '"' in f:                    # the two text-verdict cells
        f = re.sub(r"\$?([A-N])\$?(\d+)", lambda mm: "(%r)" % (ev(mm.group(1) + mm.group(2), seen),), f)
        f = re.sub(r"(?<![<>=!])=(?!=)", "==", f)
        return eval(re.sub(r"\bIF\(", "_IF(", f), {"_IF": lambda c, a, b: a if c else b})

    f = f.replace("^", "**")
    f = re.sub(r"(?<![<>=!])=(?!=)", "==", f)
    f = re.sub(r"\bIF\(", "_IF(", f)
    f = re.sub(r"\bMIN\(", "min(", f)
    f = re.sub(r"\bMAX\(", "max(", f)
    f = re.sub(r"\$?[A-N]\$?\d+", lambda mm: "(%r)" % (ev(mm.group(0).replace("$", ""), seen),), f)
    return eval(f, {"_IF": lambda c, a, b: a if c else b, "min": min, "max": max})


# ---- 1. every formula resolves ------------------------------------------
bad, n = [], 0
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            n += 1
            try:
                if ev(cell.coordinate) is None:
                    bad.append((cell.coordinate, "None"))
            except Exception as e:
                bad.append((cell.coordinate, "%s: %s" % (type(e).__name__, e)))
print("formulas evaluated : %d" % n)
print("errors             : %d %s" % (len(bad), bad[:10] if bad else ""))

# ---- 2. every scored row, wherever it sits ------------------------------
scored = []
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and VL.fullmatch(cell.value[1:] if cell.value.startswith("=") else ""):
            src = VL.fullmatch(cell.value[1:]).group(1)
            label = ws["A%d" % cell.row].value or ""
            scored.append((cell.row, str(label).strip(), ev(src), ev(cell.coordinate)))

print("\nSCORED ROWS  (%d)" % len(scored))
for _, label, ics, tier in scored:
    print("  %-52s  %5.1f  ->  %s" % (label[:52], ics, tier))

# ---- 3. spot-checks against the published tables ------------------------
checks = [
    ("USD 20 a month, perfect, five years", 100, "Sovereign"),
    ("USD 2,000 a month, perfect, five years", 100, "Sovereign"),
    ("Six payments scattered, never six in a row", 0, "No tier"),
    ("Sells a quarter of his gold", 75, "Platinum"),
    ("Sells a third", 75, "Platinum"),
    ("Sells half", 53.6, "Gold"),
    ("Sells everything", 25, "Silver"),
    ("The cycler", 25, "Silver"),
    ("one missed month", 91.7, "Platinum"),
    ("three missed months", 75, "Platinum"),
    ("four missed months", 66.7, "Gold"),
    ("seven missed months", 41.7, "Silver"),
    ("stopped for a full year", 25, "Silver"),
    ("pays every other month, forever", 50, "Gold"),
    ("Month 61", 100, "Sovereign"),
    ("Month 60", 98.96, "Platinum"),
    ("Month 50", 88.5, "Platinum"),   # sells 25% — must not move the tier
    ("Month 51", 89.6, "Platinum"),   # still 25% sold — still no effect
    ("Month 30", 67.7, "Gold"),
    ("Month 6", 25, "Silver"),
]
print("\nSPOT-CHECKS vs _draft_ics-scoring.md")
fails = 0
by_label = {lab: (ics, tier) for _, lab, ics, tier in scored}
for want_lab, want_ics, want_tier in checks:
    hit = next(((k, v) for k, v in by_label.items() if k.startswith(want_lab)), None)
    if not hit:
        print("  MISSING ROW: %s" % want_lab)
        fails += 1
        continue
    lab, (ics, tier) = hit
    ok = tier == want_tier and abs(ics - want_ics) < 0.15
    fails += 0 if ok else 1
    print("  %-46s %6.1f %-10s %s" % (want_lab[:46], ics, tier,
                                      "OK" if ok else "<<< want %.1f / %s" % (want_ics, want_tier)))
print("\nfailures: %d" % fails)

# ---- 4. structural invariants -------------------------------------------
def record(m):
    return m * 50 / 12.0 if m <= 12 else (100 if m >= 60 else 50 + (m - 12) * 50 / 48.0)


print("\nINVARIANTS")
print("  Standing >= Record for every clean month 0-240 : %s"
      % all(min(12, m) * 100 / 12.0 >= record(m) - 1e-9 for m in range(241)))
print("  Record(6)=25, so the gate opens exactly at Silver: %s" % (abs(record(6) - 25) < 1e-9))
print("  Record(12)=50, Record(36)=75, Record(60)=100   : %s"
      % all(abs(record(a) - b) < 1e-9 for a, b in [(12, 50), (36, 75), (60, 100)]))
print("  Retention == 1.000 at exactly the 30%% allowance: %s"
      % (ev("B%d" % next(rw for rw in range(1, 200)
                         if str(ws["A%d" % rw].value).startswith("Free withdrawal"))) == 0.30))
