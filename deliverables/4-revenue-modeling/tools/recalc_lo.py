"""
Recalculate an xlsx with LibreOffice and report formula errors.  WINDOWS-SAFE.

Why this exists instead of the xlsx skill's recalc.py:
  1. recalc.py's setup_libreoffice_macro() uses the LINUX profile path
     (~/.config/libreoffice/...) on every platform, so on Windows the macro is
     never installed, the soffice call silently does nothing, and the file is
     left untouched.
  2. It then loads the UNTOUCHED file with data_only=True. openpyxl-written
     files carry NO cached values, so every formula reads None, no cell can
     contain an error string, and it reports "status: success, total_errors: 0".
     That success is an artefact of the failure - it is not a passing test.

This version uses `--convert-to xlsx` with an ISOLATED user profile
(-env:UserInstallation), which forces LibreOffice to evaluate every formula and
write real cached values. It then asserts the cache actually exists, so a silent
no-op can never masquerade as a pass again.

Usage:  python recalc_lo.py <file.xlsx>
Writes: _recalc/<file.xlsx>  and prints a JSON summary.
"""
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path
from openpyxl import load_workbook

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
PROFILE = "file:///C:/Temp/lo_recalc_profile"
ERRORS = ("#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A")


def recalc(path):
    src = Path(path).resolve()
    if not src.exists():
        return {"status": "error", "detail": "%s does not exist" % src}
    if not Path(SOFFICE).exists():
        return {"status": "error", "detail": "LibreOffice not found at %s" % SOFFICE}

    outdir = src.parent / "_recalc"
    if outdir.exists():
        shutil.rmtree(outdir, ignore_errors=True)
    outdir.mkdir(parents=True, exist_ok=True)

    proc = subprocess.run(
        [SOFFICE, "-env:UserInstallation=%s" % PROFILE, "--headless", "--norestore",
         "--convert-to", "xlsx", "--outdir", str(outdir), str(src)],
        capture_output=True, text=True, timeout=600)

    out = outdir / src.name
    if not out.exists():
        return {"status": "error", "detail": "convert-to produced nothing",
                "stdout": proc.stdout[-400:], "stderr": proc.stderr[-400:]}

    wb_v = load_workbook(out, data_only=True)
    wb_f = load_workbook(out, data_only=False)

    formulas = cached = 0
    found = {}
    for name in wb_f.sheetnames:
        wsf, wsv = wb_f[name], wb_v[name]
        for row in wsf.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    formulas += 1
                    got = wsv[c.coordinate].value
                    if got is not None:
                        cached += 1
                    if isinstance(got, str) and got in ERRORS:
                        found.setdefault(got, []).append("%s!%s" % (name, c.coordinate))
    wb_v.close()
    wb_f.close()

    total = sum(len(v) for v in found.values())
    # A cache that never materialised means the recalc did not happen. Refuse to
    # call that a pass - this is precisely the recalc.py failure mode.
    if formulas and cached == 0:
        return {"status": "error", "recalculated_path": str(out),
                "detail": "NO cached values written - recalculation did not run. "
                          "Zero errors here would be meaningless.",
                "total_formulas": formulas}

    return {
        "status": "success" if total == 0 else "errors_found",
        "recalculated_path": str(out),
        "total_formulas": formulas,
        "formulas_with_cached_value": cached,
        "total_errors": total,
        "error_summary": {k: {"count": len(v), "locations": v[:20]} for k, v in found.items()},
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    res = recalc(sys.argv[1])
    print(json.dumps(res, indent=2))
    sys.exit(0 if res.get("status") == "success" else 1)
