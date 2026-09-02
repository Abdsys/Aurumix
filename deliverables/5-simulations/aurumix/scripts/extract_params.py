"""
Extract every parameter the Model sheet references from the CALCULATED workbook
into config/params.json. The deterministic engine reads that JSON, so the port
cannot drift from the workbook by transcription error.

Reference of record: Aurumix_Revenue_Model_calculated.xlsx (blueprint, At a Glance).
"""

import json
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.normpath(os.path.join(
    HERE, "..", "..", "4-revenue-modeling", "tools",
    "Aurumix_Revenue_Model_calculated.xlsx"))

# Assumptions!B cells the Model formulas reference, keyed by a readable name.
B = {
    "gold_price_m1": 6, "aed_usd": 7, "entry_fee": 8, "fab_premium": 9,
    "bar_grams": 10,
    "interchange": 14, "processor_fee_txn": 15, "decline_uplift": 16,
    "fx_margin": 17, "atm_allowance_aed": 18, "atm_fee": 19,
    "card_issuance_aed": 20, "card_replacement_aed": 21,
    "family_price": 25, "beneficiary_fee": 26,
    "credit_orig_gross": 27, "credit_orig_share": 28,
    "credit_serv_gross": 29, "credit_serv_share": 30, "ltv": 31,
    "spot_ticket_uae": 32, "spot_ticket_gulf": 33, "spot_ticket_india": 34,
    "spot_attach_uae": 35, "spot_attach_gulf": 36, "spot_attach_india": 37,
    "sip_floor": 41, "gate_payments": 42, "redemption_unit_cost": 43,
    "act_1a": 47, "act_1b": 48, "act_0": 49, "act_3": 50, "act_2": 51,
    "act_4": 52, "act_5": 53, "act_6": 54, "act_referral": 55,
    "act_gulf": 56, "act_india": 57,
    "ceiling_uae": 84, "ceiling_gulf": 85, "ceiling_india": 86,
    "gold_appreciation": 131, "float_buffer_days": 132,
    "persistency": 133, "monthly_churn": 134, "agent_productivity": 135,
    "cac_uae": 136, "cac_gulf": 137, "cac_india": 138,
    "cac_uae_y7": 139, "cac_gulf_y7": 140, "cac_india_y7": 141,
    "mkt_share_uae": 142, "mkt_share_gulf": 143, "mkt_share_india": 144,
    "referral_rate": 145, "referral_conversion": 146, "organic_share": 147,
    "seasonality_amplitude": 148,
    "ics_ever_share": 149, "ics_months_to_tier": 150,
    "self_custody_rate": 151, "redemption_rate": 152, "holder_redemption_mult": 153,
    "spot_attach_mult": 154, "spot_ticket_mult": 155, "spot_frequency": 156,
    "pm_share": 157, "facility_takeup": 158, "txn_per_draw": 159,
    "foreign_spend_mean": 160, "reissue_rate": 161, "replacement_rate": 162,
    "atm_share_1": 163, "atm_share_2": 164, "atm_share_3": 165, "atm_share_4": 166,
    "atm_mid_1": 167, "atm_mid_2": 168, "atm_mid_3": 169, "atm_mid_4": 170,
    "drawn_pct": 171, "facility_turnover": 172, "draws_per_year": 173,
    "family_attach": 174, "family_cancel": 175, "beneficiaries": 176,
    "family_churn_monthly": 177,
    "b2b_fee": 185, "partner_users": 186, "partner_adopt": 187,
    "partner_aum_user": 188, "partner_aum": 189,
    "ticket_uae": 190, "ticket_gulf": 191, "ticket_india": 192,
    "vault_fee": 193, "vault_min_daily": 194,
    "vara_supervision_aed": 195, "vara_application_aed": 196,
    "dmcc_licence_aed": 197, "dmcc_incorporation_aed": 198,
    "kyc_per_check": 199,
    "capital_issuance_aed": 200, "capital_activities_aed": 201,
    "nla_months": 202,
    "insurance": 203, "audit": 204, "tech_audit": 205, "launch_audit": 206,
    "ics_disc_entry": 207, "ics_disc_card": 208, "ics_disc_rebate": 209,
    "ics_disc_family": 210,
    "agent_commission": 211, "referral_reward": 212,
    "card_platform_fee": 213, "card_setup": 214, "card_per_card": 215,
    "card_per_auth": 216, "card_fraud": 217, "xborder_fee": 218,
    "uae_spend_abroad": 219, "prefund_days": 220, "prefund_min": 221,
    "tech_build_y1": 222, "tech_build_y2": 223, "tech_maint": 224,
    "contingency": 225, "kyc_monthly_min": 226,
    "sw_prepaid": 227, "sw_holders_keep_card": 228,
    "sw_premium_aurumix": 229, "sw_premium_gross": 230,
}

# Year-indexed vectors (rows inclusive)
VECTORS = {
    "marketing_spend": (114, 120),
    "agents_uae": (93, 99),
    "agents_gulf": (100, 106),
    "agents_india": (107, 113),
    "agent_ramp": (121, 127),
    "b2b_partners": (178, 184),
}


def main():
    wb = openpyxl.load_workbook(WB, data_only=True)
    a = wb["Assumptions"]

    params = {}
    for name, row in B.items():
        v = a.cell(row, 2).value
        if v is None:
            print(f"WARNING: Assumptions!B{row} ({name}) is empty", file=sys.stderr)
        params[name] = v

    for name, (r0, r1) in VECTORS.items():
        params[name] = [a.cell(r, 2).value for r in range(r0, r1 + 1)]

    # Seasonality vectors: Assumptions F..Q (cols 6..17) rows 61-63
    for name, row in [("season_acq", 61), ("season_card", 62), ("season_foreign", 63)]:
        params[name] = [a.cell(row, c).value for c in range(6, 18)]

    # The 29-period grid header rows off the Model sheet (period #, year, month,
    # months-in-period) so the engine's calendar matches exactly.
    m = wb["Model"]
    grid = {}
    for key, row in [("period", 2), ("year", 5), ("cal_month", 6), ("months", 7)]:
        grid[key] = [m.cell(row, c).value for c in range(3, 32)]   # C..AE
    params["grid"] = grid

    out = os.path.join(HERE, "config", "params.json")
    with open(out, "w") as f:
        json.dump(params, f, indent=1)
    print(f"wrote {out}: {len(params)} entries")
    nulls = [k for k, v in params.items() if v is None]
    if nulls:
        print("NULL params:", nulls)


if __name__ == "__main__":
    main()
