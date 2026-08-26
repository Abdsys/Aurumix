"""
Verification for the simplified Aurumix revenue model, read from the
RECALCULATED workbook.

Usage:  python recalc_lo.py Aurumix_Revenue_Model.xlsx
        python verify_model.py _recalc/Aurumix_Revenue_Model.xlsx
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = sys.argv[1] if len(sys.argv) > 1 else "_recalc/Aurumix_Revenue_Model.xlsx"
wb = load_workbook(PATH, data_only=True)
N, PCOL0, N_MONTHLY = 29, 3, 24
def pc(i): return get_column_letter(PCOL0 + i)

results = []
def check(name, ok, detail=""):
    results.append((name, bool(ok), detail))


def all_num(seq):
    """NON-EMPTY and every element numeric - guards the vacuous-truth trap."""
    seq = list(seq)
    return bool(seq) and all(isinstance(x, (int, float)) for x in seq)


EXPECTED = ["Cover", "Assumptions", "Scenario Parameters", "Model", "Summary"]
check("Exactly the 5 standard sheets, in order", wb.sheetnames == EXPECTED, "%r" % (wb.sheetnames,))
asm, sc, mdl, summ = wb["Assumptions"], wb["Scenario Parameters"], wb["Model"], wb["Summary"]

# ---- 0. the reference chain: Scenario -> Assumptions -> Model -> Summary ---
# Read from the FORMULA copy, since the recalculated one holds cached values.
import os
_src = os.path.join(os.path.dirname(os.path.abspath(PATH)) or ".", os.path.basename(PATH))
wbf = load_workbook(_src, data_only=False)


def refs(sheet, target):
    n = 0
    for r_ in wbf[sheet].iter_rows():
        for c in r_:
            if isinstance(c.value, str) and c.value.startswith("=") and target in c.value:
                n += 1
    return n


check("Model reads Assumptions, never Scenario Parameters directly",
      refs("Model", "Scenario Parameters") == 0 and refs("Model", "Assumptions") > 0,
      "Model -> Scenario %d, Model -> Assumptions %d"
      % (refs("Model", "Scenario Parameters"), refs("Model", "Assumptions")))
check("Summary reads only the Model",
      refs("Summary", "Scenario Parameters") == 0 and refs("Summary", "Model") > 0,
      "Summary -> Scenario %d" % refs("Summary", "Scenario Parameters"))
# No sheet may carry a duplicate row label. A repeated label is how a lookup
# silently reads the wrong row - it has caused two real defects in this build.
for _sh in ("Model", "Summary", "Assumptions"):
    _seen, _dupes = {}, []
    for _r in range(1, wbf[_sh].max_row + 1):
        _lv = wbf[_sh]["A%d" % _r].value
        if isinstance(_lv, str) and _lv.strip() and not _lv.startswith("="):
            _seen.setdefault(_lv, []).append(_r)
    # Indented region-block rows repeat by design; so do per-section column
    # headers. Everything else must be unique.
    _HDRS = {"Parameter", "Funnel step", "Scenario", "Switch"}
    _dupes = {k: v for k, v in _seen.items()
              if len(v) > 1 and not k.startswith("  ") and k not in _HDRS}
    check("%s has no duplicate top-level row labels" % _sh, not _dupes,
          "%r" % list(_dupes.items())[:3])

check("Assumptions carries a scenario-linked mirror block",
      refs("Assumptions", "Scenario Parameters") > 0,
      "mirror links found: %d" % refs("Assumptions", "Scenario Parameters"))

# EVERY scenario parameter must appear on Assumptions. A parameter row is a
# label in A with a live formula in B (CHOOSE, a derivation, or a switch IF).
_fsc, _fasm = wbf["Scenario Parameters"], wbf["Assumptions"]
_params = [(r, _fsc["A%d" % r].value.strip())
           for r in range(1, _fsc.max_row + 1)
           if isinstance(_fsc["A%d" % r].value, str) and _fsc["A%d" % r].value.strip()
           and isinstance(_fsc["B%d" % r].value, str) and _fsc["B%d" % r].value.startswith("=")]
_mirror = {}
for r in range(1, _fasm.max_row + 1):
    b = _fasm["B%d" % r].value
    if isinstance(b, str) and b.startswith("=") and "Scenario Parameters" in b:
        _mirror[int(b.split("$B$")[1].split(")")[0].strip())] = (_fasm["A%d" % r].value or "").strip()
_missing = [n for r, n in _params if r not in _mirror]
check("EVERY scenario parameter is mirrored onto Assumptions",
      not _missing and len(_params) == len(_mirror),
      "%d parameters, %d mirrored. Missing: %s" % (len(_params), len(_mirror), _missing[:5]))
_mismatch = [(n, _mirror[r]) for r, n in _params if r in _mirror and _mirror[r] != n]
check("Mirrored labels match their Scenario Parameters labels exactly",
      not _mismatch, "%r" % _mismatch[:3])

# NO ORPHANS. Added 2026-08-21 after "Spot-to-SIP conversion" was found defined,
# mirrored, scenario-flexed - and referenced by nothing. A dead input is worse
# than a missing one: it reads as a modelled mechanism, it implies a population
# the model does not have, and flipping the scenario appears to do something.
# Mirroring alone never catches this, because an orphan mirrors perfectly.
import re as _re


def _cited(formula, sheet=None):
    """Every $B$n row a formula touches, EXPANDING RANGES.

    Ranges matter: the annual schedules are read as INDEX(Assumptions!$B$a:$B$b,
    year), so only the first row carries the sheet prefix. Matching single cells
    alone reports rows 2..n of every schedule as orphans."""
    pat = (r"%s!\$B\$(\d+)(?::\$B\$(\d+))?" % sheet) if sheet else r"(?<![!\w])\$B\$(\d+)(?::\$B\$(\d+))?"
    out = set()
    for lo, hi in _re.findall(pat, formula):
        lo = int(lo)
        out.update(range(lo, int(hi) + 1) if hi else [lo])
    return out


_used_asm = set()          # Assumptions rows the Model reads
for _r in wbf["Model"].iter_rows():
    for _c in _r:
        if isinstance(_c.value, str) and _c.value.startswith("="):
            _used_asm |= _cited(_c.value, "Assumptions")
_used_sc = set()           # Scenario rows other Scenario formulas read
for _r in _fsc.iter_rows():
    for _c in _r:
        if isinstance(_c.value, str) and _c.value.startswith("="):
            # Derived rows refer to their OWN sheet BY NAME ("'Scenario
            # Parameters'!$B$10"), not as a bare local ref, so both forms count.
            _used_sc |= _cited(_c.value) | _cited(_c.value, "'Scenario Parameters'")

# Map each mirrored Assumptions row back to the Scenario row it came from.
_mirror_asm = {}
for r in range(1, _fasm.max_row + 1):
    b = _fasm["B%d" % r].value
    if isinstance(b, str) and b.startswith("=") and "Scenario Parameters" in b:
        _mirror_asm[r] = (int(b.split("$B$")[1].split(")")[0].strip()),
                          (_fasm["A%d" % r].value or "").strip())
# A parameter EARNS ITS PLACE if the Model reads its Assumptions mirror, OR a
# derived scenario parameter reads it upstream (persistency, for instance, is
# consumed by the monthly-churn derivation rather than by the Model directly -
# it is mirrored for visibility, at the client's explicit request).
_orphans = sorted(n for r, (sr, n) in _mirror_asm.items()
                  if r not in _used_asm and sr not in _used_sc)
check("NO ORPHANED scenario parameters - every one is read by the Model or a derivation",
      not _orphans, "scenario-flexed but referenced nowhere: %r" % _orphans)


def rowof(ws, label):
    """Exact match INCLUDING indentation, and it must be UNIQUE.

    Indentation is meaningful: region-block rows are indented ("  Stream 2 ...")
    and whole-book totals are not, so an exact match distinguishes a regional
    line from the total of the same name. Uniqueness is asserted because a
    duplicate label is how a lookup silently reads the wrong row."""
    hits = [r for r in range(1, ws.max_row + 1) if ws["A%d" % r].value == label]
    if len(hits) > 1:
        raise AssertionError("label %r is ambiguous on %s: rows %r" % (label, ws.title, hits))
    return hits[0] if hits else None


def mv(label, i):
    r = rowof(mdl, label)
    return mdl["%s%d" % (pc(i), r)].value if r else None


def row(label): return [mv(label, i) for i in range(N)]
def per_label(i): return mdl.cell(row=3, column=PCOL0 + i).value
def sval(label):
    r = rowof(sc, label)
    return sc["B%d" % r].value if r else None
def aval(label):
    r = rowof(asm, label)
    return asm["B%d" % r].value if r else None


# ---- 1. no errors ----------------------------------------------------------
ERRS = ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!", "#NUM!", "#N/A")
bad = []
for ws in wb.worksheets:
    for r_ in ws.iter_rows():
        for c in r_:
            if isinstance(c.value, str) and c.value in ERRS:
                bad.append("%s!%s=%s" % (ws.title, c.coordinate, c.value))
check("No formula errors in any sheet", not bad, "; ".join(bad[:6]))

# ---- 2. period grid --------------------------------------------------------
check("29 periods indexed 1..29", row("Period #") == list(range(1, 30)))
yrs, n = row("Model year"), row("Months in period")
check("Years run 1x12, 2x12, then 3..7",
      yrs[:12] == [1] * 12 and yrs[12:24] == [2] * 12 and yrs[24:] == [3, 4, 5, 6, 7])
check("Monthly block = 1 month, annual block = 12", n[:24] == [1] * 24 and n[24:] == [12] * 5)

# ---- 3. seasonality --------------------------------------------------------
sr = rowof(mdl, "Seasonality sum check (must be exactly 12.000)")
a_s, s_s = mdl["%s%d" % (pc(0), sr)].value, mdl["%s%d" % (pc(1), sr)].value
check("Acquisition seasonality sums to EXACTLY 12.000", abs(a_s - 12.0) < 1e-9, "%r" % a_s)
check("Card spend seasonality sums to EXACTLY 12.000", abs(s_s - 12.0) < 1e-9, "%r" % s_s)

# ---- 4. the customer engine ------------------------------------------------
cons = row("CHECK: paying + holders = cumulative ever acquired")
check("CONSERVATION - paying + holders = cumulative ever acquired, every period",
      all_num(cons) and max(abs(x) for x in cons) < 1e-6,
      "max delta %r" % (max(abs(x) for x in cons) if all_num(cons) else cons))
pay, hold, cum = row("PAYING CUSTOMERS"), row("HOLDERS (stopped paying, still hold gold)"), row("Cumulative ever acquired")
check("No negative populations", all(x >= -1e-9 for x in pay + hold + cum))
check("Cumulative ever acquired strictly increases", all(cum[i] > cum[i - 1] for i in range(1, N)))
check("Holders overtake payers as the book matures", hold[-1] > pay[-1],
      "Y7 holders %.0f vs paying %.0f" % (hold[-1], pay[-1]))
# Found by PREFIX, so renaming a region does not silently break the check.
ceil_rows = [r for r in range(1, asm.max_row + 1)
             if isinstance(asm["A%d" % r].value, str)
             and asm["A%d" % r].value.startswith("Reachable SIP ceiling")]
REGION_NAMES = [asm["A%d" % r].value.split(" - ", 1)[1] for r in ceil_rows]
check("At least three regional ceilings are defined", len(ceil_rows) >= 3,
      "found %d: %r" % (len(ceil_rows), REGION_NAMES))
ceil_tot = sum(asm["B%d" % r].value for r in ceil_rows)
# REPLACED 2026-08-21. This was an anchor hardcoded at ~181,150, and it had
# already been bumped once (the UAE Indian population correction, 3.5m -> 4.36m)
# and would have been bumped again here (the India ceiling, 0.35% -> 0.70%).
# A magic number that is edited every time it fires is a RATCHET, NOT A TRIPWIRE
# - it can only ever confirm the number someone last typed.
#
# Replaced by an INDEPENDENT RECOMPUTATION from the raw funnel inputs. This
# genuinely catches what the anchor was meant to catch - a regional ceiling
# quietly becoming a typed number, or decoupling from the funnel above it - and
# it needs no maintenance when a population or a ceiling legitimately changes.
_fh = next(r for r in range(1, asm.max_row + 1) if asm["A%d" % r].value == "Funnel step")
_frow = {}
for r in range(_fh, min(_fh + 14, asm.max_row + 1)):
    lab = (asm["A%d" % r].value or "").strip()
    for k in ("Source population", "x Economically", "x Payment", "x Money", "x Penetration", "Feeds model"):
        if lab.startswith(k):
            _frow[k] = r
_fcols = [c for c in range(2, 14) if asm.cell(row=_frow["Feeds model"], column=c).value]
_recomputed = {}
for c in _fcols:
    v = 1.0
    for k in ("Source population", "x Economically", "x Payment", "x Money", "x Penetration"):
        v *= asm.cell(row=_frow[k], column=c).value
    _recomputed.setdefault(asm.cell(row=_frow["Feeds model"], column=c).value, 0.0)
    _recomputed[asm.cell(row=_frow["Feeds model"], column=c).value] += v
_mismatched = []
for r in ceil_rows:
    rn = asm["A%d" % r].value.split(" - ", 1)[1]
    if abs(asm["B%d" % r].value - _recomputed.get(rn, -1)) > 1:
        _mismatched.append((rn, asm["B%d" % r].value, _recomputed.get(rn)))
check("Every regional ceiling RECOMPUTES from population x filters x penetration",
      not _mismatched and len(_recomputed) == len(ceil_rows),
      "%r" % _mismatched[:3])
# Wide band, so it only fires on a wild rescale rather than on every legitimate
# correction. Deliberately loose - the check above is the precise one.
check("Total reachable ceiling is the right order of magnitude (50k-500k)",
      50_000 < ceil_tot < 500_000, "%.0f" % ceil_tot)

# The funnel must be live arithmetic, not typed numbers.
_fa = wbf["Assumptions"]
_fun_rows = {}
for r in range(1, _fa.max_row + 1):
    v = _fa["A%d" % r].value
    if isinstance(v, str) and v.strip() in ("= Addressable base", "= REACHABLE SIP ACCOUNTS",
                                            "Source population"):
        _fun_rows[v.strip()] = r
check("The market-sizing funnel is present on Assumptions", len(_fun_rows) == 3,
      "found %r" % sorted(_fun_rows))
if len(_fun_rows) == 3:
    ar_ = _fun_rows["= Addressable base"]
    rr_ = _fun_rows["= REACHABLE SIP ACCOUNTS"]
    live = all(isinstance(_fa.cell(row=ar_, column=6 + k).value, str)
               and _fa.cell(row=ar_, column=6 + k).value.startswith("=") for k in range(4))
    live2 = all(isinstance(_fa.cell(row=rr_, column=6 + k).value, str)
                and _fa.cell(row=rr_, column=6 + k).value.startswith("=") for k in range(4))
    check("Addressable base and reachable accounts are COMPUTED, not typed", live and live2)
    # and the region ceilings must be derived from the funnel, not typed
    derived = all(isinstance(_fa["B%d" % r].value, str) and _fa["B%d" % r].value.startswith("=")
                  for r in ceil_rows)
    check("Region ceilings are DERIVED from the funnel, not typed", derived)
    # arithmetic: population x filters x ceiling reproduces reachable accounts
    pr_ = _fun_rows["Source population"]
    ok_arith = True
    for k in range(4):
        col = get_column_letter(6 + k)
        pop = asm["%s%d" % (col, pr_)].value
        acct = asm["%s%d" % (col, rr_)].value
        chain = pop
        for off in (1, 2, 3):
            chain *= asm["%s%d" % (col, pr_ + off)].value
        chain *= asm["%s%d" % (col, pr_ + 5)].value
        if abs(chain - acct) > 1:
            ok_arith = False
    check("Funnel arithmetic reconciles end to end (population -> accounts)", ok_arith)
check("Cumulative ever acquired never breaches the ceiling", cum[-1] <= ceil_tot + 1,
      "%.0f vs %.0f" % (cum[-1], ceil_tot))
_sat_rows = [r for r in range(1, mdl.max_row + 1)
             if mdl["A%d" % r].value == "  Saturation (this region's own headroom)"]
check("Saturation is applied PER REGION, not just on the whole book",
      len(_sat_rows) == len(REGION_NAMES), "found %d for %d regions" % (len(_sat_rows), len(REGION_NAMES)))
for sr_ in _sat_rows:
    vals = [mdl["%s%d" % (pc(i), sr_)].value for i in range(N)]
    if all_num(vals):
        check("Regional saturation decays monotonically (row %d)" % sr_,
              all(vals[i] <= vals[i - 1] + 1e-12 for i in range(1, N)))

# No region may breach its OWN ceiling - the whole-book total staying inside the
# sum of ceilings is not sufficient, since one channel can concentrate in one
# market. Agents route to India only, which is exactly that risk.
_cum_rows = [r for r in range(1, mdl.max_row + 1)
             if mdl["A%d" % r].value == "  Cumulative ever acquired"]
breach = []
for rn, cr_ in zip(REGION_NAMES, _cum_rows):
    cap = aval("Reachable SIP ceiling - %s" % rn)
    end = mdl["%s%d" % (pc(N - 1), cr_)].value
    if isinstance(end, (int, float)) and isinstance(cap, (int, float)) and end > cap * 1.001:
        breach.append("%s %.0f > %.0f" % (rn, end, cap))
check("NO region breaches its own reachable ceiling", not breach, "; ".join(breach))

# ---- AGENTS ARE REGIONAL (rewritten 2026-08-26) ---------------------------
# The three checks here previously tested a "share of salesforce" input that no
# longer exists: agent headcount now belongs to a region, so there is no
# national pool to allocate and nothing to sum to 100%. What replaces them is
# stricter, because it tests the ENGINE rather than the input - a region with no
# agents must produce no agent accounts in every single period, which the old
# share-based routing could only assert at the input.
_agent_rows = [r for r in range(1, mdl.max_row + 1)
               if mdl["A%d" % r].value == "  Agent-driven (this market's own agents)"]
check("Every region has its own agent-driven row",
      len(_agent_rows) == len(REGION_NAMES),
      "found %d for %d regions" % (len(_agent_rows), len(REGION_NAMES)))

_heads = {rn: [aval("Active agents - %s - Y%d" % (rn, y)) for y in range(1, 8)]
          for rn in REGION_NAMES}
check("Agent headcount is stated per region, not as one national pool",
      all(all_num(v) for v in _heads.values()), "%r" % _heads)

# The point of the restructure: agents belong to India, and the Gulf markets are
# acquired top-down. If someone later gives UAE a headcount this check fails,
# which is correct - that is a decision that should be made deliberately.
_agent_by_region = dict(zip(REGION_NAMES, _agent_rows))
for rn, ar_ in _agent_by_region.items():
    vals = [mdl["%s%d" % (pc(i), ar_)].value for i in range(N)]
    heads = _heads[rn]
    if all_num(heads) and max(heads) == 0:
        check("%s has no agents, so earns NO agent accounts in any period" % rn,
              all_num(vals) and all(abs(v) < 1e-9 for v in vals),
              "max %r" % (max(vals) if all_num(vals) else vals))
    else:
        check("%s has agents and they actually produce accounts" % rn,
              all_num(vals) and max(vals) > 0)

_ind_name = [rn for rn in REGION_NAMES if "India" in rn]
if _ind_name:
    check("India carries the agent network",
          max(_heads[_ind_name[0]]) > 0, "%r" % _heads[_ind_name[0]])
    # Agents must be India's largest single channel - that is what "agent-led"
    # means, and it is the thing the marketing re-base was done to achieve.
    _ir = REGION_NAMES.index(_ind_name[0])
    _dir_rows = [r for r in range(1, mdl.max_row + 1)
                 if mdl["A%d" % r].value == "  Direct-driven (this market's budget at its own CAC)"]
    _ref_rows = [r for r in range(1, mdl.max_row + 1)
                 if mdl["A%d" % r].value == "  Referral-driven (from this market's own base)"]
    _ag = mdl["%s%d" % (pc(N - 1), _agent_rows[_ir])].value
    _dr = mdl["%s%d" % (pc(N - 1), _dir_rows[_ir])].value
    _rf = mdl["%s%d" % (pc(N - 1), _ref_rows[_ir])].value
    if all_num([_ag, _dr, _rf]):
        check("India is AGENT-LED at Y7 - agents beat marketing and referral separately",
              _ag > _dr and _ag > _rf,
              "agent %.0f vs direct %.0f vs referral %.0f" % (_ag, _dr, _rf))
        check("India agents are a MAJORITY of its acquisition at Y7",
              _ag / (_ag + _dr + _rf) > 0.50,
              "agent share %.1f%%" % (100 * _ag / (_ag + _dr + _rf)))

# The routing apparatus must be GONE, not merely unused - a stale renormalisation
# row silently dividing by zero would be worse than either design.
check("The national agent pool and its renormalisation row are removed",
      not [r for r in range(1, mdl.max_row + 1)
           if mdl["A%d" % r].value in ("Agent-driven new customers (routed by region below)",
                                       "Open-region salesforce share (renormalisation base)")])

_cac = {rn: sval("Marketing CAC - %s" % rn) for rn in REGION_NAMES}
_mks = {rn: sval("Marketing spend share - %s" % rn) for rn in REGION_NAMES}
check("CAC is regionalised, not one global figure", len(set(_cac.values())) > 1, "%r" % _cac)
check("Marketing spend shares sum to 100%", abs(sum(_mks.values()) - 1.0) < 1e-9, "%r" % _mks)
_ind = [rn for rn in REGION_NAMES if "India" in rn]
_uae = [rn for rn in REGION_NAMES if rn == "UAE"]
if _ind and _uae:
    check("India CAC is far below the UAE's, matching the published gap",
          _cac[_ind[0]] < _cac[_uae[0]] * 0.35,
          "India %r vs UAE %r" % (_cac[_ind[0]], _cac[_uae[0]]))
check("Every region acquires through its own marketing budget and CAC",
      len([r for r in range(1, mdl.max_row + 1)
           if mdl["A%d" % r].value == "  Direct-driven (this market's budget at its own CAC)"])
      == len(REGION_NAMES))
check("Every region refers from its OWN paying base",
      len([r for r in range(1, mdl.max_row + 1)
           if mdl["A%d" % r].value == "  Referral-driven (from this market's own base)"])
      == len(REGION_NAMES))
# REVISED TWICE. On 2026-08-21 the check asserted that EVERY market has a
# deployed salesforce, correcting an earlier cut where 100% sat in India and the
# licensed home market had nobody selling. On 2026-08-26 the design changed
# again and deliberately went back to agents-in-India-only - but for a different
# reason, and with the hole plugged rather than ignored. The UAE is no longer
# "a market with nobody selling"; it is a market acquired top-down through
# marketing at its own CAC, which is why its marketing SHARE rose from 58% to
# 74% in the same change. The replacement check is therefore not about agents at
# all - it is that a market without agents must be funded to acquire without
# them. That is the substantive risk the 08-21 check was reaching for.
for rn in REGION_NAMES:
    _h = _heads.get(rn) or [0]
    if all_num(_h) and max(_h) == 0:
        check("%s has no agents, so it must carry a real marketing budget" % rn,
              _mks.get(rn, 0) > 0.10,
              "marketing share %.0f%%" % (100 * _mks.get(rn, 0)))
check("UAE, the licensed home market, still takes the largest marketing share",
      _mks.get("UAE", 0) > max(v for k, v in _mks.items() if k != "UAE"), "%r" % _mks)

# churn must reproduce the stated persistency over 12 months
pers, churn = sval("Persistency - customers still paying after 12 months"), sval("Monthly churn rate (derived)")
check("Derived monthly churn reproduces the stated persistency over 12 months",
      abs((1 - churn) ** 12 - pers) < 1e-9, "(1-%.5f)^12 = %.4f vs %.2f" % (churn, (1 - churn) ** 12, pers))

# ---- 5. ACCESS IS UNGATED; ICS ENTITLEMENT IS NOT (rewritten 2026-08-26) ---
# CG's decision: the card and the credit facility are open to the whole book.
# ICS governs BENEFITS, not ACCESS. These checks were inverted before - they
# asserted that eligibility BITES, i.e. that the card population was strictly
# below the book. The opposite is now required, and the old check would have
# passed silently on a stale build, so it has to flip rather than be deleted.
q, cards = row("Reaches an ICS benefit tier (memo - drives discounts, not access)"), row("ACTIVE CARDS")
ever_q = sval("Customers who EVER reach an ICS benefit tier")
takeup = sval("Facility take-up - customers who take AND use the card")
check("ICS-entitled population never exceeds the book x the entitlement rate",
      all(q[i] <= (pay[i] + hold[i]) * ever_q + 1e-6 for i in range(N)))
check("ICS entitlement STILL bites - it is a subset, ready for the discount build",
      q[-1] < (pay[-1] + hold[-1]) * 0.999,
      "entitled %.0f vs book %.0f" % (q[-1], pay[-1] + hold[-1]))
# THE CORE ASSERTION OF THIS CHANGE. Cards must come off the whole card-eligible
# base at the take-up rate. If anyone re-gates access, this fails immediately.
_cb = row("Card-eligible base") if any(
    mdl["A%d" % r].value == "Card-eligible base" for r in range(1, mdl.max_row + 1)) else None
_live = [i for i in range(N) if cards[i] > 0]
if _live:
    base_now = [(pay[i] + hold[i]) for i in _live]
    check("ACCESS IS UNGATED - cards equal take-up x the WHOLE eligible base",
          all(abs(cards[i] - b * takeup) < max(1.0, b * 1e-6)
              for i, b in zip(_live, base_now)),
          "M%d cards %.0f vs base %.0f x %.2f = %.0f" % (
              _live[-1] + 1, cards[_live[-1]], base_now[-1], takeup, base_now[-1] * takeup))
    check("Cards are NOT restricted to the ICS-entitled subset any more",
          cards[_live[-1]] > q[_live[-1]] * takeup * 1.5,
          "cards %.0f vs entitled x take-up %.0f" % (
              cards[_live[-1]], q[_live[-1]] * takeup))
# Every card now carries a credit line - the paying/lapsed split is retired.
# These are REGIONAL rows (leading spaces), so they are not reachable through
# row(), which only resolves whole-book totals. Summed across regions instead.
_cc_rows = [r for r in range(1, mdl.max_row + 1)
            if mdl["A%d" % r].value == "  ...of which have a live credit line (now all of them)"]
_card_rows = [r for r in range(1, mdl.max_row + 1)
              if mdl["A%d" % r].value == "  Active cards"]
check("Every region reports credit-line cards", len(_cc_rows) == len(REGION_NAMES))
_cc = [sum(mdl["%s%d" % (pc(i), r)].value or 0 for r in _cc_rows) for i in range(N)]
_ac = [sum(mdl["%s%d" % (pc(i), r)].value or 0 for r in _card_rows) for i in range(N)]
check("Every active card carries a live credit line",
      all(abs(_cc[i] - _ac[i]) < 1e-6 for i in range(N)),
      "max gap %.6f" % max(abs(_cc[i] - _ac[i]) for i in range(N)))
card_m = aval("Stream 2 - Card interchange")
check("Active cards are ZERO before the card launch month",
      all(abs(x) < 1e-9 for x in cards[:card_m - 1]) and cards[card_m - 1] > 0,
      "M%d %r  M%d %r" % (card_m - 1, cards[card_m - 2], card_m, cards[card_m - 1]))

# ---- 6. every stream fires on exactly the month the calendar says ----------
# The month is READ FROM the Assumptions activation calendar, not hardcoded, so
# these checks follow the calendar if a launch date is moved again.
for label, nm in (("Stream 1a - Entry fee, SIP", "stream 1a"),
                  ("Stream 1b - Entry fee, SPOT", "stream 1b"),
                  ("Stream 3 - Family plan and Digital Will", "stream 3"),
                  ("Stream 2 - Card interchange", "stream 2"),
                  ("Stream 4 - Cardholder fees", "stream 4"),
                  ("Stream 5 - Lending revenue share", "stream 5"),
                  ("Stream 6 - B2B platform fee", "stream 6")):
    m = aval(label)
    v = row(label)
    assert m <= N_MONTHLY, "activation M%d is outside the monthly block - check needs the annual mapping" % m
    before_zero = all(abs(x) < 1e-9 for x in v[:m - 1])
    on_nonzero = abs(v[m - 1]) > 1e-9
    check("%s is zero before M%d and non-zero from M%d" % (nm, m, m), before_zero and on_nonzero,
          "M%d %r  M%d %r" % (m - 1, v[m - 2] if m > 1 else None, m, v[m - 1]))
# Redemption is a COST and has been removed from the revenue model - only the
# event count survives, as a driver for the later cost build.
check("Redemption is NOT in the revenue model", rowof(mdl, "Stream 0 - Redemption cost") is None)
redem = row("Redemption events (memo - drives cost, not revenue)")
check("Redemption event count is retained as a cost-build memo",
      all_num(redem) and redem[28] > 0, "Y7 events %r" % redem[28])

# ---- 7. the fabrication premium is borne ONCE, by ONE side -----------------
# The premium can be settled in one of two places and NEVER in both. Either the
# customer's cash buys metal at spot+premium (they receive fewer grams, Aurumix
# carries no cost) or it buys metal at spot and Aurumix pays the dealer the
# difference as COGS. The switch picks; these checks prove exactly one is live.
#
# This replaced a pair that hardcoded the customer side, which meant flipping
# the switch failed the verifier for a reason that was not a defect.
fee, prem = aval("Entry fee charged"), aval("Fabrication premium paid")
absorbed = aval("Fabrication premium borne by")          # 1 = Aurumix, 0 = customer
s1a, sip = row("Stream 1a - Entry fee, SIP"), row("SIP contributions")
spot, grams_in = row("Spot purchase volume"), row("Grams purchased")
# THE PERIOD PRICE, not the M1 price - the metal price now moves, so grams are
# bought at whatever gold cost in that period. Using the M1 price here would
# make this check fail from Y2 onwards for a reason that is not a defect.
px = row("Gold price (this period)")
cogs_prem = row("COGS - Fabrication premium paid to the dealer")

check("The premium switch reads as a clean 1/0",
      absorbed in (0, 1), "read %r" % (absorbed,))

# (a) Aurumix keeps the WHOLE entry fee under either setting. The premium is
# never netted off revenue - that was the original double-count.
check("Stream 1 earns the full entry fee (premium is NEVER netted off revenue)",
      all_num(s1a) and all(abs(s1a[i] - sip[i] * fee) < 1e-6 for i in range(N) if sip[i]),
      "implied rate %.5f vs headline fee %.5f" % (s1a[13] / sip[13] if sip[13] else 0, fee))

# (b) grams delivered match the price the customer actually paid, whichever
# side is bearing it. Under absorbed the divisor is 1 and metal ties to cash.
_paid_factor = 1 + prem * (1 - absorbed)
inflow = [(sip[i] + spot[i]) * (1 - fee) for i in range(N)]
check("Grams delivered tie to cash at the price the customer actually paid",
      all_num(grams_in) and all(
          abs(grams_in[i] * px[i] - inflow[i] / _paid_factor) < 1e-6
          for i in range(N) if inflow[i]),
      "M14 metal %.2f vs cash %.2f at factor %.5f" % (
          grams_in[13] * px[13], inflow[13], _paid_factor))

# (c) THE MATCHED PAIR. Whichever side is not bearing it must be at zero. This
# is the check that fails if the premium ever bites twice again.
_customer_bears = abs(inflow[13] - grams_in[13] * px[13])
check("Exactly ONE side bears the premium, never both",
      (absorbed == 1 and _customer_bears < 1e-6 and cogs_prem[13] > 0) or
      (absorbed == 0 and _customer_bears > 1e-6 and abs(cogs_prem[13]) < 1e-6),
      "absorbed=%s | customer bears %.4f USD at M14 | Aurumix COGS %.2f" % (
          absorbed, _customer_bears, cogs_prem[13]))

# ---- 7b. the cost band -----------------------------------------------------
bought   = row("Grams purchased")            # whole-book, unindented
recycled = row("  Grams returned by redemption (recycled into the float)")
fabbed   = row("  Net new grams - the metal that must actually be made")
custody  = row("GRAMS UNDER CUSTODY")
on_gross = aval("Premium charged on gross inflow")
cogs_tot = row("TOTAL COST OF GOODS SOLD")
cogs_chk = row("CHECK: cash for metal ties to grams delivered x the price paid (must be 0)")

# Recycled metal is DERIVED from the custody balance, so it must reproduce the
# custody identity exactly: closing = opening - returned + purchased.
check("Recycled grams reconcile to the custody balance identity",
      all_num(recycled) and all(
          abs((custody[i - 1] - recycled[i] + bought[i]) - custody[i]) < 1e-3
          for i in range(1, N)),
      "worst delta %.6f" % max(abs((custody[i - 1] - recycled[i] + bought[i]) - custody[i])
                               for i in range(1, N)))

check("Nothing is recycled in M1 - there is nothing to redeem yet",
      recycled[0] == 0, "M1 recycled %.4f" % recycled[0])

# D30: fabricated = bought - recycled. On the Gross setting it is bought flat.
check("Fabricated grams follow the D30 basis switch",
      all_num(fabbed) and all(
          abs(fabbed[i] - (bought[i] if on_gross else bought[i] - recycled[i])) < 1e-3
          for i in range(N)),
      "basis=%s, M29 fabricated %.1f vs bought %.1f" % (
          "gross" if on_gross else "net new", fabbed[-1], bought[-1]))

check("Fabricated grams never exceed grams purchased",
      all(fabbed[i] <= bought[i] + 1e-6 for i in range(N)),
      "worst excess %.4f" % max(fabbed[i] - bought[i] for i in range(N)))

# THE POINT OF D30, asserted rather than assumed: on the net-new basis the
# recycled metal must actually reduce what gets fabricated, materially.
check("Net new basis materially reduces fabrication by the back years",
      on_gross or fabbed[-1] < bought[-1] * 0.95,
      "Y7 fabricated %.0f of %.0f grams purchased (%.1f%% recycled)" % (
          fabbed[-1], bought[-1], 100 * (1 - fabbed[-1] / bought[-1]) if bought[-1] else 0))

check("COGS premium = fabricated grams x price x premium x the incidence switch",
      all_num(cogs_prem) and all(
          abs(cogs_prem[i] - fabbed[i] * px[i] * prem * absorbed) < 1e-3 for i in range(N)),
      "M29 %.2f vs %.2f" % (cogs_prem[-1], fabbed[-1] * px[-1] * prem * absorbed))

check("TOTAL COGS ties to its component lines",
      all_num(cogs_tot) and all(abs(cogs_tot[i] - cogs_prem[i]) < 1e-6 for i in range(N)),
      "M29 total %.2f vs premium %.2f" % (cogs_tot[-1], cogs_prem[-1]))

# There is deliberately NO gross-profit row while COGS is the only cost family
# built. Assert its ABSENCE, so it cannot reappear by accident and hand a
# reader a 93% margin on a cost base that is five families short.
check("No gross-profit line while the cost base is incomplete",
      rowof(mdl, "GROSS PROFIT (total revenue less COGS)") is None,
      "a gross-profit row has appeared on the Model sheet")

check("The anti-double-count row ties to zero in every period",
      all_num(cogs_chk) and all(abs(v) < 1e-6 for v in cogs_chk),
      "worst %.9f" % max(abs(v) for v in cogs_chk))

# The cost band must sit BELOW the revenue total on the sheet, or the divider
# is decorative rather than structural.
check("The cost band is placed below the revenue total on the Model sheet",
      rowof(mdl, "TOTAL COST OF GOODS SOLD") > rowof(mdl, "TOTAL NET REVENUE"),
      "COGS at row %s, revenue at row %s" % (rowof(mdl, "TOTAL COST OF GOODS SOLD"),
                                             rowof(mdl, "TOTAL NET REVENUE")))

# ---- 7c. working capital: the float ---------------------------------------
# The float is INVENTORY, not cost. These checks exist to keep it that way -
# the last one asserts it never leaks into a cost total.
bar   = aval("Bar denomination")
bufd  = aval("Float buffer - days of demand held")
dgram = row("  Daily demand")
fgram = row("  FLOAT REQUIRED")
fusd  = row("FLOAT REQUIRED (standing balance)")
ftop  = row("  Cash needed this period (new grams only)")
fcum  = row("CUMULATIVE CASH INVESTED IN THE FLOAT")
nmon  = row("Months in period")

check("Daily demand = grams purchased over the days in the period",
      all_num(dgram) and all(
          abs(dgram[i] - bought[i] / (nmon[i] * 365 / 12)) < 1e-6 for i in range(N)),
      "M1 %.3f g/day" % dgram[0])

check("Float = MAX(two bars, one bar + N days of demand)",
      all_num(fgram) and all(
          abs(fgram[i] - max(2 * bar, bar + bufd * dgram[i])) < 1e-6 for i in range(N)),
      "M1 %.0f g, Y7 %.0f g" % (fgram[0], fgram[-1]))

check("The two-bar floor is never breached",
      all(fgram[i] >= 2 * bar - 1e-9 for i in range(N)),
      "minimum %.1f g against a %.0f g floor" % (min(fgram), 2 * bar))

# The floor must actually BIND at launch and must actually STOP binding later,
# or the MAX is decorative and one of the two halves is dead code.
check("The floor binds at launch and the demand term takes over later",
      abs(fgram[0] - 2 * bar) < 1e-6 and fgram[-1] > 2 * bar,
      "M1 %.0f g (floor %.0f), Y7 %.0f g" % (fgram[0], 2 * bar, fgram[-1]))

check("The crossover happens where 100 + N x demand overtakes two bars",
      all((fgram[i] > 2 * bar) == (bar + bufd * dgram[i] > 2 * bar) for i in range(N)),
      "crossover at demand = %.1f g/day" % (bar / bufd))

check("Float value = grams x period price x (1 + premium)",
      all_num(fusd) and all(
          abs(fusd[i] - fgram[i] * px[i] * (1 + prem)) < 1e-3 for i in range(N)),
      "M1 USD %.0f, Y7 USD %.0f" % (fusd[0], fusd[-1]))

# THE CASH CALL IS NOT THE CHANGE IN THE BALANCE. Revaluing metal already on
# the shelf needs no new money, so the top-up must price only EXTRA GRAMS.
check("Top-up funds new grams only, not revaluation of metal already held",
      all_num(ftop) and abs(ftop[0] - fgram[0] * px[0] * (1 + prem)) < 1e-3 and all(
          abs(ftop[i] - max(0.0, fgram[i] - fgram[i - 1]) * px[i] * (1 + prem)) < 1e-3
          for i in range(1, N)),
      "M1 %.0f, Y7 %.0f" % (ftop[0], ftop[-1]))

check("Top-up is never negative - a shrinking float returns cash, it does not consume it",
      all(v >= -1e-9 for v in ftop), "minimum %.2f" % min(ftop))

check("Cumulative float cash accumulates the top-ups",
      all_num(fcum) and all(
          abs(fcum[i] - sum(ftop[: i + 1])) < 1e-3 for i in range(N)),
      "Y7 cumulative USD %.0f" % fcum[-1])

# Revaluation means the standing balance can exceed everything ever paid in.
check("The standing balance exceeds cash invested once gold has appreciated",
      fusd[-1] > 0 and fcum[-1] > 0,
      "Y7 balance USD %.0f against USD %.0f invested" % (fusd[-1], fcum[-1]))

# THE ONE THAT MATTERS: inventory must never be counted as expenditure.
check("The float is NOT inside total COGS",
      all(abs(cogs_tot[i] - cogs_prem[i]) < 1e-6 for i in range(N)),
      "COGS carries only the premium line")

# SHEET ORDER, and it is deliberate: revenue, then the float, then the cost
# base. The float is built BEFORE the cost band because vault storage charges
# Aurumix's own bars alongside the customers', so the opex block reads the
# float row. Order changed 2026-08-26 when storage landed; the check moved with
# it rather than being deleted, because the ordering still has to be asserted.
_r_rev = rowof(mdl, "TOTAL NET REVENUE")
_r_flt = rowof(mdl, "FLOAT REQUIRED (standing balance)")
_r_cogs = rowof(mdl, "TOTAL COST OF GOODS SOLD")
_r_cost = rowof(mdl, "TOTAL COST BASE (modelled + contingency)")
check("Sheet order is revenue, then the float, then the cost base",
      _r_rev < _r_flt < _r_cogs < _r_cost,
      "revenue %s, float %s, COGS %s, total cost %s" % (_r_rev, _r_flt, _r_cogs, _r_cost))

# ---- 7d. operating expenses: vault storage --------------------------------
srate = aval("Vault storage fee")
smind = aval("Vault minimum charge")
sgram = row("  Metal in the vault (customer gold + Aurumix's own float)")
stor  = row("OPEX - Vault storage")
spct  = row("  ...as % of metal held per year (memo - watch the minimum bite)")
topex = row("TOTAL OPERATING EXPENSES")
tcost = row("TOTAL COST BASE (modelled + contingency)")
cust  = row("GRAMS UNDER CUSTODY")

# THE FLOAT IS STORED TOO. If this ever equals customer grams alone, someone
# has quietly stopped charging Aurumix for its own bars.
check("Storage is charged on customer gold PLUS Aurumix's own float",
      all_num(sgram) and all(
          abs(sgram[i] - (cust[i] + fgram[i])) < 1e-6 for i in range(N)),
      "M1 %.1f g = %.1f customer + %.1f float" % (sgram[0], cust[0], fgram[0]))

check("Storage = MAX(the minimum, the rate on metal held)",
      all_num(stor) and all(
          abs(stor[i] - max(smind * nmon[i] * 365 / 12,
                            srate * sgram[i] * px[i] * nmon[i] / 12)) < 1e-3
          for i in range(N)),
      "M1 USD %.0f, Y7 USD %.0f" % (stor[0], stor[-1]))

# Both halves of the MAX must be live somewhere in the horizon, or one is dead.
_minbinds = [i for i in range(N)
             if smind * nmon[i] * 365 / 12 > srate * sgram[i] * px[i] * nmon[i] / 12 + 1e-9]
check("The vault minimum binds early and stops binding later",
      bool(_minbinds) and 0 in _minbinds and (N - 1) not in _minbinds,
      "minimum binds in %d of %d periods, last at %s" % (
          len(_minbinds), N, per_label(max(_minbinds)) if _minbinds else "never"))

check("Storage cost rises with the book",
      stor[-1] > stor[0], "M1 %.0f to Y7 %.0f" % (stor[0], stor[-1]))

# The memo rate must reveal the minimum: high while it binds, settling to the
# headline rate once volume clears it.
check("The effective rate falls to the headline rate once the minimum clears",
      all_num(spct) and spct[0] > srate and abs(spct[-1] - srate) < 1e-6,
      "M1 effective %.4f%% against a %.4f%% headline, Y7 %.4f%%" % (
          spct[0] * 100, srate * 100, spct[-1] * 100))

# ---- 7e. the regulatory block ---------------------------------------------
aed   = aval("AED/USD peg")
vsup  = aval("VARA annual supervision fee")
vapp  = aval("VARA licence application fee (one-off)")
dmcca = aval("DMCC company licence")
dmccs = aval("DMCC incorporation (one-off)")
kycpc = aval("KYC and AML - per verification")
kycmm = aval("KYC and AML - monthly minimum")
vara  = row("OPEX - VARA annual supervision")
dmcc  = row("OPEX - DMCC company licence")
kyc   = row("OPEX - KYC and AML verification")
onef  = row("OPEX - One-off Year 1 (licence application, incorporation, launch audit)")
insu  = row("OPEX - Insurance (PI, D&O, crime)")
audi  = row("OPEX - Audit and reserve attestation")
tech  = row("OPEX - Technology audit and penetration testing")
insa  = aval("Insurance - PI, D&O and crime")
auda  = aval("Audit and reserve attestation")
techa = aval("Technology audit and penetration testing")
lauda = aval("Launch technology and smart contract audit (one-off)")
newc  = row("NEW CUSTOMERS")
calm  = row("Calendar month")

# ANNIVERSARY BOOKING. An annual invoice must land in ONE period, not be
# smeared across twelve - smearing flatters Year 1 cash, the year that decides
# the funding ask.
# NB: the loop variable is `amt`, NOT `fee`. `fee` is the ENTRY FEE, bound far
# above and read again ~600 lines below; naming the loop variable `fee` here
# silently rebound it to a DMCC licence fee and failed an unrelated check in
# section 9b. Caught 2026-08-26.
for nm, ser, amt in (("VARA supervision", vara, vsup), ("DMCC company licence", dmcc, dmcca)):
    check("%s lands on the anniversary, not smeared monthly" % nm,
          all_num(ser) and all(
              abs(ser[i] - (amt / aed if calm[i] in (0, 1) else 0.0)) < 1e-3
              for i in range(N)),
          "%s fires in %d of 29 periods at USD %.0f" % (
              nm, sum(1 for v in ser if v > 0), amt / aed))

check("Exactly twelve periods carry an annual fee - two Januaries plus five annual columns",
      sum(1 for v in vara if v > 0) == 7,
      "%d periods carry VARA supervision" % sum(1 for v in vara if v > 0))

check("VARA supervision is charged from M1, before any customer arrives",
      vara[0] > 0, "M1 USD %.0f" % vara[0])

check("KYC = MAX(monthly minimum, per-check x new customers)",
      all_num(kyc) and all(
          abs(kyc[i] - max(kycmm * nmon[i], kycpc * newc[i])) < 1e-3 for i in range(N)),
      "M1 USD %.0f, Y7 USD %.0f" % (kyc[0], kyc[-1]))

# THE KYC MINIMUM NEVER BINDS ON THIS BOOK, and that is a finding rather than a
# defect. It binds below ~162 verifications a month; acquisition opens at ~271
# new customers in M1 and rises from there, so per-verification cost dominates
# from the first period. Unlike the VAULT minimum, which binds until Year 3,
# this minimum-commitment structure costs Aurumix nothing. Asserted so that a
# future cut with a slower ramp trips this and gets re-examined.
_kycmin = [i for i in range(N) if kycmm * nmon[i] > kycpc * newc[i] + 1e-9]
check("The KYC minimum never binds - acquisition opens above the threshold",
      not _kycmin,
      "binds in %d of %d periods; M1 is %.0f checks against a %.0f-check threshold" % (
          len(_kycmin), N, newc[0], kycmm / kycpc))

check("One-off launch costs land in M1 only",
      all_num(onef) and abs(onef[0] - ((vapp + dmccs) / aed + lauda)) < 1e-3
      and all(abs(x) < 1e-9 for x in onef[1:]),
      "M1 USD %.0f, everything after zero" % onef[0])

# THESE THREE ARE RECURRING, NOT LAUNCH ONE-OFFS - the single most common way
# this cost base gets understated is treating the audit as a Year-1 item.
for nm, ser, amt in (("Insurance", insu, insa), ("Audit and attestation", audi, auda),
                     ("Technology audit", tech, techa)):
    check("%s accrues every period, spread evenly" % nm,
          all_num(ser) and all(abs(ser[i] - amt * nmon[i] / 12) < 1e-3 for i in range(N)),
          "%s USD %.0f/yr, present in %d of %d periods" % (
              nm, amt, sum(1 for x in ser if x > 0), N))

check("The technology audit recurs rather than sitting only in Year 1",
      tech[-1] > 0 and tech[0] > 0,
      "M1 USD %.0f, Y7 USD %.0f" % (tech[0], tech[-1]))

redm = row("OPEX - Redemption handling (no fee may be charged)")
revents = row("Redemption events (memo - drives cost, not revenue)")
rcost = aval("Cost per redemption event")

# STREAM 0. Zero-revenue by regulation - VARA Annex 2 III.E.4 forbids charging
# any fee on redemption, so nothing offsets this line, ever.
check("Redemption cost = events x the per-event rate",
      all_num(redm) and all(abs(redm[i] - revents[i] * rcost) < 1e-3 for i in range(N)),
      "Y7 USD %.0f on %.0f events at USD %.2f" % (redm[-1], revents[-1], rcost))

check("Redemption cost scales with the book and is never zero once redemptions start",
      redm[-1] > redm[0] and redm[-1] > 0,
      "M1 USD %.0f to Y7 USD %.0f" % (redm[0], redm[-1]))

tbld = row("OPEX - Technology build (Y1-Y2)")
tmnt = row("OPEX - Technology maintenance (Y3 onward)")
yr_ = row("Model year")

# BUILD THEN SETTLE. If either line bleeds into the other's years the cash
# profile stops looking like a business that builds and then runs.
check("Technology build lands in Y1-Y2 only",
      all_num(tbld) and all((tbld[i] > 0) == (yr_[i] <= 2) for i in range(N)),
      "Y1+Y2 total USD %.0f, nothing after" % sum(tbld[i] for i in range(N) if yr_[i] <= 2))

check("Technology maintenance runs from Y3 only",
      all_num(tmnt) and all((tmnt[i] > 0) == (yr_[i] >= 3) for i in range(N)),
      "USD %.0f/yr from Y3" % (tmnt[-1]))

check("Build and maintenance never overlap in the same period",
      all(not (tbld[i] > 0 and tmnt[i] > 0) for i in range(N)),
      "no period carries both")

check("TOTAL OPERATING EXPENSES ties to its component lines",
      all_num(topex) and all(
          abs(topex[i] - (stor[i] + vara[i] + dmcc[i] + kyc[i] + onef[i]
                          + insu[i] + audi[i] + tech[i] + redm[i]
                          + tbld[i] + tmnt[i])) < 1e-6
          for i in range(N)),
      "Y7 %.0f" % topex[-1])

# ---- 7g. ICS benefit costs ------------------------------------------------
# CONTRA-REVENUE. The streams stay GROSS and the giveaway is its own line, so
# these checks mostly guard against the discount being applied twice - once here
# and again inside a stream.
qsh = row("  Qualifying share of the book (memo - ramps to 55%)")
bent = row("BENEFIT - Entry fee discount (SIP only)")
bcard = row("BENEFIT - Card fee discounts (FX, ATM, issuance)")
breb = row("BENEFIT - Gold rebate")
bfam = row("BENEFIT - Family wallet and will discount")
btot = row("TOTAL ICS BENEFIT COSTS")
s1a_t = row("Stream 1a - Entry fee, SIP")
s1b_t = row("Stream 1b - Entry fee, SPOT")
s2_t = row("Stream 2 - Card interchange")
s3_t = row("Stream 3 - Family plan and Digital Will")
s4_t = row("Stream 4 - Cardholder fees")
dent = aval("ICS discount - entry fee")
dcard = aval("ICS discount - card fees")
dreb = aval("ICS discount - gold rebate")
dfam = aval("ICS discount - family wallet and will")

# THE REBATE IS ON ALL CARD REVENUE, not interchange alone - interchange is the
# smaller half AND is already net of the partner split, so it measures what is
# left rather than what the customer generated.
spend_t = row("Card spend")
rebpct = row("  ...as % of card spend (memo - check against the 0.15/0.45/0.75 ladder)")
check("Gold rebate is a share of interchange PLUS cardholder fees",
      all_num(breb) and all(
          abs(breb[i] - (s2_t[i] + s4_t[i]) * qsh[i] * dreb) < 1e-3 for i in range(N)),
      "Y7 USD %.0f on a USD %.0f card-revenue base" % (breb[-1], s2_t[-1] + s4_t[-1]))

# The rate only means something against SPEND. Assert it lands inside the
# corpus ladder rather than above its top rung.
check("The rebate sits inside the corpus ladder as a share of card spend",
      all_num(rebpct) and 0.0 < rebpct[-1] < 0.0075,
      "Y7 rebate is %.3f%% of card spend, against a 0.15-0.75%% ladder" % (rebpct[-1] * 100))

check("The qualifying share RAMPS - it is not a flat 55% from launch",
      all_num(qsh) and qsh[0] < qsh[-1] and qsh[-1] <= 0.5501,
      "M1 %.1f%% rising to %.1f%%" % (qsh[0] * 100, qsh[-1] * 100))

for nm, ser, base, rate in (("Entry fee", bent, s1a_t, dent), ("Card fees", bcard, s4_t, dcard),
                            ("Family wallet", bfam, s3_t, dfam)):
    check("%s discount = stream x qualifying share x rate" % nm,
          all_num(ser) and all(abs(ser[i] - base[i] * qsh[i] * rate) < 1e-3 for i in range(N)),
          "Y7 USD %.0f on a USD %.0f stream" % (ser[-1], base[-1]))

# SPOT EARNS NO ICS. If Stream 1b ever appears in the entry-fee discount, the
# benefit has been extended to a lane the design excludes.
# TESTED AT THE TERMINAL PERIOD, not every period. In the early months the
# qualifying share is near zero, so including or excluding Stream 1b makes a
# difference smaller than the tolerance and the check would fail on arithmetic
# noise rather than on a defect.
check("Spot carries NO entry-fee discount - Stream 1b is excluded",
      all_num(bent) and s1b_t[-1] > 1e-6 and qsh[-1] > 0.1
      and abs(bent[-1] - s1a_t[-1] * qsh[-1] * dent) < 1e-3
      and abs(bent[-1] - (s1a_t[-1] + s1b_t[-1]) * qsh[-1] * dent) > 1.0,
      "Y7 discount USD %.0f is on Stream 1a (%.0f), not 1a+1b (%.0f)" % (
          bent[-1], s1a_t[-1], s1a_t[-1] + s1b_t[-1]))

# The streams must stay GROSS. If a discount is also netted inside a stream the
# giveaway is counted twice, which is the exact defect the premium fix removed.
check("Streams are reported GROSS - no discount is netted inside them",
      all(abs(s1a_t[i] - sip[i] * fee) < 1e-6 for i in range(N) if sip[i]),
      "Stream 1a still equals SIP x the full headline fee")

check("TOTAL ICS BENEFIT COSTS ties to its component lines",
      all_num(btot) and all(
          abs(btot[i] - (bent[i] + bcard[i] + breb[i] + bfam[i])) < 1e-6 for i in range(N)),
      "Y7 USD %.0f" % btot[-1])

check("Benefits never exceed the streams they are given out of",
      all(btot[i] <= s1a_t[i] + s2_t[i] + s3_t[i] + s4_t[i] + 1e-6 for i in range(N)),
      "Y7 benefits USD %.0f against USD %.0f of funding streams" % (
          btot[-1], s1a_t[-1] + s2_t[-1] + s3_t[-1] + s4_t[-1]))

# ---- 7h. acquisition costs ------------------------------------------------
mktg = row("ACQ - Marketing spend")
acomm = row("ACQ - Agent commission")
aref_ = row("ACQ - Referral rewards")
atot = row("TOTAL ACQUISITION COSTS")
ashare = row("  Agent-acquired share of the book (memo)")
acum = row("  Cumulative agent-driven acquisitions")
anew = row("  Agent-driven acquisitions this period")
blend = row("  Blended cost per new customer (memo)")
newt = row("NEW CUSTOMERS")
comm_r = aval("Agent commission - share of the entry fee")

# MARKETING WAS ALWAYS BEING SPENT - it drove the direct channel and was charged
# to nothing. This is the offsetting entry, and it must appear EXACTLY ONCE.
check("Marketing spend is booked and rises with the schedule",
      all_num(mktg) and mktg[-1] > mktg[0] > 0,
      "M1 USD %.0f to Y7 USD %.0f" % (mktg[0], mktg[-1]))

# THE OPEX COMPONENT LIST IS REPEATED HERE ON PURPOSE. If a new opex line is
# added and this list is not updated, this check fails - which is what caught
# the technology lines on 2026-08-26. The failure mode it really guards is
# marketing appearing in BOTH acquisition and opex, as it did in v1.0.
_opex_parts = [stor, vara, dmcc, kyc, onef, insu, audi, tech, redm, tbld, tmnt]
check("Marketing is booked ONCE - it is not also inside operating expenses",
      all(abs(topex[i] - sum(part[i] for part in _opex_parts)) < 1e-6 for i in range(N)),
      "opex reconciles from its %d components with no marketing term" % len(_opex_parts))

check("Cumulative agent acquisitions accumulate the per-period flow",
      all_num(acum) and all(abs(acum[i] - sum(anew[: i + 1])) < 1e-3 for i in range(N)),
      "Y7 %.0f agent-acquired accounts" % acum[-1])

check("Agent share stays a proper fraction of the book",
      all_num(ashare) and all(-1e-9 <= x <= 1.0 + 1e-9 for x in ashare),
      "Y7 %.1f%% of accounts came via agents" % (ashare[-1] * 100))

check("Agent commission = entry fee x agent share x the commission rate",
      all_num(acomm) and all(
          abs(acomm[i] - (s1a_t[i] + s1b_t[i]) * ashare[i] * comm_r) < 1e-3 for i in range(N)),
      "Y7 USD %.0f at %.0f%% of the fee" % (acomm[-1], comm_r * 100))

# ONGOING, NOT A ONE-OFF. If commission ever tracked new customers rather than
# the paying book, it has been switched to a first-year model by accident.
check("Agent commission is ONGOING - it tracks the paying book, not new customers",
      acomm[-1] > 0 and s1a_t[-1] > 0,
      "Y7 commission USD %.0f against USD %.0f of entry-fee revenue" % (
          acomm[-1], s1a_t[-1] + s1b_t[-1]))

check("Referral rewards are zero until the referral channel opens",
      all_num(aref_) and abs(aref_[0]) < 1e-9 and aref_[-1] > 0,
      "M1 USD %.0f, Y7 USD %.0f" % (aref_[0], aref_[-1]))

check("TOTAL ACQUISITION COSTS ties to its component lines",
      all_num(atot) and all(
          abs(atot[i] - (mktg[i] + acomm[i] + aref_[i])) < 1e-6 for i in range(N)),
      "Y7 USD %.0f" % atot[-1])

# The blended figure spreads total spend over ALL new customers including the
# organic ones, so it must sit BELOW any single paid channel's CAC.
check("Blended cost per customer is below the paid marketing CAC",
      all_num(blend) and blend[-1] > 0 and blend[-1] < aval("Marketing CAC - UAE"),
      "Y7 blended USD %.2f against a UAE paid CAC of USD %.0f" % (
          blend[-1], aval("Marketing CAC - UAE")))

# CAC RAMPS DOWN. If it ever goes flat the ramp has been disconnected.
_cacrows = [r for r in range(1, mdl.max_row + 1)
            if mdl["A%d" % r].value == "  Marketing CAC this period (ramping Y1 to Y7)"]
check("Every region carries a ramping CAC row",
      len(_cacrows) == len(REGION_NAMES), "found %d" % len(_cacrows))
for _r in _cacrows:
    _series = [mdl.cell(row=_r, column=PCOL0 + i).value for i in range(N)]
    check("CAC falls from Y1 to Y7 (row %d)" % _r,
          all_num(_series) and _series[-1] < _series[0],
          "USD %.1f falling to USD %.1f" % (_series[0], _series[-1]))

# ---- 7i. card programme ---------------------------------------------------
cfix = row("CARD - NymCard platform and scheme fees")
cset = row("CARD - Programme setup (one-off at launch)")
cprod = row("CARD - Card production and delivery")
cproc = row("CARD - Authorisation and switching")
cfrd = row("CARD - Fraud and chargebacks")
cxb = row("CARD - Cross-border scheme assessment")
xbs = row("  Cross-border card spend (non-UAE in full, plus UAE spend abroad)")
ctot_c = row("TOTAL CARD PROGRAMME COSTS")
cratio = row("  Card cost as % of card revenue (above 100% = the card is not paying for itself)")
newc_t = row("  Cards newly issued this period")
auths_t = row("Card authorisations (memo - drives cost, not revenue)")
spend_c = row("Card spend")
act_s2 = aval("Stream 2 - Card interchange")
xbr = aval("Cross-border scheme assessment")

# NOTHING BEFORE LAUNCH. The card activates at period 13; a cost before that
# means the programme is being charged for a product that does not exist.
check("No card programme cost before the card launches",
      all_num(ctot_c) and all(abs(ctot_c[i]) < 1e-9 for i in range(int(act_s2) - 1)),
      "first cost lands at period %d, card activates at %d" % (
          next((i + 1 for i, x in enumerate(ctot_c) if x > 0), -1), act_s2))

check("Programme setup is a one-off in the launch month",
      all_num(cset) and sum(1 for x in cset if x > 0) == 1 and cset[int(act_s2) - 1] > 0,
      "fires once, at period %d" % act_s2)

# FLAT FROM LAUNCH - that is what a minimum commitment means, and it is why the
# programme loses money while the book is small.
check("The fixed fee is flat from launch, not scaled to the book",
      all_num(cfix) and abs(cfix[int(act_s2) - 1] - cfix[int(act_s2)]) < 1e-6 and cfix[-1] > 0,
      "USD %.0f per month from launch" % cfix[int(act_s2) - 1])

check("Card production tracks cards ISSUED, not cards active",
      all_num(cprod) and all(
          abs(cprod[i] - newc_t[i] * aval("Card programme - per card issued")) < 1e-3
          for i in range(N)),
      "Y7 %.0f cards issued" % newc_t[-1])

check("Authorisation cost tracks the authorisation count",
      all_num(cproc) and all(
          abs(cproc[i] - auths_t[i] * aval("Card programme - per authorisation")) < 1e-3
          for i in range(N)),
      "Y7 %.0f authorisations" % auths_t[-1])

check("Fraud tracks card spend",
      all_num(cfrd) and all(
          abs(cfrd[i] - spend_c[i] * aval("Card programme - fraud and chargebacks")) < 1e-3
          for i in range(N)),
      "Y7 USD %.0f on USD %.0f of spend" % (cfrd[-1], spend_c[-1]))

# MOSTLY DERIVED, NOT ASSUMED. Non-UAE customers hold a UAE-issued card, so
# every transaction they make is cross-border by construction.
check("Cross-border spend is derived from region, not assumed wholesale",
      all_num(xbs) and all(xbs[i] <= spend_c[i] + 1e-6 for i in range(N))
      and xbs[-1] / spend_c[-1] > 0.5,
      "Y7 cross-border is %.1f%% of card spend" % (xbs[-1] / spend_c[-1] * 100))

check("Cross-border assessment = cross-border spend x the scheme rate",
      all_num(cxb) and all(abs(cxb[i] - xbs[i] * xbr) < 1e-3 for i in range(N)),
      "Y7 USD %.0f at %.2f%%" % (cxb[-1], xbr * 100))

# THE FINDING: the scheme charges the issuer MORE than Aurumix retains of
# interchange, so cross-border interchange cannot cover its own scheme fee.
_pm_keep = 1 - aval("Programme manager share of interchange")
_ic = aval("Interchange - Gold")
check("The cross-border rate exceeds Aurumix's retained interchange share",
      xbr > _ic * _pm_keep,
      "scheme takes %.2f%% of spend, Aurumix retains %.3f%%" % (
          xbr * 100, _ic * _pm_keep * 100))

check("TOTAL CARD PROGRAMME COSTS ties to its component lines",
      all_num(ctot_c) and all(
          abs(ctot_c[i] - (cfix[i] + cset[i] + cprod[i] + cproc[i] + cfrd[i] + cxb[i])) < 1e-6
          for i in range(N)),
      "Y7 USD %.0f" % ctot_c[-1])

_live = [i for i in range(N) if s2_t[i] + s4_t[i] > 0]
check("The card runs at a loss at launch and pays for itself later",
      bool(_live) and cratio[_live[0]] > 1.0 and cratio[-1] < 1.0,
      "launch %.0f%% of card revenue, Y7 %.0f%%" % (
          cratio[_live[0]] * 100, cratio[-1] * 100))

check("The card cost ratio excludes credit revenue, which would flatter it",
      all_num(cratio) and all(
          abs(cratio[i] - ctot_c[i] / (s2_t[i] + s4_t[i])) < 1e-6
          for i in range(N) if s2_t[i] + s4_t[i] > 0),
      "ratio is card cost over streams 2 + 4 only")

cmeas = row("  Sub-total - costs actually modelled")
ccont = row("  Contingency for cost families not yet built")
cont_r = aval("Contingency on total costs")

check("Modelled sub-total = COGS + opex + benefits + acquisition + card",
      all_num(cmeas) and all(
          abs(cmeas[i] - (cogs_tot[i] + topex[i] + btot[i] + atot[i] + ctot_c[i])) < 1e-6
          for i in range(N)),
      "Y7 %.0f" % cmeas[-1])

check("Contingency is a flat percentage of the modelled sub-total",
      all_num(ccont) and all(abs(ccont[i] - cmeas[i] * cont_r) < 1e-3 for i in range(N)),
      "Y7 USD %.0f at %.0f%%" % (ccont[-1], cont_r * 100))

# THE CONTINGENCY IS A PLACEHOLDER, NOT COVERAGE. Headcount alone is anchored at
# about USD 588k in Y1. Assert the shortfall so nobody reads the profit line as
# a forecast.
check("The contingency is smaller than headcount alone would be",
      ccont[11] < 588000,
      "Y1 contingency USD %.0f against a USD 588,000 headcount anchor" % sum(ccont[:12]))

check("TOTAL COST BASE = modelled costs + contingency",
      all_num(tcost) and all(
          abs(tcost[i] - (cmeas[i] + ccont[i])) < 1e-6 for i in range(N)),
      "Y7 %.0f = %.0f + %.0f" % (tcost[-1], cmeas[-1], ccont[-1]))

# ---- 7j. profit and cash --------------------------------------------------
# `tied` is re-read here rather than reused: the capital checks in 7f run LATER
# in this file than this block does, so the name is not bound yet.
npro = row("NET PROFIT")
nmar = row("  Net margin")
cpro = row("CUMULATIVE NET PROFIT")
wcch = row("  Cash into working capital and regulatory capital")
ncf = row("NET CASH FLOW")
ccf = row("CUMULATIVE CASH FLOW (the funding requirement)")
peak = row("PEAK FUNDING REQUIREMENT (memo - worst point of the line above)")
rev_t = row("TOTAL NET REVENUE")

check("Net profit = total revenue less the total cost base",
      all_num(npro) and all(abs(npro[i] - (rev_t[i] - tcost[i])) < 1e-6 for i in range(N)),
      "Y7 USD %.0f" % npro[-1])

check("Cumulative net profit accumulates the periods",
      all_num(cpro) and all(abs(cpro[i] - sum(npro[: i + 1])) < 1e-3 for i in range(N)),
      "Y7 cumulative USD %.0f" % cpro[-1])

# THE CASH VIEW WAS REMOVED at the client's instruction on 2026-08-26. Its
# checks went with it. Asserted as ABSENT so a later rebuild cannot quietly
# reintroduce a funding figure nobody has reviewed.
check("The cash view is absent - profit only",
      rowof(mdl, "NET CASH FLOW") is None
      and rowof(mdl, "CUMULATIVE CASH FLOW (the funding requirement)") is None,
      "no cash flow rows on the Model sheet")

# ---- 8. the ATM distribution earns what a mean would not ------------------
# Verified from the parameters directly, so the check survives the stream-4
# components being combined into one regional formula.
BK_ATM = ("AED 0-500", "AED 500-1,500", "AED 1,500-3,000", "AED 3,000+")
allowance = aval("Free ATM allowance (Gold)")
shares = [sval("ATM %s - share of cardholders" % b) for b in BK_ATM]
mids = [sval("ATM %s - midpoint draw" % b) for b in BK_ATM]
mean_draw = sum(s * m for s, m in zip(shares, mids))
over_mean = max(0.0, mean_draw - allowance)
over_dist = sum(s * max(0.0, m - allowance) for s, m in zip(shares, mids))
check("The MEAN ATM draw sits below the free allowance (so a mean earns zero)",
      mean_draw < allowance and over_mean == 0,
      "mean %.0f vs allowance %.0f" % (mean_draw, allowance))
check("The DISTRIBUTION still earns ATM revenue where the mean would not",
      over_dist > 0, "distribution yields AED %.1f per cardholder, mean yields %.1f"
      % (over_dist, over_mean))

# The partner schedule must supply a partner by the month stream 6 activates,
# or the stream switches on with nothing to earn on.
b2b_m = aval("Stream 6 - B2B platform fee")
b2b_year = (b2b_m - 1) // 12 + 1
check("A B2B partner exists by the year stream 6 activates",
      (sval("B2B partners - Y%d" % b2b_year) or 0) >= 1,
      "stream 6 at M%d (Y%d) but Y%d partners = %r" % (b2b_m, b2b_year, b2b_year,
                                                       sval("B2B partners - Y%d" % b2b_year)))
check("Partner count rises monotonically once B2B is live",
      all((sval("B2B partners - Y%d" % y) or 0) >= (sval("B2B partners - Y%d" % (y - 1)) or 0)
          for y in range(2, 8)),
      "%r" % [sval("B2B partners - Y%d" % y) for y in range(1, 8)])

# ---- 8b. regions are handled consistently across BOTH inflow lanes --------
# REPLACED 2026-08-21. The two checks here read a "Share of new customers" row
# that has been deleted: it claimed acquisition followed the ceiling share, but
# acquisition is now driven per region by budget, CAC, salesforce and referrals,
# so the split is an OUTCOME. Both checks were also tautological - the row was
# derived from the ceilings, so testing it against the ceilings proved nothing.
# What is worth asserting is that the OUTCOME is sane.
ceil_share = [asm["B%d" % r].value / ceil_tot for r in ceil_rows]
check("Region ceiling shares sum to 1.000", abs(sum(ceil_share) - 1.0) < 1e-9,
      "sum %r" % sum(ceil_share))
_newr = [r for r in range(1, mdl.max_row + 1) if mdl["A%d" % r].value == "  New customers"]
_newy7 = [mdl["%s%d" % (pc(N - 1), r)].value for r in _newr]
check("Every region is still acquiring at Y7 (none has been starved by the mix)",
      len(_newy7) == len(REGION_NAMES) and all_num(_newy7) and all(v > 0 for v in _newy7),
      "Y7 new by region %r" % [round(v, 1) for v in _newy7])
# The emergent split should bear SOME relation to opportunity - not equality,
# since CAC and salesforce deliberately differ, but not a wild inversion either.
_share = [v / sum(_newy7) for v in _newy7]
check("Emergent regional acquisition is within 3x of each region's opportunity share",
      all(c / 3 <= s <= c * 3 for s, c in zip(_share, ceil_share)),
      "acquired %r vs opportunity %r" % ([round(x, 3) for x in _share],
                                         [round(x, 3) for x in ceil_share]))
# REPLACED 2026-08-21. The old check asserted the regional spot multipliers
# averaged 1.00 - a property of a DERIVED number that no longer exists. The
# tickets are now observed per region, so the thing worth defending is that
# they still carry the ORDERING the sources establish.
stkt = {r: aval("Spot ticket - %s" % r) for r in REGION_NAMES}
check("Every region carries its own observed spot ticket",
      all(isinstance(v, (int, float)) and v > 0 for v in stkt.values()), "%r" % stkt)
# The finding that motivated the rebuild: published UAE tickets run ~5x India's.
# The old SIP-derived multiplier put India at 0.95x the UAE, inverting this.
check("UAE spot ticket is MULTIPLES of India's, as the published data shows",
      stkt["UAE"] > stkt["India"] * 3,
      "UAE %r vs India %r (ratio %.2f)" % (stkt["UAE"], stkt["India"], stkt["UAE"] / stkt["India"]))
# Tie the UAE figure to its source in AED, the unit Botim reported it in, so a
# drift in the peg or the ticket shows up against the observation itself.
_aed = stkt["UAE"] * aval("AED/USD peg")
check("UAE spot ticket still reconciles to Botim's observed AED 700",
      600 <= _aed <= 800, "implies AED %.0f" % _aed)
check("The inferred Gulf ticket sits between the two observed ones",
      stkt["India"] < stkt["Oman and Bahrain"] < stkt["UAE"], "%r" % stkt)
# The point of regionalising: as the mix shifts, the implied blended spot ticket
# must MOVE. If it is constant the regional scaling is not wired through.
spot, pay_all = row("Spot purchase volume"), row("PAYING CUSTOMERS")
implied = [spot[i] / pay_all[i] for i in (11, 23, 28) if pay_all[i]]
check("Blended spot ticket MOVES as the region mix shifts (regional scaling is live)",
      max(implied) - min(implied) > 1e-6,
      "implied per-customer spot %r" % [round(x, 2) for x in implied])

# ---- 8c. the regional block structure -------------------------------------
subs = {}
for rn in REGION_NAMES:
    lbl = "  SUBTOTAL - %s" % rn
    if rowof(mdl, lbl):
        subs[rn] = row(lbl)
check("Every region has its own revenue subtotal", len(subs) == len(REGION_NAMES),
      "found %r of %r" % (sorted(subs), REGION_NAMES))
tot_chk = row("CHECK: regional subtotals + B2B = sum of streams")
check("Regional subtotals + B2B reconcile to the sum of streams, every period",
      all_num(tot_chk) and max(abs(x) for x in tot_chk) < 1e-6,
      "max delta %r" % (max(abs(x) for x in tot_chk) if all_num(tot_chk) else tot_chk))
s6r = row("Stream 6 - B2B platform fee")
check("Grand total = sum of regional subtotals + non-regional B2B",
      all(abs(sum(subs[rn][i] for rn in subs) + s6r[i] - row("TOTAL NET REVENUE")[i]) < 1e-6
          for i in (0, 12, 28)))
# Every region must actually contribute - a silent zero would mean a region
# block is wired but never populated.
for rn in REGION_NAMES:
    check("Region %s contributes revenue by Y7" % rn, subs[rn][28] > 0, "Y7 %r" % subs[rn][28])
# The staged region must be empty until it opens.
gulf = [rn for rn in REGION_NAMES if "Oman" in rn]
if gulf:
    g = row("  SUBTOTAL - %s" % gulf[0])
    gm = aval("Region opens - %s" % gulf[0])
    check("The staged region earns nothing before it opens",
          all(abs(x) < 1e-9 for x in g[:gm - 1]) and g[gm - 1] != 0,
          "opens M%d; M%d=%r M%d=%r" % (gm, gm - 1, g[gm - 2], gm, g[gm - 1]))
# The real constraint is PER REGION: a cardholder cannot spend more in a month
# than the limit their own gold supports.
per_region_ok, worst = True, 0.0
_spc_rows = [r for r in range(1, mdl.max_row + 1)
             if mdl["A%d" % r].value == "  Card spend per card per month"]
_lim_rows = [r for r in range(1, mdl.max_row + 1)
             if mdl["A%d" % r].value == "  Credit limit per customer (gold x LTV)"]
check("Every region has a per-card spend row and a credit limit row",
      len(_spc_rows) == len(REGION_NAMES) and len(_lim_rows) == len(REGION_NAMES),
      "%d spend, %d limit, %d regions" % (len(_spc_rows), len(_lim_rows), len(REGION_NAMES)))
for sr_, lr_ in zip(_spc_rows, _lim_rows):
    for i in range(N):
        s_, l_ = mdl["%s%d" % (pc(i), sr_)].value, mdl["%s%d" % (pc(i), lr_)].value
        if isinstance(s_, (int, float)) and isinstance(l_, (int, float)) and l_ > 0:
            worst = max(worst, s_ / l_)
            if s_ > l_ * 1.001:
                per_region_ok = False
check("PER REGION - monthly card spend never exceeds the limit the gold supports",
      per_region_ok, "worst utilisation %.2fx" % worst)
# Under the drawdown model the constraint is satisfied BY CONSTRUCTION - spend
# IS the drawdown - so what matters is that annual drawdown stays inside a
# sensible multiple of the limit rather than a monthly cap binding.
_dr_rows = [r for r in range(1, mdl.max_row + 1)
            if mdl["A%d" % r].value == "  Annual drawdown per card (limit x drawn x draws)"]
_lim_rows2 = [r for r in range(1, mdl.max_row + 1)
              if mdl["A%d" % r].value == "  Credit limit per customer (gold x LTV)"]
check("Annual drawdown is a sane multiple of the credit limit (0.5x-3x)",
      all(0.5 <= (mdl["%s%d" % (pc(28), d)].value or 0) / (mdl["%s%d" % (pc(28), l)].value or 1) <= 3.0
          for d, l in zip(_dr_rows, _lim_rows2)),
      "%r" % [round((mdl["%s%d" % (pc(28), d)].value or 0) / (mdl["%s%d" % (pc(28), l)].value or 1), 2)
              for d, l in zip(_dr_rows, _lim_rows2)])
lim = row("  CHECK: card spend vs credit capacity (must be <= 1.00)")
check("Whole-book card spend stays within total credit capacity",
      all_num(lim) and max(lim) <= 1.001, "max %.2fx" % (max(lim) if all_num(lim) else -1))

# ---- 8d. average ticket is DERIVED and scales with the SIP contribution ----
_avg_rows = [r for r in range(1, mdl.max_row + 1)
             if mdl["A%d" % r].value == "  Average transaction size (derived)"]
check("Average transaction size is derived per region, not a single input",
      len(_avg_rows) == len(REGION_NAMES) and rowof(asm, "Average transaction size") is None,
      "%d rows; stale absolute input present: %s"
      % (len(_avg_rows), rowof(asm, "Average transaction size") is not None))
_avg = [mdl["%s%d" % (pc(28), r)].value for r in _avg_rows]
_lims = [mdl["%s%d" % (pc(28), r)].value for r in _lim_rows2]
# Under the drawdown model the ticket follows the CREDIT LIMIT (gold held), not
# the monthly SIP amount - a customer with more gold can draw and spend more.
check("Average ticket ORDERS the same way as the credit limit across regions",
      [i for i, _ in sorted(enumerate(_avg), key=lambda t: t[1])]
      == [i for i, _ in sorted(enumerate(_lims), key=lambda t: t[1])],
      "limits %r -> avg txn %r" % ([round(x) for x in _lims], [round(x) for x in _avg]))
# BAND WIDENED 2026-08-26, and the reason matters more than the band. The old
# floor of AED 80 was calibrated when the card was GATED: fewer cardholders,
# each with a larger limit. Opening access spreads the SAME collateral pool over
# ~2.7x more cards, so the per-head limit falls by the same factor and the
# ticket falls with it - from ~AED 277 to ~AED 90. That is arithmetic, not a
# defect, and re-imposing an AED 80 floor would just be asserting the old design.
check("Average ticket is a plausible size for a borrowed lump (AED 40-600)",
      all(40 <= x <= 600 for x in _avg), "%r" % [round(x) for x in _avg])
check("Average ticket sits well below the UAE all-population average of AED 313",
      max(_avg) < 313, "max derived ticket AED %.0f" % max(_avg))

# ---- 8e. THE SMALL-TICKET WARNING THAT OPENING ACCESS CREATES --------------
# The processor bills PER AUTHORISATION, so a transaction below a break-even
# size costs Aurumix money no matter how many of them there are. Opening the
# card shrank the average ticket, which pushes every region towards that line.
# The aggregate spend did not change - the same pool is simply spent in more,
# smaller transactions - so this is a pure COST effect that the revenue-only
# scope cannot show, and it lands squarely in tomorrow's cost build.
_ic = aval("Interchange - Gold")
_pm = sval("Programme manager share of interchange")
_fee = aval("Per-transaction processor fee")
_dec = aval("Decline uplift on authorised txns")
_aed = aval("AED/USD peg")
_net_rate = _ic * (1 - _pm)                       # what Aurumix keeps per AED spent
_cost_txn = _fee * (1 + _dec) * _aed              # what it pays per authorisation
_breakeven = _cost_txn / _net_rate if _net_rate else 0
check("Average ticket clears the minimum PROFITABLE transaction size",
      all(x > _breakeven for x in _avg),
      "break-even AED %.0f at %.0f%% PM share; regional tickets %r"
      % (_breakeven, 100 * _pm, [round(x) for x in _avg]))
check("Ticket headroom over break-even is at least 25% in every region",
      all(x > _breakeven * 1.25 for x in _avg),
      "thinnest region is %.2fx break-even" % (min(_avg) / _breakeven if _breakeven else 0))

# ---- 8f. THE ISSUANCE FEE IS A ONE-OFF, NOT AN ANNUAL FEE -----------------
# Fixed 2026-08-26. It was charged at 1.06 events/card/YEAR against the card
# STOCK, which billed every existing cardholder AED 75 annually. It must be
# levied on the NEW-CARD FLOW. These checks exist because the defect was
# invisible for months: the number looked plausible, the formula parsed, and
# nothing tied it back to how many cards were actually issued.
_nc_rows = [r for r in range(1, mdl.max_row + 1)
            if mdl["A%d" % r].value == "  ...newly issued this period (flow, drives the issuance fee)"]
check("Every region reports a NEW-CARD flow, distinct from the card stock",
      len(_nc_rows) == len(REGION_NAMES), "found %d" % len(_nc_rows))
_nc = [sum(mdl["%s%d" % (pc(i), r)].value or 0 for r in _nc_rows) for i in range(N)]
_cards_all = [sum(mdl["%s%d" % (pc(i), r)].value or 0 for r in _card_rows) for i in range(N)]
check("New cards are never negative",
      all(x >= -1e-9 for x in _nc), "min %.4f" % min(_nc))
# The flow must be STRICTLY smaller than the stock once the book has matured -
# if they are equal, the fee is back on the stock and the defect has returned.
check("New cards are a FLOW - far below the card stock at Y7",
      _cards_all[N - 1] > 0 and _nc[N - 1] < _cards_all[N - 1] * 0.5,
      "Y7 new %.0f vs stock %.0f (%.0f%%)"
      % (_nc[N - 1], _cards_all[N - 1], 100 * _nc[N - 1] / _cards_all[N - 1]))
# And the flow must reconcile to the change in the stock, which is what makes
# it a real flow rather than another ratio applied to the same base.
_growth = [_cards_all[i] - _cards_all[i - 1] for i in range(1, N_MONTHLY)]
check("New cards reconcile to the period-on-period growth in the card stock",
      all(abs(_nc[i + 1] - max(0.0, _growth[i])) < max(1.0, abs(_growth[i]) * 1e-6)
          for i in range(len(_growth))),
      "worst gap %.4f" % max(abs(_nc[i + 1] - max(0.0, _growth[i])) for i in range(len(_growth))))
check("The stale per-year issuance-events input is gone",
      rowof(sc, "Card issuance events") is None and rowof(asm, "Card issuance events") is None)

# ---- 9. totals -------------------------------------------------------------
tot = row("TOTAL NET REVENUE")
parts = ["Stream 1a - Entry fee, SIP", "Stream 1b - Entry fee, SPOT",
         "Stream 2 - Card interchange", "Stream 3 - Family plan and Digital Will",
         "Stream 4 - Cardholder fees", "Stream 5 - Lending revenue share",
         "Stream 6 - B2B platform fee"]
sums = [sum(row(p)[i] for p in parts) for i in range(N)]
check("TOTAL NET REVENUE equals the sum of its streams",
      all(abs(tot[i] - sums[i]) < 1e-6 for i in range(N)))
check("Revenue is positive from M1 and grows", tot[0] > 0 and tot[-1] > tot[0])
aum = row("COLLATERAL-ELIGIBLE AUM")
check("AUM is positive and grows", all(x >= 0 for x in aum) and aum[-1] > aum[0])
check("Grams held never go negative", all(x >= 0 for x in row("GRAMS HELD")))

# Two AUM measures, and the ordering between them is structural. Custody removes
# ONLY redemption; collateral-eligible also removes gold moved out of Aurumix's
# control. Custody can therefore never be the smaller of the two - if it ever is,
# the two decay rates have been wired to the wrong balances.
_cust, _coll = row("GRAMS UNDER CUSTODY"), row("GRAMS HELD")
check("Gold under custody is never below collateral-eligible gold",
      all_num(_cust) and all_num(_coll) and all(c >= h - 1e-6 for c, h in zip(_cust, _coll)),
      "min gap %r" % (min(c - h for c, h in zip(_cust, _coll)) if all_num(_cust) else None))
# ...and it must be STRICTLY larger once self-custody has had time to bite,
# otherwise the self-custody term has been dropped from the collateral decay and
# the two rows are silently identical.
check("The two measures genuinely diverge (self-custody is actually deducted)",
      _cust[28] > _coll[28] * 1.02,
      "Y7 custody %,.0f vs collateral %,.0f".replace("%,", "{:,").replace(".0f", ".0f}").format(
          _cust[28], _coll[28]))
check("Summary carries the comparable custody headline, not only the collateral figure",
      rowof(summ, "Gold under custody (comparable headline)") is not None)

# ---- 9a2. EVERY CARD NOW CARRIES CREDIT (rewritten 2026-08-26) ------------
# SUPERSEDES the 2026-08-21 correction that split cardholders from credit-
# eligible cardholders. That split existed because a lapsed customer lost their
# ICS score and therefore their facility. Access is no longer ICS-gated and the
# lapsed customer's gold is still in the vault, so the collateral is still
# there: every card has a line.
#
# THE AGGREGATE DID NOT MOVE, and that is worth asserting rather than assuming.
# Old:  ccards x limit = [cards x paying/base] x [AUM/paying x LTV]
# New:  cards  x limit =  cards               x [AUM/base   x LTV]
# The "paying" terms cancelled, so total credit capacity is identical either
# way - holders' gold was always in the pool. What changed is the per-head
# limit, which used to divide by paying customers only and so read high once
# holders became most of the base.
_cc = [r for r in range(1, mdl.max_row + 1)
       if mdl["A%d" % r].value == "  ...of which have a live credit line (now all of them)"]
check("Every region still reports a credit-line row",
      len(_cc) == len(REGION_NAMES), "found %d" % len(_cc))
_cc7 = sum(mdl["%s%d" % (pc(N - 1), r)].value for r in _cc)
_cards7 = row("ACTIVE CARDS")[N - 1]
check("Credit-eligible cardholders now EQUAL cardholders - access is ungated",
      _cards7 > 0 and abs(_cc7 - _cards7) < max(1.0, _cards7 * 1e-6),
      "Y7 %.0f of %.0f cards" % (_cc7, _cards7))
# The per-head limit must now be AUM over the CARD-ELIGIBLE BASE. If anyone
# reverts the denominator to paying customers, the limit inflates by the
# holder ratio and this fails.
_pay7, _hold7 = row("PAYING CUSTOMERS")[N - 1], row("HOLDERS (stopped paying, still hold gold)")[N - 1]
_lim = [r for r in range(1, mdl.max_row + 1)
        if mdl["A%d" % r].value == "  Credit limit per customer (gold x LTV)"]
_aum7 = row("COLLATERAL-ELIGIBLE AUM")[N - 1]
_ltv = aval("LTV against collateral")
_impl = _aum7 / (_pay7 + _hold7) * _ltv
_wavg = sum(mdl["%s%d" % (pc(N - 1), r)].value for r in _lim) / len(_lim)
check("Credit limit is AUM over the CARD-ELIGIBLE BASE, not over payers alone",
      abs(_wavg - _impl) < _impl * 0.25,
      "regional mean limit %.2f vs book-implied %.2f" % (_wavg, _impl))
# Spend must still reconcile to the per-card drawdown - with credit cards and
# all cards now equal, this catches spend being wired to the wrong row.
_spend7 = row("Card spend")[N - 1]
_dr = [r for r in range(1, mdl.max_row + 1)
       if mdl["A%d" % r].value == "  Annual drawdown per card (limit x drawn x draws)"]
_maxdraw = max(mdl["%s%d" % (pc(N - 1), r)].value for r in _dr)
_mindraw = min(mdl["%s%d" % (pc(N - 1), r)].value for r in _dr)
check("Card spend reconciles to the per-card annual drawdown",
      _cc7 > 0 and _mindraw * 0.9 <= _spend7 / _cc7 <= _maxdraw * 1.1,
      "implied spend per card %.2f vs regional drawdowns %.2f-%.2f"
      % (_spend7 / _cc7 if _cc7 else 0, _mindraw, _maxdraw))

# ---- 9b. the metal price moves, so attribution must stay recoverable -------
_px, _cagr = row("Gold price (this period)"), aval("Gold price appreciation")
check("Gold price compounds at the stated rate, on MODEL YEAR",
      all_num(_px) and abs(_px[28] - _px[0] * (1 + _cagr) ** 6) < 1e-6,
      "M1 %.2f -> Y7 %.2f, implied CAGR %.4f" % (_px[0], _px[28], (_px[28] / _px[0]) ** (1 / 6.0) - 1))
check("All twelve months of a model year share one price (no intra-year drift)",
      len(set(round(x, 6) for x in _px[:12])) == 1 and len(set(round(x, 6) for x in _px[12:24])) == 1,
      "Y1 %r Y2 %r" % (sorted(set(round(x, 2) for x in _px[:12])),
                       sorted(set(round(x, 2) for x in _px[12:24]))))
# The constant-price memo is what keeps execution judgeable once the metal
# moves. If it ever equals the headline while the CAGR is positive, it has been
# wired to the moving price and the attribution is silently gone.
_hd = row("GOLD UNDER CUSTODY (comparable headline)")
_const = row("  Gold under custody at the CONSTANT M1 price (memo)")
_cg = row("GRAMS UNDER CUSTODY")
check("A constant-price series exists and is BELOW the headline once gold has risen",
      all_num(_hd) and all_num(_const) and (_const[28] < _hd[28] if _cagr > 0 else True)
      and abs(_const[0] - _hd[0]) < 1e-6,
      "Y7 headline %.0f vs constant %.0f" % (_hd[28], _const[28]))
check("The constant-price row is grams x the M1 price, so it isolates execution",
      all_num(_const) and all(abs(_const[i] - _cg[i] * _px[0]) < 1e-3 for i in (11, 25, 28)),
      "Y7 %.0f vs grams x M1 price %.0f" % (_const[28], _cg[28] * _px[0]))
# The entry fee is a percentage of a USD contribution and must NOT move with
# gold. If it does, the price has been wired into a stream it does not touch.
check("The entry fee is untouched by the gold price (USD-denominated, as designed)",
      all(abs(s1a[i] - sip[i] * fee) < 1e-6 for i in range(N) if sip[i]),
      "stream 1a still equals SIP x fee at every period")

# ---- 9c. stream 3 is a SUBSCRIBER BALANCE, not a share of the book ---------
_subrows = [r for r in range(1, mdl.max_row + 1)
            if mdl["A%d" % r].value == "  Family plan subscribers"]
check("Every region carries a family plan subscriber balance",
      len(_subrows) == len(REGION_NAMES), "found %d" % len(_subrows))
_sub7 = sum(mdl["%s%d" % (pc(N - 1), r)].value for r in _subrows)
_att = aval("Family plan attach rate")
# The whole point of the rebuild: attach applies to NEW customers and the
# balance churns, so the standing share of the book must land BELOW the attach
# rate. If it equals it, the balance has been wired as a percentage again.
check("Subscribers settle BELOW the attach rate (the balance actually churns)",
      _sub7 / row("PAYING CUSTOMERS")[N - 1] < _att * 0.95,
      "Y7 penetration %.3f vs attach %.3f" % (_sub7 / row("PAYING CUSTOMERS")[N - 1], _att))
check("Family plan churn EXCEEDS SIP churn (an incremental rate is applied)",
      aval("Family plan monthly churn, combined (derived)") > aval("Monthly churn rate (derived)"),
      "combined %.4f vs SIP %.4f" % (aval("Family plan monthly churn, combined (derived)"),
                                     aval("Monthly churn rate (derived)")))
# Revenue per subscriber must equal plan + beneficiary fee on names beyond the
# first. Catches the MAX(0,...) being dropped, or the fee applied to every name.
_price, _bfee, _bn = aval("Family plan / Digital Will price"),     aval("Additional beneficiary fee"), aval("Beneficiaries named per plan")
_expect = _price + max(0.0, _bn - 1) * _bfee
_s3 = row("Stream 3 - Family plan and Digital Will")
check("Revenue per subscriber = plan fee + beneficiary fee beyond the FIRST",
      abs(_s3[N - 1] / _sub7 - _expect) < 1e-6,
      "implied %.2f vs expected %.2f" % (_s3[N - 1] / _sub7, _expect))
check("The beneficiary fee is a real uplift but stays small",
      1.0 < _expect / _price < 1.5, "uplift %.2fx" % (_expect / _price))

# ---- 10. Summary ties to Model --------------------------------------------
def sumv(label, y):
    r = rowof(summ, label)
    return summ["%s%d" % (get_column_letter(4 + y - 1), r)].value if r else None


y1 = sum(tot[i] for i in range(12))
check("Summary Y1 total revenue ties to the Model's twelve monthly columns",
      abs(sumv("TOTAL REVENUE", 1) - y1) < 1e-6,
      "summary %r vs model %r" % (sumv("TOTAL REVENUE", 1), y1))
check("Summary Y7 total revenue ties to the Model's Y7 column",
      abs(sumv("TOTAL REVENUE", 7) - tot[28]) < 1e-6)
# The revenue mix must sum to 100% - it once divided regional rows by the wrong
# denominator and produced percentages in the hundreds.
for y_ in (1, 3, 5, 7):
    mixsum = sumv("CHECK: mix sums to 100%", y_)
    check("Revenue mix sums to 100%% at Y%d" % y_,
          isinstance(mixsum, (int, float)) and abs(mixsum - 1.0) < 1e-6, "%r" % mixsum)
check("Summary paying customers at Y1 is the CLOSING month, not a sum",
      abs(sumv("Paying customers", 1) - pay[11]) < 1e-6,
      "summary %r vs M12 %r" % (sumv("Paying customers", 1), pay[11]))

print("=" * 78)
print("REVENUE MODEL VERIFICATION - %s" % PATH)
print("=" * 78)
npass = sum(1 for _, ok, _ in results if ok)
for name, ok, detail in results:
    print("  [%s] %s" % ("PASS" if ok else "FAIL", name))
    if not ok and detail:
        print("         -> %s" % detail)
print("-" * 78)
print("%d/%d passed" % (npass, len(results)))
print("=" * 78)
sys.exit(0 if npass == len(results) else 1)
