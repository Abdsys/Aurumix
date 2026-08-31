"""
Aurumix Revenue Model - Excel Workbook Builder
===============================================

Built from : deliverables/4-revenue-modeling/Aurumix_Revenue_Model_Architecture_Brief.md
Standard   : DRODE_Revenue_Model_V3.xlsx (the firm's reference build)

FIVE SHEETS: Cover | Assumptions | Scenario Parameters | Model | Summary

    Assumptions  A=Parameter  B=Value(BLUE)  C=Unit  D=Source
    Scenario     A=Parameter  B=Active  C=Base  D=Aggressive  E=Conservative
    Model        R2=Period #  R3=Period,  periods in C..AE (29)
    Summary      year columns D..J (Y1..Y7)

Colours: BLUE = input, BLACK = in-sheet formula, GREEN = cross-sheet reference.

Periods: 29 = 24 monthly (M1-M24) + 5 annual (Y3-Y7). M1 = January 2027.

-----------------------------------------------------------------------------
SIMPLIFICATION AGREED 2026-08-20. This is a REVENUE model.
-----------------------------------------------------------------------------
DELETED (moved to the Phase 5 simulation, where heterogeneity belongs):
  - five payment archetypes and their weights / pay probabilities / hazards
  - the run-of-6 first-passage gate chain
  - the 84-month lifecycle curves and the convolution
  - the six-state machine, the REDUCED state, the withdrawal-bucket distribution

REPLACED BY a rolling balance the client can read:
    opening + new - churned = closing

KEPT DELIBERATELY, because dropping each changes the answer by more than the
simplification saves:
  1. ELIGIBILITY as two input cells - "% who ever qualify" and "months to
     qualify". The card streams are ~83% of gross profit and all hang off the
     six-payment gate; assuming everyone qualifies overstates by ~59%. The
     archetype engine's OUTPUT (55%, M8) becomes these two inputs.
  2. A HOLDERS balance. Customers who stop paying KEEP THEIR GOLD - they stay
     in AUM, custody, collateral and the B2B base, and reach ~81% of everyone
     ever acquired by Y6. Without it AUM and streams 5/6 are badly understated.
  3. The FABRICATION PREMIUM inside stream 1. The entry fee is a MARGIN, not a
     fee: 5% charged less ~1.5% to buy the metal. Gross would read ~43% high.
     This is cost OF REVENUE, not opex.
  4. The ATM draw distribution (4 rows). The mean draw sits just below the free
     allowance BY DESIGN, so a mean returns EXACTLY ZERO over-allowance revenue.

OUT OF SCOPE for now, to be bolted on later: opex, headcount, tax, working
capital, cash, funding, break-even. The model reports revenue and gross margin.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

HERE = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(HERE, "Aurumix_Revenue_Model.xlsx")

# ============================================================================
# BRAND
# ============================================================================
WARM_CHARCOAL, GOLD, DARK_GOLD, STONE, STONE_LIGHT = "1A1714", "B8956E", "96734A", "D4CFC8", "E8E4DF"
BLUE, BLUE_BOLD = Font(color="0000FF"), Font(color="0000FF", bold=True)
BLACK, BLACK_BOLD = Font(color="000000"), Font(color="000000", bold=True)
GREEN, GREEN_BOLD = Font(color="009B73"), Font(color="009B73", bold=True)
RED_BOLD = Font(color="C00000", bold=True)

SECTION_FILL = PatternFill("solid", fgColor=WARM_CHARCOAL)
HEADER_FILL = PatternFill("solid", fgColor=STONE)
BANNER_FILL = PatternFill("solid", fgColor=STONE_LIGHT)

COVER_TITLE = Font(size=24, bold=True, color=GOLD)
COVER_SUB = Font(size=16, color=WARM_CHARCOAL)
SHEET_TITLE = Font(size=18, bold=True, color=GOLD)
SHEET_SUB = Font(size=14, bold=True, color=WARM_CHARCOAL)
SECTION_FONT = Font(size=11, bold=True, color="FFFFFF")
BANNER_FONT = Font(size=11, bold=True, color=WARM_CHARCOAL)
HEADER_FONT = Font(size=10, bold=True, color=WARM_CHARCOAL)
SECONDARY = Font(size=9, color=DARK_GOLD)
NOTE_FONT = Font(size=9, italic=True, color=DARK_GOLD)

STONE_THIN = Side(style="thin", color=STONE)
BORDER = Border(top=STONE_THIN, bottom=STONE_THIN, left=STONE_THIN, right=STONE_THIN)
TOTALS_BORDER = Border(top=Side(style="thin", color=GOLD), bottom=Side(style="double", color=GOLD))

FMT_USD, FMT_USD2 = '$#,##0;($#,##0);"-"', '$#,##0.00;($#,##0.00);"-"'
FMT_NUM, FMT_NUM2, FMT_NUM3 = '#,##0;(#,##0);"-"', '#,##0.00;(#,##0.00);"-"', '#,##0.000'
FMT_PCT, FMT_PCT2, FMT_PCT3 = '0.0%', '0.00%', '0.000%'
UNFILLED = "{{UNFILLED}}"

# ============================================================================
# PERIOD GRID
# ============================================================================
N_MONTHLY, N_ANNUAL = 24, 5
N_PERIODS = N_MONTHLY + N_ANNUAL
PCOL0, START_YEAR = 3, 2027
SUMCOL0, N_YEARS = 4, 7
REGIONS = ["uae", "gulf", "india"]
RLAB = {"uae": "UAE", "gulf": "Oman and Bahrain", "india": "India"}
RDESC = {
    "uae": "UAE residents - Indian plus other South Asian (Pakistani, Bangladeshi, Sri Lankan, Nepali). "
           "COMBINED 2026-08-20 from two separate regions: both sit under one licence, one set of rails and "
           "one agent channel, both open at M1, and both are UAE-resident for VAT. They differed only on "
           "average ticket (USD 38 vs USD 26), which the blended USD 33.60 preserves at the book level. The "
           "core market - the gold-savings habit and the agent network originate here.",
    "gulf": "Expatriate workers in Oman and Bahrain. ~600,000 addressable. NOT UAE-resident, so VAT and "
            "reporting treatment differ. Requires local authorisation in a second and third jurisdiction, "
            "which is why it is the only staged region.",
    "india": "India-resident retail investors. ~12.5m addressable, sized BEHAVIOURALLY from gold ETF folios "
             "intersected with active digital-gold holders rather than demographically. The payment route is "
             "ASSUMED SOLVED in this model - if it is not, this region does not open at all.",
}


def pcol(i):  return get_column_letter(PCOL0 + i)
def pcell(r, i): return "%s%d" % (pcol(i), r)
def ycol(y):  return get_column_letter(SUMCOL0 + y - 1)
def is_monthly(i): return i < N_MONTHLY
def plabel(i): return "M%d" % (i + 1) if is_monthly(i) else "Y%d" % (i - N_MONTHLY + 3)
def pyear(i): return (i // 12) + 1 if is_monthly(i) else (i - N_MONTHLY + 3)
def cal_month(i): return (i % 12) + 1 if is_monthly(i) else 0
def n_months(i): return 1 if is_monthly(i) else 12
def is_fy_end(i): return (not is_monthly(i)) or ((i + 1) % 12 == 0)


MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
NAMED = []


def declare(nm, sheet, addr):
    NAMED.append((nm, "'%s'!%s" % (sheet, addr) if " " in sheet else "%s!%s" % (sheet, addr)))


def section(ws, row, text, span=8):
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    cell = ws.cell(row=row, column=1)
    cell.value, cell.font = text, SECTION_FONT
    ws.row_dimensions[row].height = 18
    return row


def banner(ws, row, text, span=31):
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = BANNER_FILL
    cell = ws.cell(row=row, column=1)
    cell.value, cell.font = text, BANNER_FONT
    return row


def divider(ws, row, text, span=31):
    """A HARD visual break, distinct from banner().

    banner() is a pale sub-heading INSIDE a band. This is the charcoal bar that
    separates one side of the P&L from the other, so a reader scrolling the
    Model sheet can never be in doubt whether the row under the cursor is
    something the business earns or something it pays.
    """
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    cell = ws.cell(row=row, column=1)
    cell.value, cell.font = text, SECTION_FONT
    ws.row_dimensions[row].height = 20
    return row


def headers(ws, row, cols):
    for col, h in cols:
        c = ws["%s%d" % (col, row)]
        c.value, c.fill, c.font, c.border = h, HEADER_FILL, HEADER_FONT, BORDER
    return row


def note(ws, row, text, col="A"):
    c = ws["%s%d" % (col, row)]
    c.value, c.font = text, NOTE_FONT


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


# ============================================================================
wb = Workbook()
ws_cover = wb.active
ws_cover.title = "Cover"
ws_assum = wb.create_sheet("Assumptions")
ws_scen = wb.create_sheet("Scenario Parameters")
ws_model = wb.create_sheet("Model")
ws_summ = wb.create_sheet("Summary")
for s in wb.worksheets:
    s.sheet_view.showGridLines = False

# ============================================================================
# COVER
# ============================================================================
ws_cover["B2"] = "Aurumix Revenue Model"
ws_cover["B2"].font = COVER_TITLE
ws_cover["B3"] = "Hybrid Monthly + Annual View  |  7-Year Projection"
ws_cover["B3"].font = COVER_SUB
for r, (lbl, val) in enumerate([
    ("Date", "=TODAY()"),
    ("Scope", "REVENUE ONLY. Six revenue streams, reported net of cost of revenue. Operating costs, tax, "
              "cash and funding are added in a later build."),
    ("Periods", "24 monthly (M1-M24) + 5 annual (Y3-Y7) = 29 total.  M1 = January %d" % START_YEAR),
    ("Engine", "Rolling customer balance: opening + new - churned = closing, by region. Customers who stop "
               "paying keep their gold and move to a HOLDERS balance."),
    ("Metal", "100 g bars throughout. One entry fee, one fabrication premium."),
    ("Prepared by", "Tokenomics.net: Custom Data Room Engagement"),
], start=5):
    ws_cover["B%d" % r] = lbl
    ws_cover["B%d" % r].font = BLACK_BOLD
    ws_cover["C%d" % r] = val
ws_cover["C5"].number_format = "mmmm yyyy"

ws_cover["B12"] = "COLOR LEGEND"
ws_cover["B12"].font = BLACK_BOLD
for r, (t, d, f) in enumerate([("BLUE text", "Hardcoded inputs (Assumptions & Scenario Parameters)", BLUE),
                               ("BLACK text", "In-sheet formulas", BLACK),
                               ("GREEN text", "Cross-sheet references", GREEN)], start=13):
    ws_cover["B%d" % r] = t
    ws_cover["B%d" % r].font = f
    ws_cover["C%d" % r] = d

ws_cover["B17"] = "SHEETS"
ws_cover["B17"].font = BLACK_BOLD
for r, t in enumerate([
    "Cover: This page",
    "Assumptions: Fixed inputs, activation calendar, seasonality, market ceilings",
    "Scenario Parameters: Base / Aggressive / Conservative, plus structural switches",
    "Model: Customer engine, AUM and the six revenue streams (29 periods)",
    "Summary: Annual revenue by stream (Y1-Y7), revenue mix, key metrics",
], start=18):
    ws_cover["B%d" % r] = t

ws_cover["B24"] = "SIMPLIFIED BUILD - cohort and archetype detail moves to the Phase 5 simulation."
ws_cover["B24"].font = RED_BOLD
ws_cover["B25"] = ("Kept as inputs rather than engines: card eligibility (% who qualify, months to qualify), "
                   "the holders balance, and the fabrication premium inside stream 1.")
ws_cover["B25"].font = NOTE_FONT
widths(ws_cover, {"A": 3, "B": 30, "C": 104})

# ============================================================================
# ASSUMPTIONS
# ============================================================================
ws_assum["A1"] = "=Cover!$B$2"
ws_assum["A1"].font = SHEET_TITLE
ws_assum["A2"] = "Inputs and Assumptions"
ws_assum["A2"].font = SHEET_SUB
widths(ws_assum, {"A": 54, "B": 16, "C": 24, "D": 120})
AROW, _ar = {}, [4]


def a_section(t):
    r = _ar[0]
    section(ws_assum, r, t, span=4)
    headers(ws_assum, r + 1, [("A", "Parameter"), ("B", "Value"), ("C", "Unit"), ("D", "Source")])
    _ar[0] = r + 2
    return r


def a_row(key, param, value, unit, source, fmt=FMT_NUM2, name=None):
    r = _ar[0]
    ws_assum["A%d" % r] = param
    c = ws_assum["B%d" % r]
    if value is None:
        c.value, c.font = UNFILLED, RED_BOLD
    else:
        c.value, c.font = value, BLUE
        if fmt and not isinstance(value, str):
            c.number_format = fmt
    ws_assum["C%d" % r] = unit
    ws_assum["D%d" % r] = source
    AROW[key] = r
    if name:
        declare(name, "Assumptions", "$B$%d" % r)
    _ar[0] = r + 1
    return r


def aref(k): return "Assumptions!$B$%d" % AROW[k]


a_section("PRICING AND METAL")
a_row("gold_price", "Gold price at M1", 141.46, "USD/g",
      "USD 4,400/oz verified 2026-08-17. THE STARTING PRICE - it no longer holds for the whole horizon; see "
      "the appreciation rate below and the live 'Gold price (this period)' row on the Model. CITED.",
      FMT_NUM2, "gold_price")
a_row("aed_usd", "AED/USD peg", 3.6725, "AED/USD", "CBUAE peg. CITED.", FMT_NUM2, "aed_usd")
a_row("entry_fee", "Entry fee charged", 0.050, "% of contribution",
      "Decision 9, client-stated. A SINGLE RATE throughout. CLIENT INPUT.", FMT_PCT2, "entry_fee")
a_row("fab_premium", "Fabrication premium paid", 0.01500, "% over spot",
      "100 g bars. OBSERVED (D28): goldtrade.ae 19 Aug 2026 measured 100 g at +1.71% (PAMP 1.75, Valcambi "
      "1.67), less the published 25 bp bulk gradient. THE ENTRY FEE IS A MARGIN, NOT A FEE - this is netted "
      "off stream 1, because it is cost OF REVENUE. Reporting stream 1 gross would read ~43% high.",
      FMT_PCT2, "fab_premium")
a_row("bar_grams", "Bar denomination", 100, "grams", "100 g throughout, by agreement. OBSERVED.",
      FMT_NUM, "bar_grams")
_ar[0] += 1

a_section("CARD ECONOMICS")
a_row("ic_gold", "Interchange - Gold", 0.01800, "% of spend",
      "Visa UAE IRF schedule, 18 Oct 2025. CITED, PRIMARY. The live model reads the FLAT GOLD RATE: replacing "
      "the full 1.80/2.05/2.10 ladder with it moves stream 2 by only ~2% of gross profit at Y7, because "
      "Sovereign is 1.2% of tiered accounts.", FMT_PCT2, "ic_gold")
a_row("txn_fee", "Per-transaction processor fee", 0.10, "USD/authorised txn",
      "WHAT THE PROCESSOR CHARGES AURUMIX to authorise and switch each transaction - NOT passed to the "
      "cardholder, and it cannot be: no consumer card charges per swipe on normal purchases, the card here is "
      "a REWARD for saving rather than a product sold, and a declined authorisation cannot be billed to "
      "anyone. Issuing economics are structurally: earn interchange, pay scheme and processing costs out of "
      "it. USD 0.10 is Stripe Issuing's PUBLISHED RACK RATE, used as a proxy because NO UAE PROCESSOR "
      "PUBLISHES PRICING. At volume, processors commonly reach USD 0.02-0.05 - a term-sheet negotiation, and "
      "at USD 0.03 the minimum profitable transaction falls from ~AED 77 to ~AED 23. CITED as a proxy.",
      FMT_USD2, "txn_fee")
a_row("decline_uplift", "Decline uplift on authorised txns", 0.06, "% uplift",
      "The processor fee is charged per AUTHORISATION, not per settled transaction, so declined attempts are "
      "billable to Aurumix and chargeable to nobody. ASSUMPTION.", FMT_PCT, "decline_uplift")
a_row("fx_margin", "FX margin on foreign spend", 0.020, "% of foreign spend",
      "Market rate ~2%, converged across four comparables. CITED.", FMT_PCT2, "fx_margin")
a_row("atm_allowance", "Free ATM allowance (Gold)", 1000, "AED/month",
      "Monthly and NOT rolling. Set deliberately ABOVE the median draw - that is the finding. CITED.",
      FMT_NUM, "atm_allowance")
a_row("atm_fee", "ATM fee over the allowance", 0.020, "% of the excess",
      "Sector converged. CITED.", FMT_PCT2, "atm_fee")
a_row("issuance_fee", "Card issuance fee", 75, "AED one-off",
      "Charged at base level, waived at upper tiers. CITED structure, ASSUMPTION rate.", FMT_NUM, "issuance_fee")
a_row("replacement_fee", "Card replacement fee", 100, "AED per event",
      "Market-normal UAE replacement is AED 75-150. ASSUMPTION.", FMT_NUM, "replacement_fee")
_ar[0] += 1

a_section("OTHER STREAM PRICING")
a_row("family_price", "Family plan / Digital Will price", 50, "USD/yr",
      "RE-PRICED 2026-08-21 from USD 120/yr. THE OLD PRICE WAS ABOVE TRUST & WILL'S PREMIUM US MEMBERSHIP "
      "(USD 49/yr) AND WAS 33% OF EVERYTHING THE CUSTOMER SAVES IN A YEAR - a USD 30/month saver puts away "
      "USD 360, and we were charging 120 of it. "
      "THE INDUSTRY MODEL IS A ONE-OFF FEE PLUS A SMALL ANNUAL UPDATE LAYER, and it is the annual layer this "
      "row represents: Trust & Will USD 199 one-off + USD 49/yr; Farewill GBP 100 + GBP 10/yr; Epilogue "
      "C$139 with NO subscription at all; other platforms renew updates at ~C$4.50/yr. Formal UAE will "
      "registration is a different product entirely and one-off - DIFC AED 10,000 single / 15,000 mirror, "
      "Abu Dhabi Civil Court AED 950 as the cheapest formal route. Indian online wills run INR 500-4,000 "
      "basic. Family add-ons on finance apps are usually BUNDLED FREE; where charged it is USD 5-15/month "
      "for a WHOLE HOUSEHOLD (Monarch 99.99/yr, Zeta 59.99/yr, Greenlight 5.99/mo). "
      "SET AT USD 50/yr, client instruction 2026-08-21. That is essentially TRUST & WILL'S USD 49/yr "
      "MEMBERSHIP - a defensible anchor in that it matches the best-known product in the category, but "
      "NOTE WHOSE PRODUCT IT IS: Trust & Will sells to US customers who paid 199-599 up front for an "
      "estate plan, not to a migrant worker saving USD 30 a month. At 50 the plan is 14% OF EVERYTHING "
      "THE CUSTOMER SAVES IN A YEAR, against 10% at 36 and 33% at the original 120. It is no longer "
      "absurd, but it is priced at a developed-market benchmark for an emerging-market customer, and "
      "that is the line an adviser would push on. A ONE-OFF SETUP FEE IS NOT MODELLED - if the client wants the full "
      "industry structure, that is a second row, not a bigger number here. CITED comparators, ASSUMPTION on "
      "the point in the range.", FMT_USD, "family_price")
a_row("benef_fee", "Additional beneficiary fee", 6, "USD/yr each",
      "ADDED 2026-08-21, client instruction. Charged per named beneficiary BEYOND THE FIRST, which the plan "
      "includes - the standard shape for this kind of add-on, and the reason the formula carries a MAX(0, "
      "n-1) rather than charging every beneficiary. "
      "NO COMPARABLE EXISTS: will services price PER WILL, not per beneficiary. DIFC charges more for a "
      "MIRROR will (two people) and more again for guardianship, but nobody publishes a per-beneficiary "
      "tariff. So this is priced on COST RECOVERY, which is the defensible basis available: every named "
      "beneficiary needs identity capture and AML screening, and Sumsub screening already sits in this "
      "model at ~USD 1.85 per check in the redemption cost memo. USD 6/yr covers that with margin without "
      "becoming a second subscription. "
      "DELIBERATELY SMALL. At 1.5 chargeable beneficiaries it adds USD 9/yr to a USD 50 plan - an 18% uplift "
      "on a stream worth ~3.5% of revenue, so it cannot move the model and should not be tuned to try. "
      "ASSUMPTION on a cost-recovery basis.", FMT_USD, "benef_fee")
# B2B platform fee MOVED TO SCENARIO PARAMETERS 2026-08-21 - see Group E.
a_row("origination_gross", "Credit origination fee - gross", 0.0100, "% of draw",
      "Finance House UAE gold loan Key Facts Statement carries a 1% processing fee - the one gross rate with a "
      "real UAE anchor. CITED.", FMT_PCT2, "origination_gross")
a_row("origination_share", "Credit origination - Aurumix share", 0.50, "% of gross", "ASSUMPTION.",
      FMT_PCT, "origination_share")
a_row("servicing_gross", "Credit servicing fee - gross", 0.0050, "%/yr of drawn",
      "ASSUMPTION.", FMT_PCT2, "servicing_gross")
a_row("servicing_share", "Credit servicing - Aurumix share", 0.70, "% of gross",
      "The highest of the heads: servicing is where Aurumix does the actual work. ASSUMPTION.",
      FMT_PCT, "servicing_share")
a_row("ltv_gold", "LTV against collateral", 0.50, "% of collateral",
      "_draft_ics-scoring.md sec 6.2, settled 2026-08-13. CITED.", FMT_PCT, "ltv_gold")
# ---- Spot ticket: OBSERVED per region, replacing a derived multiplier -------
# REBUILT 2026-08-21. Until now the regional spot ticket was a multiplier
# derived from the regional SIP ticket, on the reasoning that both proxy the
# same person's income. That was internally consistent and EXTERNALLY WRONG.
#
# SIP tickets barely differ across regions (33.60 / 26.00 / 30.00 - a 12%
# spread), so the derivation compressed the real gap into 12% AND POINTED IT
# THE WRONG WAY: it put India at 0.95x the UAE, where the published data puts
# it near 0.20x. Deriving a number from a related one imports that number's
# flatness. These are now OBSERVED per region.
#
# EACH IS THE SIZE OF ONE PURCHASE, not an annual sum. Annual spend per
# buying customer is this x spot frequency (1.7/yr). Both sources below quote
# average ticket PER TRANSACTION, so the basis matches with no conversion.
SPOT_TICKET = (
    ("uae", 190,
     "OBSERVED, PRIMARY. Botim gold: average ticket AED 700, Khaleej Times, Nov 2025 (= USD 191 at the "
     "peg). CROSS-CHECKS against the same platform's own volume disclosure - AED 100m over 128,000 trades "
     "implies AED 781. THE STRONGEST COMPARABLE IN THIS MODEL: Botim is a messaging app that became the "
     "default for Gulf expats and added gold in Aug 2025, so its users ARE Aurumix's customers - same "
     "country, same segment, same product, same period. Note 64% of its users buy under AED 500, so the "
     "mean sits above the median and this figure is, if anything, generous."),
    ("gulf", 145,
     "INFERRED - THE WEAKEST OF THE THREE. No GCC platform outside the UAE publishes an average ticket; "
     "this is a CONFIRMED NEGATIVE, not a gap in the search. Taken as the UAE figure scaled by the regional "
     "income gap (SIP 26.00/33.60 = 0.77x). Do not present this as sourced."),
    ("india", 40,
     "OBSERVED, PROXY. Augmont: average festive-season purchase Rs 3,300, Times of India, Sep 2025 "
     "(= ~USD 38). Augmont is a licensed refiner and digital gold platform; a festive lump is structurally "
     "the same act as a spot purchase, which is why it is preferred to the alternatives. REJECTED: SafeGold "
     "Rs 5,000-8,000 (May 2026) - that book skews to investors, while Aurumix's Indian customer saves USD 30 "
     "a month; MMTC-PAMP Rs 800-1,200 - closer to routine micro-saving than a lump; Rs 230 industry-wide - a "
     "LinkedIn aggregate, not a citable source. WEAKER THAN THE UAE FIGURE: right country and right "
     "behaviour, but Aurumix's Indian customer is poorer than Augmont's average buyer, so 40 may be generous."),
)
for key, tkt, _tnote in SPOT_TICKET:   # NB: not "note" - that shadows note()
    a_row("spotticket_" + key, "Spot ticket - %s" % RLAB[key], tkt, "USD/purchase",
          _tnote, FMT_USD, "spotticket_" + key)

# ---- Spot attach: REGIONALISED on affordability ----------------------------
# REGIONALISED 2026-08-21, from a single global 14%.
#
# STILL NO SOURCE - v2.6 records it as a confirmed negative: "there is no
# spot-attach benchmark for any comparable gold or savings product anywhere."
# What changed is not the evidence but the STRUCTURE: a flat rate asserted that
# a UAE customer and an Indian customer face the same decision, and they do not.
#
# THE DRIVER IS AFFORDABILITY, and it is observable. A spot purchase costs:
#     UAE              190 / 33.60 = 5.7 MONTHS of that customer's saving
#     Oman & Bahrain   145 / 26.00 = 5.6 months
#     India             40 / 30.00 = 1.3 months
# Committing six months of savings in one go is a rare act; committing five
# weeks' worth is an ordinary one. That gap - not appetite for gold - is why
# attach should differ, and it is why INDIA IS THE HIGH ONE despite being the
# poorest market. The population is also self-selected: these are people who
# already chose to buy gold every month, so a low attach was always hard to
# defend for them, and hardest of all in India where festive buying is close to
# universal.
# India raised 0.25 -> 0.35 on 2026-08-21, client instruction. Defensible on the
# same affordability logic that set the split: at 1.3 months of saving per
# purchase a festive lump is an ORDINARY act in India, so a third of customers
# doing it once a year is unremarkable. UAE and Oman & Bahrain DELIBERATELY LEFT
# ALONE - at 5.7 and 5.6 months per purchase, asserting that a fifth or a
# quarter of those customers commit most of half a year's savings in one go
# every year is not something that survives being asked out loud.
SPOT_ATTACH = (("uae", 0.12), ("gulf", 0.10), ("india", 0.35))
for key, att in SPOT_ATTACH:
    a_row("spotattach_" + key, "Spot attach rate - %s" % RLAB[key], att, "% of customers/yr",
          "Share of paying customers who make ANY spot purchase in a year - NOT a monthly rate, and NOT how "
          "often a buyer buys (that is the separate frequency input). Set on the affordability ratio above: "
          "the spot ticket as a multiple of the same customer's monthly SIP. NO PUBLISHED BENCHMARK EXISTS "
          "- confirmed negative. KNOWN SIMPLIFICATION: v2.6 applied a TENURE UPLIFT here ('a 3-year account "
          "is ~2x as likely to buy spot as a 6-month account') which v3.0 dropped, so this is flat from day "
          "one and OVERSTATES SPOT IN THE EARLY YEARS, when nearly every account is young. Restore the uplift "
          "in the Phase 5 simulation. ASSUMPTION, structured on an observable driver.",
          FMT_PCT, "spotattach_" + key)
# CONFIRMED NEGATIVE, recorded so it is not re-searched: no published average
# ticket exists anywhere for South Asian expatriate or blue-collar customers in
# the Gulf specifically - the same gap already hit on transaction frequency.
_ar[0] += 1

a_section("SIP RULES")
a_row("sip_floor", "SIP hard floor", 20, "USD/month",
      "Rejected outright below; never partially credited. CITED.", FMT_USD2, "sip_floor")
a_row("gate_run", "Confirmed SIP gate", 6, "consecutive payments",
      "Client's own figure. The gate is why eligibility is NOT universal - see the two eligibility parameters "
      "on the Scenario sheet. CLIENT INPUT.", FMT_NUM, "gate_run")
a_row("redemption_cost", "Cost per redemption event", 1.85, "USD",
      "CUT TWICE, AND BOTH CUTS WERE DELIBERATE. From 4.20 to 3.20 on 2026-08-20, removing the outbound bank "
      "transfer fee (USD 1.00-2.50) as the customer's cost to bear. From 3.20 to 1.85 on 2026-08-26, client "
      "instruction, removing the ~1.35 of operational handling. "
      "WHAT REMAINS IS SUMSUB AML RE-SCREENING AT 1.85, the same published per-verification rate the "
      "onboarding line uses - a redemption triggers a re-screen. NOT a double count with the KYC line: that "
      "one charges new customers at onboarding, this one charges a different event. "
      "WHAT WAS REMOVED IS NOT ZERO. Operational handling - the staff time to check, approve and settle a "
      "redemption - is real work; it is now assumed to sit inside headcount rather than being priced per "
      "event. WHEN HEADCOUNT LANDS, CONFIRM IT IS ACTUALLY CARRIED THERE, or this cost has quietly vanished "
      "from the model rather than moved. NOT IN THE REVENUE MODEL: "
      "redemption is a COST and arrives with the cost build. This row is carried so the driver is documented "
      "and the unit rate already agreed. THE FINDING IT CARRIES IS UNCHANGED AND IS THE POINT - VARA Annex 2 "
      "III.E.4 forbids charging ANY fee on redemption, verified verbatim at primary source, so the cost is "
      "100% absorbed, no offsetting revenue exists or can exist, and THERE CAN NEVER BE AN EXIT FEE. DERIVED.",
      FMT_USD2, "redemption_cost")
_ar[0] += 1

# ----------------------------------------------------- activation calendar --
a_section("ACTIVATION CALENDAR - the single source of every activation period")
ACTIVATIONS = [
    ("s1a", 1, "from", "Stream 1a - Entry fee, SIP",
     "M1. The core product - live at launch."),
    ("s1b", 1, "from", "Stream 1b - Entry fee, SPOT",
     "M1. A sub-stream of 1: the same fee on the same gold through a different rail, so it needs no "
     "additional build."),
    ("s0", 1, "from", "Stream 0 - Redemption cost",
     "M1. A mandatory cost from day one. VARA Annex 2 III.E.4 forbids charging any fee on redemption, so "
     "there is no offsetting revenue and there can never be an exit fee."),
    ("s3", 7, "from", "Stream 3 - Family plan and Digital Will",
     "M7 (client decision 2026-08-20). RATIONALE: this is a paid add-on sold INTO an existing base, not an "
     "acquisition product. At M1 there is no base to attach it to, and the Digital Will needs the legal "
     "template and beneficiary flow finished. Six months lets the first cohort establish before they are "
     "asked to buy a second product."),
    ("s2", 13, "from", "Stream 2 - Card interchange",
     "M13 (client decision 2026-08-20, moved earlier from M18). RATIONALE FOR ANY DELAY AT ALL: the card is "
     "not ours to launch. It needs a sponsor bank (BIN sponsorship), Visa scheme certification and a "
     "processor integration - all commercial gates outside our control. It also needs customers who have "
     "cleared the six-payment gate, which takes ~8 months on average, so a card launched much earlier would "
     "have almost nobody eligible to hold it."),
    ("s4", 13, "from", "Stream 4 - Cardholder fees",
     "M13, necessarily with stream 2 - FX, ATM and issuance fees cannot exist before the card does."),
    ("s5", 13, "from", "Stream 5 - Lending revenue share",
     "M13 (client decision 2026-08-20, moved earlier from Y3). RATIONALE FOR ANY DELAY AT ALL: lending needs "
     "a partner lender under contract, and it needs collateral that has SEASONED 90 days. A customer joining "
     "at M1 has enough seasoned gold by M13; one joining at M10 does not, which the model handles because "
     "borrowers are drawn from the gate-cleared population, not from everyone."),
    ("s6", 13, "from", "Stream 6 - B2B platform fee",
     "M13 (client decision 2026-08-20, moved earlier from Y3). RATIONALE FOR ANY DELAY AT ALL: it needs a "
     "signed partner and multi-tenant capability at register and mint. NOTE: at Y3 this stream had to be "
     "placed at the START of the annual block to avoid a one-twelfth-year stub column sitting beside a full "
     "annual one. At M13 it sits safely inside the monthly block, so that complication disappears."),
    ("referral", 13, "from", "Referral channel begins",
     "M13, and this one is STRUCTURAL rather than commercial: a referrer must clear their own six-payment "
     "gate before they can refer, and the referee must then clear theirs. Two six-month gates IN SERIES. "
     "Year 1 pays nothing and the channel does not reach steady state until roughly M25."),
    ("gulf", 13, "from", "Region opens - Oman and Bahrain",
     "M13 (client decision 2026-08-20). THE ONLY STAGED REGION, and the reason is regulatory: it requires "
     "local authorisation in a second and third jurisdiction. These customers are also NOT UAE-resident, so "
     "the VAT and reporting treatment differs."),
    ("india", 1, "from", "Region opens - India",
     "M1 (client decision 2026-08-20, moved earlier from M13). The India payment route is ASSUMED SOLVED in "
     "this model. That is a live assumption, not a settled fact - if the route is not solved, this region "
     "does not open at all and roughly a quarter of terminal revenue goes with it."),
]
for key, per, kind, label, src in ACTIVATIONS:
    a_row("act_" + key, label, per, "period # (%s)" % kind, src, FMT_NUM, "act_" + key)
_ar[0] += 1

# ---------------------------------------------------------- seasonality -----
SEAS_HDR = _ar[0]
section(ws_assum, SEAS_HDR, "SEASONALITY VECTORS (raw - the Model normalises to exactly 12.000)", span=4)
SEAS_COL0 = 6
for k, mn in enumerate(MONTHS):
    c = ws_assum.cell(row=SEAS_HDR + 1, column=SEAS_COL0 + k)
    c.value, c.fill, c.font = mn, HEADER_FILL, HEADER_FONT
    c.alignment = Alignment(horizontal="center")
    ws_assum.column_dimensions[get_column_letter(SEAS_COL0 + k)].width = 8
ws_assum["A%d" % (SEAS_HDR + 1)] = "Vector"
_ar[0] = SEAS_HDR + 2
SEAS_ROW = {}
for key, label, vals, src in [
    ("acq", "Acquisition seasonality",
     [1.01, 0.96, 0.92, 1.15, 1.04, 0.88, 0.85, 0.85, 0.93, 1.25, 1.15, 1.01],
     "SOURCED TIMING, JUDGED AMPLITUDE. Peaks at Akshaya Tritiya (Apr/May), the most concentrated gold-buying "
     "day globally by value, and Dhanteras/Diwali (Oct/Nov), India's peak gold occasion with demand up 8-12% "
     "in the fortnight before. Wedding season Nov-Mar supports the winter. Trough Jul-Sep, the sourced annual "
     "low. Jan carries the Dubai Shopping Festival. DAMPED to ~+/-25% against the ~+/-33% swing in raw "
     "jewellery demand, because a USD 20/month savings mandate is far less impulsive than buying jewellery. "
     "Two traps avoided: India's Q3 2024 spike was an IMPORT DUTY CUT, not seasonality; and Ramadan/Eid moves "
     "~11 days a year, so a fixed-month Eid bump would be wrong by construction over 7 years."),
    ("spend", "Card spend seasonality",
     [1.10, 1.00, 0.98, 0.98, 1.00, 0.95, 0.88, 0.90, 1.00, 1.02, 1.05, 1.14],
     "UAE RESIDENT CARD SPEND, not gold. Peaks Dec-Jan on the Dubai Shopping Festival and year-end holidays; "
     "troughs Jul-Aug when residents leave for the summer. NARROWER than acquisition (+/-14% vs +/-25%) "
     "because everyday spend is less culturally spiky. Note the deliberate OPPOSITION to the foreign-spend "
     "vector: Jul-Aug is the weakest total-spend month and the strongest foreign-share month."),
    ("foreign", "Foreign-spend share vector",
     [30, 30, 30, 32, 34, 55, 60, 56, 36, 42, 40, 32],
     "The one seasonality vector the brief itself tabulates. Raw mean 39.75, rescaled on the Model to the "
     "foreign-spend-share parameter. THE SUMMER TRAVEL SEASON IS THE LARGEST STREAM-4 MONTH on the weakest "
     "total-spend base."),
]:
    r = _ar[0]
    ws_assum["A%d" % r] = label
    for k, v in enumerate(vals):
        c = ws_assum.cell(row=r, column=SEAS_COL0 + k)
        c.value, c.font, c.number_format = v, BLUE, FMT_NUM2
        c.alignment = Alignment(horizontal="center")
    ws_assum["D%d" % r] = src
    SEAS_ROW[key] = r
    _ar[0] += 1
_ar[0] += 1

# ------------------------------------------------------ market ceilings -----
a_section("REGION KEY - who each region is")
for key in REGIONS:
    r = _ar[0]
    ws_assum["A%d" % r] = RLAB[key]
    ws_assum["A%d" % r].font = BLACK_BOLD
    # R1 carries no activation row: it is the baseline market, open from M1 by
    # definition, and the region renormalisation treats it as always open.
    _opens = dict((k, p) for k, p, _a, _b, _c in ACTIVATIONS).get(key, 1)
    ws_assum["B%d" % r] = "opens M%d" % _opens
    ws_assum["B%d" % r].font = SECONDARY
    ws_assum["D%d" % r] = RDESC[key]
    _ar[0] += 1
_ar[0] += 1

# ---------------------------------------------------------------------------
# MARKET SIZING - the funnel, shown as live arithmetic rather than a typed
# ceiling. Four source populations feed three model regions; every step is a
# visible cell, so a reader can see exactly how 12.5m people become 43,750
# reachable accounts and can argue with any single filter.
# ---------------------------------------------------------------------------
FUN_HDR = _ar[0]
section(ws_assum, FUN_HDR, "MARKET SIZING - how each region's reachable SIP account ceiling is derived", span=4)
FUN0 = 6                                    # column F
FUNNEL = [
    # CEILINGS RAISED 2026-08-21, client instruction. UAE-Indian 9.5% -> 10%,
    # UAE-other-South-Asian 6% -> 10%, Oman & Bahrain 4% -> 6%.
    #
    # THE OTHER-SOUTH-ASIAN MOVE IS THE ONE TO NOTICE. 6% vs 9.5% was not
    # arbitrary - it encoded that this is an Indian-built product, sold through
    # Indian agents, along Indian remittance corridors. Equalising the two at
    # 10% ASSERTS THAT AURUMIX REACHES PAKISTANI, BANGLADESHI, NEPALI AND SRI
    # LANKAN WORKERS AS EFFECTIVELY AS INDIAN ONES. That may be true, but it is
    # a claim about distribution, not a number - and the v2.6 brief argued the
    # opposite, that the agent network "must recruit non-Indian agents before it
    # reaches the other South Asian market at all". If that recruitment does not
    # happen, this column belongs back near 6%.
    ("uae_ind", "UAE - Indian", 4360000, 0.80, 0.57, 0.40, 0.100, "uae"),
    ("uae_osa", "UAE - other South Asian", 3460000, 0.80, 0.57, 0.40, 0.100, "uae"),
    ("gulf", "Oman and Bahrain", 2630000, 1.00, 0.57, 0.40, 0.060, "gulf"),
    # India ceiling RAISED 2026-08-21 from 0.35%. It was THE BINDING CONSTRAINT
    # IN THE MODEL and nothing else was close: at Y7 India's raw demand ran at
    # ~21,900 new customers a year and saturation cut it to ~5,900 - 73% of the
    # demand the acquisition engine generated was thrown away, against 19% in
    # the UAE and 25% in Oman & Bahrain. The model wanted to grow in India and
    # was not allowed to.
    #
    # RAISED AGAIN to 1.00% on 2026-08-21, client instruction: 125,000 of
    # 12.5m active digital-gold users, or 1 in 100 over seven years. Jar (20m users) and Augmont (42m registered) prove the
    # behaviour exists at that scale, so 1 in 285 was hard to defend as a
    # CEILING - a number meant to bound what is ever reachable, not to forecast.
    #
    # THIS IS THE MOST DANGEROUS NUMBER IN THE FUNNEL. Revenue moves almost
    # linearly with it and it is the least falsifiable input here, so it must
    # never be moved to make an output look better. The real question it answers
    # is not "how big is Indian gold demand" - plainly enormous - but "what is
    # Aurumix's RIGHT TO WIN": a Dubai-vaulted product competing with entrenched
    # domestic platforms and sovereign gold bonds, with FEMA/LRS unresolved and
    # currently assumed away. If that route fails, this belongs back at 0.35% or
    # lower, and the India block should be switched off entirely.
    ("india", "India", 12500000, 1.00, 1.00, 1.00, 0.0100, "india"),
]
for k, (key, nm, pop, ea, pay, mon, ceil, rgn) in enumerate(FUNNEL):
    c = ws_assum.cell(row=FUN_HDR + 1, column=FUN0 + k)
    c.value, c.fill, c.font = nm, HEADER_FILL, HEADER_FONT
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws_assum.column_dimensions[get_column_letter(FUN0 + k)].width = 15
ws_assum["A%d" % (FUN_HDR + 1)] = "Funnel step"
ws_assum["A%d" % (FUN_HDR + 1)].fill = HEADER_FILL
ws_assum["A%d" % (FUN_HDR + 1)].font = HEADER_FONT
_ar[0] = FUN_HDR + 2

FUN_ROW = {}


def fun_row(key, label, vals, fmt, src, font=BLUE):
    r = _ar[0]
    ws_assum["A%d" % r] = label
    for k, v in enumerate(vals):
        c = ws_assum.cell(row=r, column=FUN0 + k)
        c.value = v
        c.font = font
        c.number_format = fmt
        c.alignment = Alignment(horizontal="center")
    ws_assum["D%d" % r] = src
    FUN_ROW[key] = r
    _ar[0] += 1
    return r


fun_row("pop", "Source population", [f[2] for f in FUNNEL], FMT_NUM,
        "UPDATED 2026-08-20. UAE-Indian is 4.36m - the Indian consulate's official count at December 2024, "
        "also given as 4.3m by the Embassy of India. The brief carried 3.5m, which is roughly a 2021-22 "
        "vintage: the community was 3.89m at December 2023 and 4.36m a year later, having doubled from 2.2m "
        "over the decade. UAE-other South Asian is 3.46m, the sum of Pakistani 1.90m, Bangladeshi 0.84m, "
        "Nepali 0.36m and Sri Lankan 0.36m - the brief's 3.4m was already correct. These are TOTAL RESIDENT "
        "NATIONALS including dependants, which is why the economically-active filter below applies. Oman and "
        "Bahrain is an expatriate WORKER count. INDIA IS NOT A DEMOGRAPHIC BASE - it is sized BEHAVIOURALLY "
        "from gold ETF folios intersected with active digital-gold holders, so its filters are 100%. "
        "HELD STATIC over the horizon, which is conservative: the UAE Indian community grew ~12% in the "
        "single year to Dec 2024.")
fun_row("ea", "  x Economically active", [f[3] for f in FUNNEL], FMT_PCT,
        "Oman and Bahrain is 100% because the source counts EXPAT WORKERS, who are working-age and "
        "economically active by definition - applying a labour-force filter would double-count.")
fun_row("pay", "  x Payment capable (IBAN able to carry a direct debit)", [f[4] for f in FUNNEL], FMT_PCT,
        "A SIP needs a standing mandate, not just an account. This is the filter that replaced an unsourced "
        "gold-savings propensity filter, which was flagged as the weakest link in the entire sizing and had "
        "no published source. This one has a stated mechanism behind it.")
fun_row("mon", "  x Money capable (USD 20/month after remittances)", [f[5] for f in FUNNEL], FMT_PCT,
        "Discretionary capacity for a RECURRING USD 240/year commitment, tested after remittance obligations "
        "rather than before them.")
_ab = _ar[0]
fun_row("addr", "  = Addressable base", [None] * 4, FMT_NUM,
        "Population x the three filters. The UAE-Indian column reproduces the brief's worked figure of "
        "~640,000.", BLACK_BOLD)
for k in range(4):
    col = get_column_letter(FUN0 + k)
    c = ws_assum["%s%d" % (col, _ab)]
    c.value = "=%s%d*%s%d*%s%d*%s%d" % (col, FUN_ROW["pop"], col, FUN_ROW["ea"],
                                        col, FUN_ROW["pay"], col, FUN_ROW["mon"])
    c.font, c.number_format = BLACK_BOLD, FMT_NUM
    c.alignment = Alignment(horizontal="center")
fun_row("ceilpct", "  x Penetration ceiling", [f[6] for f in FUNNEL], FMT_PCT2,
        "The share of the addressable base the product could ever reach. 'base x ceiling' is the INVARIANT - "
        "changing one without the other silently re-scales the whole model.")
_rc = _ar[0]
fun_row("acct", "  = REACHABLE SIP ACCOUNTS", [None] * 4, FMT_NUM,
        "This is what the acquisition engine saturates against.", BLACK_BOLD)
for k in range(4):
    col = get_column_letter(FUN0 + k)
    c = ws_assum["%s%d" % (col, _rc)]
    c.value = "=%s%d*%s%d" % (col, _ab, col, FUN_ROW["ceilpct"])
    c.font, c.number_format = BLACK_BOLD, FMT_NUM
    c.alignment = Alignment(horizontal="center")
r = _ar[0]
ws_assum["A%d" % r] = "  Feeds model region"
for k, f in enumerate(FUNNEL):
    c = ws_assum.cell(row=r, column=FUN0 + k)
    c.value = RLAB[f[7]]
    c.font = SECONDARY
    c.alignment = Alignment(horizontal="center")
_ar[0] += 2

a_section("MARKET CEILINGS AND ACQUISITION DRIVERS")
for key in REGIONS:
    cols = [get_column_letter(FUN0 + k) for k, f in enumerate(FUNNEL) if f[7] == key]
    src = ("DERIVED from the market sizing funnel above - the sum of its source populations (%s), not a typed "
           "number. Change any filter and this follows."
           % ", ".join(FUNNEL[k][1] for k, f in enumerate(FUNNEL) if f[7] == key))
    r = _ar[0]
    ws_assum["A%d" % r] = "Reachable SIP ceiling - %s" % RLAB[key]
    c = ws_assum["B%d" % r]
    c.value = "=" + "+".join("%s%d" % (cl, _rc) for cl in cols)
    c.font, c.number_format = BLACK_BOLD, FMT_NUM
    ws_assum["C%d" % r] = "accounts"
    ws_assum["D%d" % r] = src
    AROW["ceil_" + key] = r
    declare("ceil_" + key, "Assumptions", "$B$%d" % r)
    _ar[0] += 1
r = _ar[0]
ws_assum["A%d" % r] = "Total reachable SIP accounts"
ws_assum["A%d" % r].font = BLACK_BOLD
ws_assum["B%d" % r] = "=" + "+".join(aref("ceil_" + k).replace("Assumptions!", "") for k in REGIONS)
ws_assum["B%d" % r].font = BLACK_BOLD
ws_assum["B%d" % r].number_format = FMT_NUM
ws_assum["C%d" % r] = "accounts"
ws_assum["D%d" % r] = ("~181,150, against the brief's 165,750. THE BRIEF'S 'base x ceiling is the "
                       "invariant' RULE IS DELIBERATELY BROKEN HERE, and the distinction matters: that rule "
                       "guards against a METHODOLOGICAL re-cut quietly inflating the market by re-tuning "
                       "filters, which proves nothing. This is not that - it is a DATA CORRECTION. The UAE "
                       "Indian population really is 4.36m rather than 3.5m, so the addressable market really "
                       "is larger. Holding the invariant would mean asserting that penetration falls exactly "
                       "as population rises, which has no basis. The penetration ceilings are untouched.")
AROW["ceil_total"] = r
declare("ceil_total", "Assumptions", "$B$%d" % r)
_ar[0] += 1

# ---- CROSS-CHECK against observed platform users ---------------------------
# Added 2026-08-21. The funnel derives the market from population and filters;
# this tests that answer against what a real platform in the same market has
# actually seen. THESE ROWS FEED NOTHING - they are evidence, not inputs.
#
# WHY THIS IS A CROSS-CHECK AND NOT THE METHOD. Replacing the funnel with
# platform user counts was considered and rejected on the data:
#   (a) "USERS" IS NOT A DEFINED UNIT. For the SAME UAE product the published
#       figures are 8.5m (can access), 1.5m (OGold users), 775,000 (explored the
#       feature) and 75,000 (active). A 113x spread - choosing between them is a
#       judgement at least as large as any filter it would replace, but it would
#       LOOK like an observation, which is worse.
#   (b) INDIA CANNOT BE DEDUPLICATED. Jar reports 35m registered (Sep 2025) and
#       Augmont 42m; the same person is on several. No industry estimate of
#       UNIQUE digital gold investors exists - a confirmed negative. Registered
#       is not active either: Augmont runs 0.85 transactions per registered user
#       per year, so most are dormant.
#   (c) OMAN & BAHRAIN HAS NO PUBLISHED USER COUNT AT ALL, so the method could
#       not be applied uniformly - which was its main appeal.
#   (d) BEHAVIOUR IS NOT CAPACITY, and this is the one that decides it. A Jar
#       user saving Rs 10/day proves they will buy gold digitally; it does NOT
#       prove they can commit Rs 2,640/month to a SIP. The funnel's filters are
#       not measuring appetite for gold - appetite is not in doubt - they are
#       measuring capacity for a RECURRING commitment. Swapping in platform
#       users would delete that test and replace it with nothing.
r = _ar[0]
ws_assum["A%d" % r] = "CROSS-CHECK vs observed market evidence (memo - feeds nothing)"
ws_assum["A%d" % r].font = BLACK_BOLD
_ar[0] += 1
a_row("xchk_uae_obs", "  O Gold active users, UAE (observed)", 75000, "users",
      "OBSERVED. O Gold, the gold engine behind Botim, reported at 75,000 ACTIVE users. The same product is "
      "also reported at 775,000 'explored the feature', 1.5m 'users' and 8.5m 'can access' - ACTIVE is the "
      "only one of the four comparable to a reachable-accounts figure, and it is the most conservative. "
      "CITED.", FMT_NUM, "xchk_uae_obs")
r = _ar[0]
ws_assum["A%d" % r] = "  Our UAE ceiling as a multiple of observed active users"
c = ws_assum["B%d" % r]
c.value = "=%s/%s" % (aref("ceil_uae").replace("Assumptions!", ""),
                      aref("xchk_uae_obs").replace("Assumptions!", ""))
c.font, c.number_format = BLACK_BOLD, '0.00"x"'
ws_assum["C%d" % r] = "x observed"
ws_assum["D%d" % r] = (
    "LIVE, SO IT MOVES IF ANYONE RAISES A CEILING - which is the point. Two INDEPENDENT methods landing "
    "within ~2x is the strongest defence of the funnel in this model: the funnel says the UAE can reach "
    "~113,000 SIP accounts, and a real UAE gold app in the same segment has ~75,000 active users today. "
    "Our addressable base of ~1.43m sits against 775,000 who explored that product - same order again. "
    "If this multiple drifts far above ~2x, the ceiling has stopped being defensible by evidence and is "
    "being set by the answer someone wanted. DERIVED from a cited observation.")
AROW["xchk_uae_ratio"] = r
_ar[0] += 1
r = _ar[0]
ws_assum["A%d" % r] = "  Oman & Bahrain / India (no comparable figure)"
ws_assum["B%d" % r] = "n/a"
ws_assum["B%d" % r].font = SECONDARY
ws_assum["C%d" % r] = "-"
ws_assum["D%d" % r] = (
    "CONFIRMED NEGATIVE, recorded so it is not re-searched. No GCC platform outside the UAE publishes a "
    "user count. India publishes only REGISTERED counts (Jar 35m, Augmont 42m) which cannot be deduplicated "
    "or converted to active users, so no comparable cross-check can be built. NOTE THAT INDIA'S OWN BASE "
    "IS ALREADY user-derived - its 12.5m is 'holds a gold ETF folio or actively buys digital gold', which is "
    "why its three filters all sit at 1.00 - AND IT IS THE LEAST TRACEABLE BASE IN THE MODEL. That is the "
    "evidence that a user-based method does not automatically produce a better-sourced number.")
_ar[0] += 2

# REGIONALISED 2026-08-26. Agents now BELONG to a region rather than sitting in
# a national pool routed by a "share of salesforce" input. UAE and Oman &
# Bahrain are ZERO by design: those markets are acquired top-down through
# marketing at their own CAC, plus referral. India carries the agent network.
AGENTS_BY_REGION = {
    "uae":   [0, 0, 0, 0, 0, 0, 0],
    "gulf":  [0, 0, 0, 0, 0, 0, 0],
    "india": [40, 40, 60, 177, 298, 356, 420],
}
MKTG = [90000, 270000, 375000, 600000, 900000, 1275000, 1650000]   # was 60k..1.1m
RAMP = [0.60, 0.85, 1.00, 1.00, 1.00, 1.00, 1.00]
AGENT_NOTE = (
    "REGIONALISED 2026-08-26, replacing a single national pool split 45/10/45 by a 'share of salesforce' "
    "input. AGENTS NOW BELONG TO A REGION. UAE and Oman & Bahrain are set to ZERO: those markets are "
    "acquired top-down through marketing at their own CAC, plus referral. India carries the agent network. "
    "THE RAMP IS NO LONGER TYPED - and the schedule it replaces was never the client's. The old "
    "5/15/40/90/200 entered the corpus in the v1.0 brief as 'Active agents | Client input REQUIRED', a "
    "placeholder flagged as a question to ask; by the v2.6 rebuild spec the word 'required' had dropped and "
    "the source column read 'CLIENT INPUT'. That sequence appears in four documents, all of them ours, and "
    "in no client source: the 100 G Business Model gives the three-tier agent STRUCTURE (S11.1) and the 15% "
    "commission, but no headcount anywhere. The only genuine client instruction was a +50% raise on "
    "2026-08-21, which scaled a placeholder rather than replacing it. "
    "THIS ROW IS NOW OBSERVED. It is Angel One's DISCLOSED Authorised Person count - 42 / 40 / 63 / 186 / "
    "313 across FY18-FY22, a 7.45x rise over four years from a small base - indexed to Y1 and anchored at "
    "40 (client decision 2026-08-26). An Angel One AP is a SEBI-registered independent person who acquires "
    "and services retail investors on commission, which is the same animal as an Aurumix agent. The tail "
    "(Y6-Y7) is extended at Prudent Corporate Advisory's mature rate, whose MFD base went 14,007 (FY20) to "
    "33,308 (FY25), ~19%/yr. "
    "THE OBSERVED Y2 DIP IS FLATTENED, client instruction: Angel One's own FY19 fell 5% on FY18, which is "
    "what real networks do - you recruit, most wash out, the survivors compound from year 3. Held FLAT at "
    "Y2 instead, which is not a free year: at 45% attrition it still means ~18 hires purely to stand still. "
    "TWO CAVEATS ON THE COMPARATORS. Angel One's FY18 is NOT that company's year one - it was founded in "
    "1996 and was rebuilding the AP channel, so this is 'a distributor network scaling from a small base', "
    "not 'a startup's first five years'. And both comparators RECRUIT PRE-LICENSED PROFESSIONALS onto a "
    "platform rather than creating agents from nothing; gold is not a securities product so Aurumix needs "
    "no ARN or SEBI registration, which removes the barrier but also removes the ready pool to poach from. "
    "CROSS-CHECKED, AND THIS IS THE STRONGEST PART: sizing agents bottom-up so they are India's primary "
    "channel independently gives ~470 at Y7 against this curve's 420 - two unrelated methods, 12% apart. "
    "Y4 IS THE YEAR TO QUESTION: 60 to 177 needs 144 hires against a base of 60, nearly tripling the "
    "network in twelve months. It is inherited from Angel One's real FY20-FY21 jump (63 to 186), so it is "
    "observed rather than invented, but it is the year the plan has to actually deliver. "
    "AND IT IS A COST THIS MODEL CANNOT SEE: 857 hires across seven years to end with 420 active, ~224 in "
    "Y7 alone, needing 5-6 full-time recruiters and ~28 supervisors at a 1:15 span. "
    "OBSERVED (Angel One FY18-FY22, Prudent FY20-FY25); the Y1 anchor is a CLIENT DECISION.")
for key in REGIONS:
    for y in range(1, 8):
        a_row("agents_%s_y%d" % (key, y),
              "Active agents - %s - Y%d" % (RLAB[key], y),
              AGENTS_BY_REGION[key][y - 1], "agents", AGENT_NOTE, FMT_NUM,
              "agents_%s_y%d" % (key, y))
for y in range(1, 8):
    a_row("mktg_y%d" % y, "Marketing spend - Y%d" % y, MKTG[y - 1], "USD/yr",
          "RAISED 50% on 2026-08-21, client instruction. A DECISION VARIABLE and an input to acquisition, "
          "not an output of a cost table. THE SINGLE LARGEST LEVER IN THE MODEL AND THE ONLY ONE THAT IS "
          "PURELY A SPENDING CHOICE: measured sensitivity puts +50% here at +28.5% customers and +29.3% "
          "gold. IT IS ALSO THE ONE THE REVENUE-ONLY SCOPE FLATTERS MOST. This model has no cost side yet, "
          "so an extra USD 1.9m of marketing across the horizon appears here as pure upside. It is not - it "
          "is USD 1.9m of spend, and whether it is worth making is a question only the cost build can "
          "answer. DO NOT PRESENT THE UPLIFT FROM THIS ROW AS PROFIT.", FMT_USD, "mktg_y%d" % y)
for y in range(1, 8):
    a_row("ramp_y%d" % y, "Agent blended ramp - Y%d" % y, RAMP[y - 1], "x productivity",
          "Six months to full productivity, which happens to match the six-payment gate.", FMT_NUM2, "ramp_y%d" % y)
# DELETED 2026-08-26: "Share of salesforce deployed - <region>" (45/10/45).
# Those three inputs, and the open-region renormalisation row on the Model that
# divided by them, existed ONLY to route a single national agent pool to the
# markets it worked in. Once headcount belongs to a region there is nothing to
# route: the agent block above IS the allocation. Three inputs and one model row
# removed - this change makes the model smaller, not bigger.
_ar[0] += 1
ws_assum.freeze_panes = "B6"

# ============================================================================
# SCENARIO PARAMETERS
# ============================================================================
ws_scen["A1"] = "=Cover!$B$2"
ws_scen["A1"].font = SHEET_TITLE
ws_scen["A2"] = "Scenario Parameters"
ws_scen["A2"].font = SHEET_SUB
widths(ws_scen, {"A": 50, "B": 15, "C": 14, "D": 14, "E": 14, "F": 24, "G": 118})

section(ws_scen, 4, "SCENARIO SELECTOR", span=5)
ws_scen["A6"] = "Select Scenario:"
ws_scen["A6"].font = BLACK_BOLD
ws_scen["B6"] = "Base"
ws_scen["B6"].font = BLUE_BOLD
ws_scen["C6"] = '=IF($B$6="Base",1,IF($B$6="Aggressive",2,3))'
ws_scen["D6"] = "<- index"
ws_scen["D6"].font = SECONDARY
ws_scen["G6"] = "One switch, three columns. Every parameter below resolves through it."
ws_scen["G6"].font = NOTE_FONT
dv = DataValidation(type="list", formula1='"Base,Aggressive,Conservative"', allowBlank=False)
ws_scen.add_data_validation(dv)
dv.add(ws_scen["B6"])
declare("scenario_index", "Scenario Parameters", "$C$6")

SROW, _sr = {}, [8]
SIDX = "$C$6"
# Every scenario parameter is MIRRORED into Assumptions as a green link, and the
# Model reads only from there. Recorded here so the mirror can be generated with
# the same label, unit, format and rationale - one source in code, no drift.
SORDER, SMETA = [], {}


def s_section(t):
    r = _sr[0]
    section(ws_scen, r, t, span=5)
    headers(ws_scen, r + 1, [("A", "Parameter"), ("B", "Active Value"), ("C", "Base"),
                             ("D", "Aggressive"), ("E", "Conservative"), ("F", "Unit")])
    _sr[0] = r + 2
    return r


def sp(key, param, base, agg, cons, unit, why, fmt=FMT_NUM2, name=None):
    r = _sr[0]
    ws_scen["A%d" % r] = param
    ws_scen["B%d" % r] = "=CHOOSE(%s,C%d,D%d,E%d)" % (SIDX, r, r, r)
    ws_scen["B%d" % r].font = BLACK_BOLD
    ws_scen["B%d" % r].number_format = fmt
    for col, val in (("C", base), ("D", agg), ("E", cons)):
        c = ws_scen["%s%d" % (col, r)]
        if val is None:
            c.value, c.font = UNFILLED, RED_BOLD
        else:
            c.value, c.font, c.number_format = val, BLUE, fmt
    ws_scen["F%d" % r] = unit
    ws_scen["F%d" % r].font = SECONDARY
    ws_scen["G%d" % r] = why
    ws_scen["G%d" % r].font = NOTE_FONT
    SROW[key] = r
    SORDER.append(key)
    SMETA[key] = (param, unit, why, fmt)
    if name:
        declare(name, "Scenario Parameters", "$B$%d" % r)
    _sr[0] = r + 1
    return r


def s_derived(key, param, formula, unit, why, fmt=FMT_PCT2, name=None):
    r = _sr[0]
    ws_scen["A%d" % r] = param
    ws_scen["B%d" % r] = formula
    ws_scen["B%d" % r].font = BLACK_BOLD
    ws_scen["B%d" % r].number_format = fmt
    ws_scen["F%d" % r] = unit
    ws_scen["F%d" % r].font = SECONDARY
    ws_scen["G%d" % r] = why
    ws_scen["G%d" % r].font = NOTE_FONT
    SROW[key] = r
    SORDER.append(key)
    SMETA[key] = (param, unit, why, fmt)
    if name:
        declare(name, "Scenario Parameters", "$B$%d" % r)
    _sr[0] = r + 1
    return r


# While the Scenario sheet is being built, references point at itself. Once the
# Assumptions mirror exists this is REBOUND to point there instead, so every
# Model formula resolves through Assumptions - see the mirror block below.
def sref(k): return "'Scenario Parameters'!$B$%d" % SROW[k]


s_section("GROUP A0: THE METAL")
sp("gold_cagr", "Gold price appreciation", 0.081, 0.120, 0.000, "%/yr",
   "ADDED 2026-08-21, client instruction, replacing a flat price. 8.1% is the COMPOUND ANNUAL GROWTH RATE "
   "SINCE 1971, the end of the gold standard - chosen because it is the LONGEST window and therefore the "
   "least cherry-picked. The shorter the window the better gold looks: ~8.6% over 50 years, ~9-10% over 20, "
   "~13-14% over 10, and over 20% over the last 5. Taking the recent decade would have flattered the model "
   "on a rally. CITED. "
   "THIS IS A NOMINAL RATE AND NOTHING ELSE IN THE MODEL INFLATES. SIP tickets, spot tickets, CAC, card "
   "fees and the B2B fee are all frozen in 2027 dollars. There is NO CONSENSUS POSITIVE LONG-RUN REAL RETURN "
   "for gold - it is generally treated as a store of value with a near-zero real yield - so most of this 8.1% "
   "IS INFLATION. Applying it to the metal while freezing everything else quietly asserts that gold outruns "
   "wages, prices and costs by 8% a year for seven years, which is not a claim anyone would defend out loud. "
   "The Model therefore also carries AUM AT THE CONSTANT M1 PRICE, and that is the honest series for judging "
   "whether the BUSINESS grew. Conservative is set at 0.0% deliberately: it restores the original flat-price "
   "design, under which every revenue change is attributable to the business rather than the metal.",
   FMT_PCT2, "gold_cagr")
sp("float_buffer_days", "Float buffer - days of demand held", 10, 6, 20, "days",
   "S50, the SOFT HALF of the float sizing rule. The corpus states the rule as 'one bar denomination plus a "
   "buffer of N days trailing inflow' and NEVER SETS N. "
   "WHAT THE BUFFER IS FOR: how long it takes to get REPLACEMENT metal into the vault - dealer lead time, "
   "weekend and holiday gaps, payment clearing, and the days a dealer cannot fill 100 g. It is a "
   "REPLENISHMENT question. "
   "NOT the fill window, which is how long demand takes to CONSUME a bar. The brief anchors the 10 days on "
   "the fill window and that is the wrong anchor - how fast the shelf empties and how fast it can be "
   "restocked are different questions that only happen to give a similar answer at launch. Corrected here "
   "2026-08-26; the brief still carries the old reasoning. "
   "10 days is a working number for a dealer who has not been chosen yet, and it is DEALER-DEPENDENT: a "
   "T+2 commitment would cut the float by roughly two-thirds. "
   "IT CHANGES NOTHING AT LAUNCH. Below 10 grams a day - which is about month six - the two-bar floor is "
   "larger, so M1 is USD 28,716 for ANY buffer from 0 to 49 days. It binds only at scale, where it is worth "
   "USD 64,669 per day of buffer by Y7 and USD 905k between the Aggressive and Conservative settings. "
   "ASSUMPTION, and the one to flex: the only number in the float calculation that was neither observed nor "
   "settled in the design docs.",
   FMT_NUM, "float_buffer_days")
_sr[0] += 1

s_section("GROUP A: CUSTOMERS AND CHURN")
sp("persistency_m13", "Persistency - customers still paying after 12 months", 0.55, 0.65, 0.45, "%",
   "THE HEADLINE RETENTION NUMBER, and the one a client can argue with. Everything about churn is derived "
   "from it. Governs lifetime value, referral economics and agent commission at once. DERIVED from the "
   "archetype work: the simulation can refine it, but 55% is what that engine produced.", FMT_PCT, "persistency_m13")
s_derived("monthly_churn", "Monthly churn rate (derived)",
          "=1-%s^(1/12)" % sref("persistency_m13"), "%/month",
          "DERIVED, not typed: the monthly rate that reproduces the persistency above over twelve months. "
          "Change persistency and this follows. Base 55% implies ~4.9%/month.", FMT_PCT2, "monthly_churn")
sp("agent_productivity", "Agent productivity", 6, 9, 3, "accounts/agent/month",
   "RE-EVIDENCED 2026-08-26. Base HOLDS at 6; the BANDS widen from 8/3 to 9/3 and the justification is "
   "rebuilt on measured comparators instead of resting on a single insurance-agency analogy. "
   "WHAT THE EVIDENCE SAYS. IRDAI's Annual Report FY25 gives 270.22 lakh new individual policies against "
   "31.23 lakh individual agents - 8.65 policies per agent per YEAR, or 0.72/month, across the whole book "
   "INCLUDING dormant agents. Adjusting for dormancy (MicroSave measures up to 43% inactive across the top "
   "four BC network managers) puts an ACTIVE Indian life agent at roughly 1.5-3 per month. "
   "BEWARE ONE WIDELY-CIRCULATED FIGURE: an analyst piece states Indian productivity as '0.87 policies per "
   "agent per year'. That is an ARITHMETIC SLIP - it divides 27.02 lakh by 31.23 lakh when the numerator is "
   "270.22 lakh. The correct figure is 8.65. Do not cite the 0.87. "
   "THE CLOSER COMPARATOR IS THE BANK MITRA / BC NETWORK, which opens a light, form-based account for the "
   "same demographic through the same field model: 20-30 accounts per month when the account is FREE, and "
   "about 2 per month once a fee is introduced - a 93% collapse, measured by MicroSave. AURUMIX CHARGES 5% "
   "ON ENTRY. That is the datapoint arguing 6 is at the OPTIMISTIC end rather than the conservative one, "
   "and it is why the conservative band moves from 3 to... 3, but now for a stated reason rather than as a "
   "round number. "
   "THE OLD NOTE argued 6 was conservative because a gold SIP signup is a fraction of the work of an "
   "underwritten insurance policy. That is true, and it is why 6 survives as Base - but it argues only "
   "against the insurance figure and is silent on the fee-friction evidence, which points the other way. "
   "COMPOUNDS WITH HEADCOUNT: both are execution assumptions, and if either disappoints the other does not "
   "compensate. TRIANGULATED at 6, with the band now evidence-led in both directions.",
   FMT_NUM2, "agent_productivity")
# RE-ANCHORED 2026-08-21 after the client asked for the evidence BEFORE cutting.
# The research says the cut is justified and that the old number was anchored on
# the wrong segment entirely - which the note below had already recorded without
# anyone acting on it.
#
# USD 120 sat inside the GCC WEALTHTECH band of 80-220, and that band is for
# FUNDED ACCOUNTS OVER USD 10,000. Aurumix's customer saves USD 33.60 a month.
# The mass-market comparators are far lower: KSA retail neobank account opening
# at USD 35-70 PAID-ONLY (9-22 blended), and UAE-expat remittance apps at USD
# 6-14 install-to-first-send.
#
# UAE is set at 85 rather than at the 35-70 band, DELIBERATELY ABOVE IT: opening
# a neobank account or installing a remittance app is a one-off act, while this
# product needs a RECURRING commitment, KYC, and six consecutive payments before
# the customer is worth anything. That is a harder sale and should cost more.
# But it is not a USD 10,000 wealthtech account, and it should never have been
# priced like one.
#
# India 20 -> 15. Published cost of acquiring a financial user in India is
# INR 850-1,200 (~USD 10-14, and reportedly 3x its 2020 level); 15 is INR ~1,320,
# still just above the top of that range. Oman & Bahrain has no published figure
# and is scaled off the UAE as before.
CAC_BY_REGION = {"uae": (85, 55, 140), "gulf": (75, 50, 125), "india": (15, 9, 26)}
for key in REGIONS:
    b, a, c = CAC_BY_REGION[key]
    sp("cac_" + key, "Marketing CAC - %s" % RLAB[key], b, a, c, "USD per customer",
       "REGIONALISED 2026-08-20. A single global USD 120 was applied everywhere, which is roughly right for "
       "the UAE and badly wrong for India. PUBLISHED BENCHMARKS: Dubai consumer fintech USD 65-220; retail "
       "neobank account paid-only USD 35-70; GCC wealthtech funded accounts USD 80-220; 'true' fully-loaded "
       "CAC per ACTIVATED customer USD 85-350 against USD 50-100 reported, once KYC, onboarding and "
       "non-activating signups are counted. INDIA IS A DIFFERENT MARKET ENTIRELY: consumer fintech "
       "INR 500-1,500 (~USD 6-18), neobank savings activation INR 1,100-1,800, wealth and broking apps "
       "INR 2,500-5,500 (~USD 30-65); Jupiter reports a USD 5-6 paid CAC. USD 120 is ~INR 10,500 - some 7-20x "
       "the Indian mass-market benchmark and still 2-4x its premium wealth apps. UAE is held at 120 as a "
       "FULLY-LOADED figure; note it sits at the top of the range for a LOW-TICKET product, and the "
       "published low-ticket band is USD 4-70. TRIANGULATED. "
       "THIS IS THE YEAR-1 VALUE. CAC now ramps DOWN to the Y7 parameter below - see that note.",
       FMT_USD, "cac_" + key)
CAC7_BY_REGION = {"uae": (55, 36, 91), "gulf": (45, 30, 75), "india": (10, 6, 17)}
for key in REGIONS:
    b, a, c = CAC7_BY_REGION[key]
    sp("cac7_" + key, "Marketing CAC at Y7 - %s" % RLAB[key], b, a, c, "USD per customer",
       "CLIENT INSTRUCTION 2026-08-26: CAC RAMPS DOWN over the horizon rather than sitting flat. UAE 85->55, "
       "Oman and Bahrain 75->45, India 15->10, interpolated LINEARLY on model year. Aggressive and "
       "Conservative hold the same proportional decline as Base. "
       "WHY IT FALLS: brand recognition compounds, organic and word-of-mouth grow off a larger base, and "
       "paid channels get cheaper per acquisition as creative, targeting and funnel conversion are tuned. "
       "IT CUTS AGAINST THE MODEL'S OWN CAUTION ELSEWHERE - S25's convexity switch says CAC RISES with spend "
       "because you buy progressively worse audiences. Both are real and they pull in opposite directions; "
       "this ramp asserts that the learning effect wins over seven years. "
       "IT IS NOT A COST SAVING - IT IS A REVENUE INCREASE. The marketing budget is fixed by schedule, so a "
       "lower CAC buys MORE CUSTOMERS for the same money. Marketing spend does not fall by a dollar.",
       FMT_USD, "cac7_" + key)
MKT_SHARE = {"uae": 0.74, "gulf": 0.18, "india": 0.08}
for key in REGIONS:
    sp("mktshare_" + key, "Marketing spend share - %s" % RLAB[key], MKT_SHARE[key], MKT_SHARE[key],
       MKT_SHARE[key], "% of budget",
       "RE-BASED 2026-08-26 from 58/16/26, when India moved to an agent-led channel. India falls to 8% - a "
       "SUPPORT budget, not an acquisition engine - and UAE and Oman & Bahrain absorb the residual, split "
       "roughly 80:20 by each region's share of its own reachable ceiling. "
       "THE REASON IS STRUCTURAL, NOT A PREFERENCE. UAE and Oman & Bahrain now have NO agent channel at "
       "all, so marketing plus referral is the WHOLE of their acquisition and the budget has to carry it. "
       "India's 420 agents carry ~57% of its new accounts by Y7, so its budget does not. "
       "WHAT THIS FIXES: holding India at 26% meant a USD 1.65m budget at a USD 15 CAC buying 35,750 "
       "accounts a year, which is what the model did before. India LOOKED agent-led while actually being "
       "carried by a cheap-CAC assumption - marketing was 67% of Indian raw demand at Y7 and agents only "
       "11%, the reverse of how the business is described. "
       "Does not vary by scenario - it is a management decision, not an uncertainty. Must sum to 100%.",
       FMT_PCT, "mktshare_" + key)
sp("referral_rate", "Referral rate", 0.60, 1.10, 0.25, "referrals/customer/yr",
   "RAISED 2026-08-21 from 0.45, client instruction. Cap removed deliberately, so the distribution is "
   "right-skewed - MODEL THE MEAN, NOT THE MEDIAN. 0.60 means the average customer produces roughly three "
   "referrals every five years; the mean is pulled up by a small number of highly social referrers, which is "
   "the normal shape for a trust-based product inside a tight diaspora community. IT IS STILL AN "
   "ASSUMPTION WITH NO EXTERNAL ANCHOR, and it compounds: referrals feed the paying base, which feeds more "
   "referrals. That loop is the reason this parameter is more load-bearing than its size suggests. "
   "ASSUMPTION.", FMT_NUM2, "referral_rate")
sp("referral_conversion", "Referral conversion", 0.62, 0.72, 0.48, "%",
   "The M7 survival of the referred cohort, uplifted ~1.1x. DERIVED.", FMT_PCT, "referral_conversion")
sp("organic_share", "Organic share of direct", 0.25, 0.40, 0.10, "% of direct",
   "RAISED 2026-08-21 from 0.12, client instruction. Kept separate from CAC deliberately, so the acquisition "
   "cost never applies to it - THAT IS ALSO WHY IT IS AN ATTRACTIVE NUMBER TO RAISE, and why it deserves "
   "scepticism. 0.25 says a quarter of paid-driven signups arrive again for free through word of mouth, app "
   "store discovery and community effects. Defensible for a product sold inside a dense diaspora network "
   "where the paid campaign and the community overlap almost completely. NO EXTERNAL ANCHOR AT EITHER "
   "LEVEL, and it is arithmetically identical to a 12% CAC reduction - so raising this AND cutting CAC in "
   "the same pass is close to taking the same benefit twice. ASSUMPTION.", FMT_PCT, "organic_share")
sp("seasonality_amplitude", "Seasonality amplitude", 1.00, 1.40, 0.60, "x deviation from 1.0",
   "Festival TIMING is sourced; how hard a Dubai savings signup responds to Dhanteras is the open question. "
   "Applied to the deviation from 1.0, then renormalised, so the vectors stay at exactly 12.000.",
   FMT_NUM2, "seasonality_amplitude")
_sr[0] += 1

s_section("GROUP B: ICS BENEFIT ENTITLEMENT - no longer gates ACCESS, only BENEFITS")
sp("ever_qualify", "Customers who EVER reach an ICS benefit tier", 0.55, 0.65, 0.45, "% of customers",
   "REPURPOSED 2026-08-26 (CG decision). THIS CELL NO LONGER GATES ACCESS TO ANYTHING. The card and the "
   "credit facility are now open to the whole book regardless of SIP status or ICS tier; what ICS still "
   "governs is BENEFITS - fee discounts, better FX rates, family-wallet pricing, gold reward rebates. "
   "IT FEEDS NO REVENUE ROW TODAY. It is carried as a memo on the Model so the ICS-entitled population is "
   "already computed when the cost and discount build lands, because every one of those benefits is a "
   "give-back that has to be sized against a population. Deleting it would mean rebuilding it a day later. "
   "WHAT IT USED TO DO, AND WHY THE CHANGE IS LARGE: it restricted the card streams to the 55% who ever "
   "cleared a six-payment gate, so removing it widens the card population by 1/0.55 = 1.82x AND removes the "
   "~8-month lag before anyone qualified. The old note called this 'the single most load-bearing parameter "
   "in the model' and claimed the card streams were ~83% of gross profit; BOTH STATEMENTS ARE NOW STALE - "
   "there is no gross-profit line in this build at all, and on the current numbers the card streams are a "
   "far smaller share of revenue than that. "
   "The 55% itself is unchanged and still the OUTPUT of the run-of-6 first-passage engine validated against "
   "the archetype mix, which now lives in the Phase 5 simulation. Whether ICS ENTITLEMENT should use the "
   "same threshold as the old payment gate is an OPEN QUESTION for the mechanism-design update - it is "
   "inherited here, not decided.", FMT_PCT, "ever_qualify")
sp("months_to_qualify", "Average months to reach an ICS benefit tier", 8, 6, 11, "months",
   "REPURPOSED 2026-08-26 with the row above - it no longer delays card access, only benefit entitlement. "
   "Tier arrival is a DISTRIBUTION, not a date: a customer who misses month 4 cannot qualify before month 9. "
   "The validated mean is M8.0, against the naive assumption of M6. UNDER THE OLD DESIGN those two extra "
   "months pushed every card ladder date to the right and were worth real revenue; under the new one they "
   "delay only the point at which a customer starts COSTING money in discounts, which is the opposite sign. "
   "DERIVED.", FMT_NUM, "months_to_qualify")
_sr[0] += 1

s_section("GROUP C: AUM, LEAKAGE AND SPOT")
sp("self_custody_leakage", "Gold moved out of Aurumix's control", 0.06, 0.03, 0.18,
   "% of AUM/yr",
   "HALVED 2026-08-21 from 0.12, client instruction. NOTE IT HAS ZERO EFFECT ON GOLD UNDER CUSTODY - measured, not asserted - because self-custodied metal never left the vault and was never in that measure. It moves ONLY the collateral-eligible base and therefore the card credit limit, worth ~0.2% of Y7 revenue. RELABELLED earlier the same day after the client correctly objected to the old name, 'self-custody leakage'. "
   "THE GOLD DOES NOT LEAVE THE VAULT: a token moving to a customer's own wallet sells nothing and burns "
   "nothing, so calling it an AUM decrease was wrong. What it does leave is Aurumix's COLLATERAL-ELIGIBLE "
   "base - Aurumix cannot foreclose on a token sitting in a private wallet, so that gold can no longer back "
   "a credit limit. In this model AUM has exactly ONE consumer, the card credit limit, so the arithmetic was "
   "always right and only the label was misleading. THAT ONE-CONSUMER PROPERTY IS WHAT MAKES THIS SAFE: if "
   "a later build points anything else at AUM - a custody fee, a reported-AUM headline - this row must be "
   "split in two first. Contrast the redemption line, which IS gold genuinely gone. ASSUMPTION.",
   FMT_PCT, "self_custody_leakage")
sp("redemption_rate", "Redemption rate", 0.06, 0.035, 0.12, "% of AUM/yr",
   "LOWERED 2026-08-21 from 0.08, client instruction - AND THIS ONE MOVES TOWARDS THE EVIDENCE, not away "
   "from it. PAXG turnover of 5.9% is the only comparator that exists, and 8% sat above it for no stated "
   "reason; 6% sits essentially on it. A DIFFERENT EVENT from gold moving out of Aurumix's control, and a "
   "separate line: this is metal actually sold and tokens burned. VARA forbids charging any fee on it. "
   "ASSUMPTION, now aligned to the single available comparator.", FMT_PCT, "redemption_rate")
sp("lapsed_redemption_mult", "Holder redemption multiplier", 1.6, 1.3, 2.8, "x the paying rate",
   "LOWERED 2026-08-21 from 2.2, client instruction. THIS IS THE WEAKEST OF THE THREE DECAY CHANGES. The "
   "LOGIC behind a multiplier above 1.0 is untouched and still holds: customers who stopped paying redeem "
   "faster, because they have no accruing benefit left to protect. Lowering it to 1.6 weakens that argument "
   "without new evidence for the smaller number - it says lapsed customers are stickier than previously "
   "assumed, which is the convenient direction. Because holders become the majority of the book by Y4, this "
   "remains the DOMINANT AUM DECAY TERM from roughly Y4, so the change is not cosmetic. ASSUMPTION, and the "
   "first one to revisit if AUM is challenged.", FMT_NUM2, "lapsed_redemption_mult")
sp("spot_attach_mult", "Spot attach scenario multiplier", 1.00, 1.70, 0.50, "x the regional attach",
   "REGIONALISED 2026-08-21. The attach LEVEL now sits per region on Assumptions (UAE 12%, Oman & Bahrain "
   "10%, India 25%); this dial carries only the uncertainty around it, preserving the old 14/24/7 spread "
   "(1.00 / 1.70 / 0.50). ONE DIAL, NOT THREE TRIPLETS - nine numbers to defend, and a scenario could "
   "otherwise silently invert the regional ordering that affordability establishes. See the regional rows on "
   "Assumptions for the reasoning. ASSUMPTION about spread.", FMT_NUM2, "spot_attach_mult")
sp("spot_ticket_mult", "Spot ticket scenario multiplier", 1.00, 1.35, 0.70, "x the observed ticket",
   "REPLACED the global 'Average spot ticket' on 2026-08-21. The LEVEL is no longer a scenario input: each "
   "region now carries its own OBSERVED ticket on Assumptions (UAE 190 from Botim, India 40 from Augmont, "
   "Oman & Bahrain 145 inferred). This parameter carries only the UNCERTAINTY around those observations, as "
   "a single dial. WHY ONE DIAL AND NOT THREE: three regional triplets would be nine numbers to defend and "
   "would let a scenario silently invert the regional ordering that the sources establish. One multiplier "
   "flexes the level and PRESERVES the ordering. Aggressive 1.35 puts the UAE at ~257, still under the "
   "AED 1,000 mark; conservative 0.70 puts it at ~133, below Botim's own sub-AED-500 majority. ASSUMPTION "
   "about spread, applied to observed levels.", FMT_NUM2, "spot_ticket_mult")
sp("spot_frequency", "Spot frequency", 1.7, 2.4, 1.2, "events/attacher/yr",
   "PROVENANCE CORRECTED 2026-08-21 - THE NUMBER STANDS, ITS OLD CITATION DID NOT. The previous note read "
   "'Botim's 128,000 trades against ~45,000 buyers implies ~1.9/yr'. THE 45,000 IS NOT A BUYER COUNT, IT IS "
   "A TRANSACTION COUNT: Khaleej Times, Nov 2025, says 775,000 users EXPLORED the feature, 'completing over "
   "45,000 transactions'. The 128,000 is also transactions, to Feb 2026. So the old derivation divided "
   "transactions by transactions across two different windows - the result was the growth in cumulative "
   "trades, not a frequency. CONFIRMED NEGATIVE: Botim has never published a unique-buyer count, so no "
   "frequency can be derived from it, and no published frequency exists ANYWHERE for ad-hoc lump gold "
   "purchases. Comparators are 0.85/yr (Augmont, per REGISTERED user, at an average transaction of Rs 331 - "
   "micro-saving, not lumps) and 264/yr (Jar, Rs 10/day auto-roundups) - a 300x spread, neither being this "
   "product. WHAT DOES SUPPORT 1.7 is a cross-check on the COMBINED basis, attach x frequency x ticket, "
   "which is comparable: Aurumix India implies ~Rs 838/yr of gold per paying customer against Augmont's "
   "~Rs 281/yr per registered user. 3x, and defensible - our customers are engaged SIP savers, not dormant "
   "registrations. Behaviourally 1.7 is also about right for festive lumps: Dhanteras, Akshaya Tritiya and a "
   "wedding is roughly two events. ASSUMPTION, cross-checked on the combined basis, NOT directly sourced.",
   FMT_NUM2, "spot_frequency")
# REMOVED 2026-08-21: "Spot-to-SIP conversion", 8%/yr. It was defined and
# REFERENCED NOWHERE - a dead input that implied a population the model does not
# have. Converting spot buyers into SIP customers only means something if
# spot-only buyers EXIST as a separate balance, and in this model they do not:
# spot is a cross-sell to people who are already paying, so the arrow would have
# pointed from SIP customers back to SIP customers.
#
# THE DECISION NOT TO BUILD THAT POPULATION IS ECONOMIC, NOT LAZY. A spot-only
# buyer is worth 1.7 x ticket x 5% a year - USD 16 in the UAE, USD 3.40 in
# India - against a CAC of 120 and 20. That is a 6-8 YEAR PAYBACK ON SPOT
# REVENUE ALONE, so paid spot acquisition is dominated by simply acquiring a SIP
# customer with the same money. A spot funnel is only worth building if the
# buyers arrive at near-zero marginal cost (festive, organic, referral), and its
# value would then sit entirely in a conversion rate that HAS NO SOURCE - a
# confirmed negative, re-checked 2026-08-20. Client decision 2026-08-21: leave
# spot as a cross-sell and let the Phase 5 simulation carry the two-population
# question, which is where heterogeneity belongs under the v3.0 scope decision.
_sr[0] += 1

s_section("GROUP D: CARD")
sp("pm_share", "Programme manager share of interchange", 0.60, 0.36, 0.85, "%",
   "LOWERED 2026-08-21 from 0.72, client instruction - AURUMIX NOW KEEPS 40% RATHER THAN 28%. "
   "0.60 sits comfortably inside the 36-85% observed range and is roughly its midpoint, so the base case "
   "moves from the pessimistic end to a neutral one. It assumes a NEGOTIATED deal rather than a "
   "take-it-or-leave-it one, which is reasonable for a programme with a gold-collateral proposition the "
   "manager cannot source elsewhere - but it IS an assumption about a contract that does not exist yet, "
   "and the only honest way to hold it is to treat 40% as a target to negotiate to, not a rate already "
   "won. BANDS CORRECTED earlier the same day - THEY WERE INVERTED. Aggressive was 0.85 and conservative 0.55, meaning the "
   "OPTIMISTIC scenario handed MORE of the interchange to the programme manager and cut Aurumix's share from "
   "28% to 15%. Every other cost-like parameter in this model improves under Aggressive - redemption, CAC, "
   "self-custody, the holder multiplier - and this one moved the other way. Flipping the scenario made "
   "interchange fall 46% in the good case and rise 61% in the bad one. "
   "Aggressive is now 0.36, THE DOCUMENTED FLOOR, which is the best deal observed anywhere and the right "
   "walk-away to take into a sponsor conversation - at that level Aurumix keeps 64% rather than 28%, and "
   "interchange more than doubles. Conservative is 0.85. "
   "WHY AURUMIX ONLY GETS A SLICE AT ALL: it is not a bank and cannot issue cards, so a programme manager "
   "holds the BIN, the scheme membership, settlement and dispute handling, and takes the majority for it. "
   "That much is normal. 72% is near the BAD END of the 36-85% range, so the base case is deliberately "
   "unfavourable - THIS IS A COMMERCIAL NEGOTIATION, NOT A MODELLING ASSUMPTION, and it is one of the few "
   "places where a better contract directly doubles a revenue line. No UAE/MENA figure is published. "
   "TRIANGULATED.", FMT_PCT, "pm_share")
sp("card_activation", "Facility take-up - customers who take AND use the card", 0.18, 0.30, 0.08,
   "% of card-eligible base",
   "RE-BASED AGAIN 2026-08-26 (CG decision): the denominator changed from GATE-CLEARED customers to the "
   "WHOLE CARD-ELIGIBLE BASE, because access is no longer gated. The RATE holds at 18% on client "
   "instruction, and that decision deserves to be stated plainly rather than buried. "
   "18% WAS CALIBRATED ON A PRE-SELECTED POPULATION. It is Indian gold-loan penetration - under 10% at a "
   "point in time - uplifted BECAUSE the customers it applied to had already proven six consecutive "
   "payments. Removing the gate removes that pre-selection but keeps the uplift, so the rate is now applied "
   "to a materially less committed population than the one it was derived from. The arithmetic is "
   "unambiguous: the base widens 1.82x and the conversion rate does not fall, so the card population rises "
   "1.82x on this change alone. IF CARD REVENUE IS CHALLENGED, THIS IS THE FIRST CELL TO LOOK AT - the "
   "defensible alternative is ~10%, the unadjusted benchmark. "
   "AGAINST THAT: open access plus ICS discounts is a genuinely different proposition from a gated facility, "
   "and there is no benchmark at all for an open gold-collateralised card, so 18% is not refutable either. "
   "It is a judgement, and it should be presented as one. CLIENT DECISION on the level; DERIVED on the "
   "benchmark beneath it. "
   "CARRIED FORWARD FROM 2026-08-20, and still true: THE CARD IS A DRAWDOWN ON THE GOLD-COLLATERALISED "
   "FACILITY, not a salary-repaid revolving card. Aurumix has no balance sheet to lend from - it can only "
   "extend credit against collateral it holds, which is the customer's gold. Using the card IS borrowing, "
   "so card activation and credit take-up are the SAME behaviour; they were once modelled as two different "
   "populations (50% activating a card, 18% taking credit) doing what is in fact one thing. THE OLD 50% "
   "CAME FROM NEOBANK COMPARABLES (PULSE 68.2%, Monzo 68%) WHERE THE CARD IS THE PRODUCT - wrong in kind "
   "for a card that only lets you borrow against your own savings, and worth remembering now that access is "
   "open, because an open card invites exactly that comparison again.", FMT_PCT, "card_activation")
sp("card_txns_per_draw", "Transactions per drawdown event", 4, 6, 2, "transactions/draw",
   "RE-BASED 2026-08-20 from 12 transactions/MONTH once the card became a drawdown on the gold facility. "
   "Under that model the customer borrows a lump (limit x drawn share) a couple of times a year and spends "
   "it down, so the meaningful unit is transactions PER DRAW, not per month. At 12/month the implied ticket "
   "collapsed to ~AED 12, which is a coffee, not a reason to pledge your gold. THE AVERAGE TICKET FALLS OUT "
   "AS drawn amount / this number, and the draw events cancel - so the ticket depends only on how large a "
   "draw is and how many purchases it is spent across. IT ALSO CHANGES WHAT THE CARD IS FOR: not daily "
   "groceries, but occasional larger needs - school fees, a medical bill, an emergency - funded by borrowing "
   "against savings rather than liquidating them. That is a materially different product story and should be "
   "put to the client in those words. ASSUMPTION.", FMT_NUM, "card_txns_per_draw")
# REMOVED 2026-08-21: "Card transactions per active card per month", 12/mo.
# SUPERSEDED when the card became a drawdown on the gold facility - the
# meaningful unit became transactions PER DRAW (above), and this row was left
# behind, still scenario-flexed and read by nothing. At 12/month the implied
# ticket collapsed to ~AED 12, which is a coffee, not a reason to pledge your
# gold. The Visa anchor it carried (~175 txns/card/yr regionally, 14-15/month)
# is retained in the note on card_txns_per_draw, which is where it now bears.
sp("foreign_spend_share", "Foreign spend share (mean)", 0.34, 0.45, 0.24, "% of card spend",
   "Applied through the seasonal vector, not as a constant. ASSUMPTION.", FMT_PCT, "foreign_spend_share")
sp("reissue_rate", "Card REISSUE rate (excludes first issue)", 0.06, 0.04, 0.10, "reissues/card/yr",
   "CORRECTED 2026-08-26, and this replaces 'Card issuance events' at 1.06/card/yr. "
   "THE OLD ROW WAS A REAL DEFECT, not a re-cut. 1.06 events per card per YEAR was applied to the STOCK of "
   "active cards, which charged the AED 75 issuance fee to every existing cardholder every single year - an "
   "ANNUAL FEE, not an issuance fee, and the row's own note said so: '1.00 at activation plus reissues'. The "
   "1.00 belongs to the year a card is ACTIVATED and must be charged on NEW cards; only the 0.06 residual is "
   "a recurring per-card event. "
   "WHAT IT WAS WORTH. At Y7 it charged ~33,980 issuance events against ~9,100 real ones, roughly 3.7x too "
   "many, or ~USD 507,000 - about 11% of total revenue. It predates the 2026-08-26 access change, but that "
   "change multiplied the card stock by 1.82x and so multiplied the error with it, taking issuance from "
   "~6% of revenue to ~15% before this fix. IT WAS THE LARGEST SINGLE ERROR IN THE MODEL. "
   "0.06 is what remains once first issue is removed: a tier upgrade forces a physical reissue, and under "
   "the ICS design tier changes are the main trigger. ASSUMPTION. "
   "STILL NOT IMPLEMENTED, and it belongs to the discount build: the fee input's own note says issuance is "
   "'waived at upper tiers'. Nothing waives it. Under the new design that waiver is exactly the kind of ICS "
   "benefit being priced tomorrow, so wire it there rather than here.",
   FMT_NUM2, "reissue_rate")
sp("replacement_events", "Card replacement events", 0.11, 0.07, 0.18, "events/card/yr",
   "Industry-normal loss/theft/damage is 8-15% annually. ASSUMPTION.", FMT_NUM2, "replacement_events")
ATM = [("a500", "AED 0-500", 0.60, 0.42, 0.74, 250, 300, 200),
       ("a1500", "AED 500-1,500", 0.25, 0.30, 0.18, 1000, 1100, 900),
       ("a3000", "AED 1,500-3,000", 0.12, 0.20, 0.06, 2250, 2400, 2100),
       ("a3000p", "AED 3,000+", 0.03, 0.08, 0.02, 4000, 4500, 3600)]
for key, nm, b, a, c, mb, ma, mc in ATM:
    sp("atm_" + key, "ATM %s - share of cardholders" % nm, b, a, c, "%",
       "KEPT AS A DISTRIBUTION DELIBERATELY. The mean draw (~AED 940) sits just BELOW the free allowance of "
       "AED 1,000 BY DESIGN, so applying the fee to the mean returns EXACTLY ZERO and stream 4 silently loses "
       "a component. The distribution returns materially more, generated almost entirely by a small high-cash "
       "tail. ASSUMPTION.", FMT_PCT, "atm_" + key)
for key, nm, b, a, c, mb, ma, mc in ATM:
    sp("atmm_" + key, "ATM %s - midpoint draw" % nm, mb, ma, mc, "AED/month",
       "Raising the allowance at the top tier waives revenue from only the top ~3% of cardholders.",
       FMT_NUM, "atmm_" + key)
_sr[0] += 1

s_section("GROUP E: CREDIT, B2B AND FAMILY")
# Credit take-up is GONE - merged into facility take-up in Group D, because
# under the gold-drawdown model taking a card and taking credit are one act.
sp("drawn_share", "Drawn as % of permitted limit", 0.50, 0.70, 0.30, "%",
   "Revolving facilities draw 40-55% of permitted. This now drives CARD SPEND as well as the loan balance, "
   "because the card is the drawdown mechanism. TRIANGULATED.", FMT_PCT, "drawn_share")
sp("facility_turnover", "Facility turnover, peak -> average", 0.42, 0.55, 0.30, "x peak drawn",
   "AN ARITHMETIC CORRECTION, not a sensitivity, derived from Manappuram's 71-day realised tenor. It cuts "
   "the lending stream by 1.88x. DERIVED.", FMT_NUM2, "facility_turnover")
sp("draw_events", "Draw events per borrower per year", 2.1, 3.2, 1.3, "events/yr",
   "MOVES WITH facility turnover - do not flex independently. DERIVED.", FMT_NUM2, "draw_events")
sp("family_attach", "Family plan attach rate", 0.15, 0.25, 0.08, "% of NEW customers",
   "SET AT 0.15 on 2026-08-21, client instruction - the same day it came DOWN from 0.20 to 0.08, so it has "
   "landed close to where it started. What has genuinely changed is not the number but the STRUCTURE. THE UNIT ALSO CHANGED, AND THAT MATTERS: it is now the share of NEW "
   "customers who take the plan when they join, feeding a SUBSCRIBER BALANCE that then churns - not a "
   "standing share of the whole book that could never fall. The steady-state share of paying customers "
   "holding a plan therefore lands BELOW this number. "
   "NOTHING IS STATED ANYWHERE IN THE CORPUS and no attach benchmark exists for this kind of add-on; 20% "
   "is at the top of the normal range for an OPTIONAL PAID extra on a savings product, which is low single "
   "digits to low teens. IT IS NOT EQUIVALENT TO THE OLD 20% AND SHOULD NOT BE READ AS A PARTIAL "
   "REVERSAL: 20% was a permanent share of the whole book that could never decay, while 15% applies once, "
   "on joining, to a balance that then churns at ~7.1% a month. The standing share of paying customers "
   "holding a plan settles near 10.5%, not 15%. NOTHING IS STATED ANYWHERE IN THE CORPUS and no attach "
   "benchmark exists for this kind of add-on. ASSUMPTION.", FMT_PCT, "family_attach")
sp("family_churn", "Family plan cancellation rate (incremental)", 0.25, 0.15, 0.40, "%/yr",
   "ADDED 2026-08-21. INCREMENTAL IS THE KEY WORD: this is cancellation by customers who STAY WITH AURUMIX. "
   "Subscribers already disappear when they stop paying their SIP, and that was the only churn stream 3 had "
   "- which implicitly assumed nobody ever cancels the plan while remaining a customer. "
   "NO PRODUCT-SPECIFIC DATA EXISTS - not for digital will memberships, not for finance-app family plans; "
   "Trust & Will, Farewill and Epilogue publish none. The only benchmark is generic mobile subscription "
   "apps: above 10%/month poor, 6-10%/month average, under 6%/month good. 25%/yr is ~2.4%/month, INSIDE the "
   "'good' band, which is the right place for an add-on bundled into an ongoing savings relationship rather "
   "than a standalone app fighting for attention every month. ASSUMPTION on a generic benchmark.",
   FMT_PCT, "family_churn")
sp("benef_count", "Beneficiaries named per plan", 2.5, 3.5, 1.6, "beneficiaries",
   "ADDED 2026-08-21. THE FIRST IS INCLUDED IN THE PLAN PRICE, so only the excess over 1.0 is charged - at "
   "2.5 that is 1.5 chargeable. "
   "No platform publishes beneficiaries-per-will. What anchors it is who these customers are: South Asian "
   "migrant workers who remit ~80% of income to families at home, typically supporting a spouse, children "
   "and parents. Household sizes in the source countries run ~4.4 in India, ~4.5 in Bangladesh and ~6.8 in "
   "Pakistan, so naming 2-3 beneficiaries is conservative against the family they actually support. "
   "NOT THE SAME AS HOUSEHOLD SIZE - a beneficiary is a named, identity-verified person on a legal "
   "instrument, and people name fewer than they support. ASSUMPTION anchored on demographics.",
   FMT_NUM2, "benef_count")
s_derived("family_churn_m", "Family plan monthly churn, combined (derived)",
          "=1-(1-%s)*(1-%s)^(1/12)" % (sref("monthly_churn"), sref("family_churn")),
          "%/month",
          "TWO WAYS TO LOSE A SUBSCRIBER, COMBINED MULTIPLICATIVELY, NOT ADDED: they leave Aurumix entirely "
          "(the SIP churn on the left) or they stay and cancel the plan (the incremental rate on the right). "
          "Adding the two would double-count the overlap - a customer cannot both leave and separately "
          "cancel. DERIVED.", FMT_PCT2, "family_churn_m")
# ---- B2B REBUILT 2026-08-21 -------------------------------------------------
# Was 6 partners x a flat USD 32m "mature AUM per partner". Two problems, and
# they pointed in OPPOSITE directions, which is why arguing about the count
# alone could never have fixed it:
#
#   - THE COUNT WAS TOO LOW. Research found ~20 candidate wallets, digital banks
#     and exchange houses across the UAE, Oman and Bahrain, and EXACTLY ONE OF
#     THEM ALREADY OFFERS GOLD (Botim, via O Gold). The Gulf slot is essentially
#     empty.
#   - THE AUM PER PARTNER WAS TOO HIGH. A flat 32m asserts every marginal
#     partner is Botim-scale. Botim has 8.2-8.5m UAE users and 1.7m ACTIVE
#     fintech users; the next Gulf wallet down has ~1.5m subscribers and the
#     ones after that are far smaller. Six Botim-scale wallets do not exist.
#
# THE FIX IS TO STOP ASSERTING THE AUM AND DERIVE IT, so the only thing left to
# argue about is the partner count - which is the argument we can actually win,
# because it rests on a nameable candidate list.
#
# INDIA IS EXCLUDED FROM THE B2B PARTNER POOL ENTIRELY (client decision).
# Every large Indian app that wants gold already has one of three entrenched
# providers on a 2-5 year contract: PhonePe (450-500m users) and MobiKwik,
# Amazon Pay and Jar run SafeGold; Paytm (300-350m) and Google Pay run
# MMTC-PAMP; Groww runs Augmont. That is displacement, not greenfield, into a
# 3-6% spread that compresses hard for large partners. AND ZERODHA EXITED THE
# CATEGORY OUTRIGHT on regulatory grounds - digital gold is not SEBI-regulated -
# which is a second India regulatory problem on top of FEMA/LRS. The partner
# pool here is GULF ONLY.
for y in range(1, 8):
    b, a, c = ([0, 1, 3, 5, 7, 9, 11][y - 1], [0, 2, 5, 8, 11, 14, 17][y - 1],
               [0, 1, 1, 2, 3, 4, 5][y - 1])
    sp("partners_y%d" % y, "B2B partners - Y%d" % y, b, a, c, "partners",
       "RAISED 2026-08-21. Base now reaches 11 by Y7, against ~20 nameable Gulf candidates without a gold "
       "product - roughly a 55% win rate on the addressable partner market, which is aggressive but "
       "arguable for a category with one incumbent. Partner 1 signs in Y2, matching the M13 B2B go-live. "
       "THE CANDIDATE LIST IS THE JUSTIFICATION AND IT IS FINITE: e& money, Payit, Careem Pay, Al Ansari, "
       "LuLu Money, Rise, NOW Money, Wio, Liv, Mashreq Neo, Klip, PayBy, Jingle Pay, Thawani, BenefitPay, "
       "CWallet, Amanat and a handful of banks and exchange houses. Going much above this count means "
       "claiming partners that have not been named. ASSUMPTION on a sourced candidate set.",
       FMT_NUM, "partners_y%d" % y)
sp("b2b_fee", "B2B platform fee", 0.0075, 0.0090, 0.0050, "% of partner AUM/yr",
   "RAISED 2026-08-21 from 0.625%, client instruction, and PROMOTED FROM A HARDCODED ASSUMPTION TO A "
   "SCENARIO PARAMETER at the same time - it drives the largest stream in the model and had no band at all, "
   "which is the wrong place to be certain. "
   "0.75% IS THE TOP OF THE CITED 0.5-0.75% BAND, NOT THE MIDDLE OF IT. That has a consequence worth stating "
   "plainly: the aggressive case at 0.90% NOW SITS OUTSIDE THE EVIDENCE, because there is no headroom left "
   "inside the band. Conservative at 0.50% is the band's floor. So the downside is sourced and the upside is "
   "not - the reverse of how a scenario range should usually be built, and the direct result of putting the "
   "base at the ceiling. "
   "The fee is COUPLED TO THE PARTNER SET: a platform fee is what a partner will pay for gold "
   "infrastructure it does not have to build, and the more bargaining power the partner has the lower it "
   "goes. Indian precedent is the warning - the three incumbent providers work on a 3-6% SPREAD that "
   "compresses hard for large partners, so a partner the size of Botim would push back on 0.75% harder than "
   "a small wallet would. Raising the partner count and the fee together assumes neither effect bites. "
   "ASSUMPTION at the ceiling of a cited band.", FMT_PCT2, "b2b_fee")
sp("partner_users", "Average partner user base", 900000, 1400000, 500000, "active users",
   "THE MIX, NOT THE BIGGEST. A realistic six-partner Gulf book is roughly one Botim-scale (1.7m ACTIVE "
   "fintech users, of 8.2-8.5m app users), two at e& money scale (~1.5m subscribers), and three small "
   "wallets at a few hundred thousand - which averages near 900k. USE ACTIVE USERS, NOT REGISTERED: "
   "Botim's own numbers separate 8.5m app users, 3m KYC'd and 1.7m active fintech users, and adopting the "
   "headline would overstate this by 5x. CITED for the two largest, ASSUMPTION for the tail.",
   FMT_NUM, "partner_users")
sp("partner_adoption", "Partner users adopting gold (mature)", 0.06, 0.10, 0.03, "% of partner users",
   "OBSERVED ANCHOR: O Gold reports 75,000 ACTIVE users against Botim's 1.7m active fintech users - 4.4% - "
   "and 775,000 who merely explored the feature. Base is set at 6%, ABOVE the observed 4.4%, because that "
   "figure is roughly six months after the gold feature launched and adoption should mature; 6% is a 1.4x "
   "maturation, not a doubling. THE 4.4% IS THE ONLY OBSERVED GOLD-ADOPTION RATE FOR A GULF WALLET THAT "
   "EXISTS. TRIANGULATED.", FMT_PCT, "partner_adoption")
sp("aum_per_partner_user", "AUM per adopting partner user", 350, 500, 220, "USD",
   "DERIVED FROM BOTIM'S OWN DISCLOSURES AND CROSS-CHECKED AGAINST OUR OWN BOOK - two independent routes "
   "to nearly the same number, which is why this is the firmest input in the B2B block. Botim: AED 100m+ "
   "of gold across 128,000 trades since Aug 2025, against 75,000 active O Gold users, implies ~USD 363 per "
   "active user. Aurumix's own direct book runs at ~USD 302 of gold per customer. A partner's users are "
   "buying the same product with the same wallets. DERIVED.", FMT_USD, "aum_per_partner_user")
s_derived("aum_per_partner", "AUM per partner (derived)",
          "=%s*%s*%s" % (sref("partner_users"), sref("partner_adoption"),
                         sref("aum_per_partner_user")), "USD",
          "NO LONGER A TYPED NUMBER. Was a flat USD 32m asserted as 'mature AUM per partner' with no "
          "derivation. Now users x adoption x AUM per adopting user, every term of which has an external "
          "anchor. THE POINT OF THE CHANGE: the only remaining argument is the PARTNER COUNT, which rests "
          "on a nameable candidate list and is the argument that can actually be won. DERIVED.",
          FMT_USD, "aum_per_partner")
_sr[0] += 1

s_section("GROUP F: REGIONS")
RDATA = [("uae", 33.6, 41.0, 26.5, 0), ("gulf", 26, 32, 21, 0), ("india", 30, 36, 24, 0)]
for key, tb, ta, tc, mix in RDATA:
    sp("ticket_" + key, "Average monthly ticket - %s" % RLAB[key], tb, ta, tc, "USD/month",
       "THE SIP AMOUNT - the recurring monthly gold contribution. SAVINGS-CAPACITY anchored, not remittance "
       "anchored, because remittance is a committed obligation rather than discretionary money; a recurring "
       "gold debit realistically captures 10-15% of monthly savings capacity. Anchors: Joyalukkas has "
       "PRICE-DISCOVERED AED 100 (USD 27) as the viable mass-market monthly instalment for this exact "
       "demographic, and Malabar sits at AED 200. The blend lands BELOW AMFI's USD 34 Indian retail SIP "
       "average (Rs 31,961 crore over 10.63 crore accounts, current at July 2026), which is the correct "
       "direction since AMFI reflects domestic urban investors with more discretionary income than a Gulf "
       "blue-collar book - it is an upper bound, not a proxy. UAE is the ceiling-weighted blend of the former "
       "Indian (USD 38) and other South Asian (USD 26) rows. Book-weighted average ~USD 31.45. TRIANGULATED.",
       FMT_USD, "ticket_" + key)
# REMOVED 2026-08-21: "Share of new customers - <region>", derived as each
# region's share of the addressable ceiling.
#
# IT WAS NOT MERELY UNUSED, IT WAS MISLEADING. The row asserted that a
# region's share of acquisition equals its share of the ceiling. That stopped
# being true when acquisition went region-by-region: new customers now fall out
# of each market's OWN marketing budget, OWN CAC, OWN salesforce share and OWN
# referring base, so the regional split is an EMERGENT RESULT, not an input. A
# reader would have taken a stale derived row for the driver. Its note also
# claimed the shares were "RENORMALISED on the Model for regions not yet open",
# which stopped being true when that renormalisation was rewired to the
# salesforce split. Found by the orphan check added the same day.
_sr[0] += 1

section(ws_scen, _sr[0], "STRUCTURAL SWITCHES - the two that move revenue", span=5)
headers(ws_scen, _sr[0] + 1, [("A", "Switch"), ("B", "Active (1/0)"), ("C", "State"), ("D", "ON value")])
_sr[0] += 2
SWROW = {}
s_section("GROUP G: THE COST BASE")
sp("storage_rate", "Vault storage fee", 0.00120, 0.00026, 0.00400, "%/yr of metal held",
   "WHAT AURUMIX PAYS THE VAULT, not what it charges customers - decision 42 makes retail storage free. "
   "BASE 0.12% IS BULLIONVAULT'S PUBLISHED TARIFF, insurance included, minimum USD 4/month "
   "(bullionvault.com/help/tariff.html, retrieved 2026-08-26). Used as an UPPER BOUND on wholesale cost "
   "because it is a RETAIL price and therefore already carries BullionVault's own margin. "
   "AGGRESSIVE 0.026% is the DGCX Spot Gold vaulting tariff read literally - USD 0.10 per kilo per day at "
   "today's gold price. IT IS CARRIED ONLY AS THE OPTIMISTIC CASE AND SHOULD NOT BE THE BASE: that notice "
   "is dated 7 DECEMBER 2015, it is a DELIVERY-CYCLE tariff for taking metal off a futures contract (three "
   "free days per cycle) rather than a custody contract, and it is a FIXED DOLLAR-PER-KILO charge set when "
   "gold was about USD 34/g. The same physical rate was 0.107%/yr in 2015 and only reads as 0.026% now "
   "because the metal quadrupled. No vault reprices that slowly. "
   "CONSERVATIVE 0.40% is the low end of the range secondary sources quote for Dubai institutional storage. "
   "For reference the Royal Mint publishes 1% + VAT for bar storage. "
   "BENCHMARKED, NOT OBSERVED: no Dubai operator publishes an institutional custody tariff. Brink's runs the "
   "DGCX-approved Dubai vaults (Gold Souk, DAFZ, Almas Tower) and quotes privately. GET A QUOTE - it is the "
   "same conversation as naming the dealer.",
   FMT_PCT3, "storage_rate")
sp("storage_min_day", "Vault minimum charge", 25.0, 5.0, 65.0, "USD/day",
   "Professional vault contracts carry a minimum invoice, and it BITES WHILE THE BOOK IS SMALL - which is "
   "most of this horizon. BASE USD 25/day is the only MINIMUM published by any Dubai-approved vault "
   "operator (DGCX Spot Gold vaulting notice, 7 December 2015). AGGRESSIVE USD 5/day is nearer "
   "BullionVault's USD 4/MONTH. CONSERVATIVE USD 65/day is roughly USD 2,000/month, a plausible "
   "institutional floor. AT BASE THE MINIMUM BINDS UNTIL ABOUT YEAR 3, which is the point: early-year unit "
   "economics look worse than steady state because of minimum-commitment structures, and this is one of "
   "three in the cost base - the others are the KYC platform minimum and the card programme minimums. "
   "ASSUMPTION. Replace with the Brink's quote.",
   FMT_NUM2, "storage_min_day")

sp("vara_supervision_aed", "VARA annual supervision fee", 400000, 200000, 600000, "AED/yr",
   "VARA Schedule 2 - Supervision and Authorisation Fees, Virtual Assets and Related Activities Regulations "
   "2023 (rulebooks.vara.ae, retrieved 2026-08-26). PUBLISHED AND EXACT: AED 200,000 per year for each of "
   "Category 1 VA Issuance, Broker-Dealer, Exchange and Custody; AED 80,000 for VA Transfer and Settlement. "
   "Payable IN ADVANCE of conducting the activity. "
   "VARA LICENSES ACTIVITIES, NOT THE TOKEN - you need each licence only if you actually perform that "
   "activity, so the ladder is a BUSINESS-MODEL question, not a risk band. "
   "BASE AED 400,000 = ISSUANCE + BROKER-DEALER. Issuance is mandatory for a gold-backed ARVA and is not "
   "negotiable. Broker-Dealer is added because Aurumix SELLS DIRECTLY TO RETAIL through its own app and "
   "agent network, and the definition catches 'soliciting or accepting orders for Virtual Assets and "
   "accepting fiat currency for such orders' - which is the SIP exactly. There is NO CARVE-OUT letting an "
   "issuer distribute its own token without it; the footnote checked on 2026-08-26 only requires issuers to "
   "comply with the Issuance Rulebook. "
   "AGGRESSIVE AED 200,000 = ISSUANCE ALONE, and it is a REAL STRATEGIC OPTION rather than optimism: mint "
   "and let a VARA-licensed broker-dealer handle distribution. Saves USD 54,459/yr of supervision and "
   "AED 600,000 of locked capital, at the cost of margin and control. "
   "CONSERVATIVE AED 600,000 adds CUSTODY, which bites only if Aurumix safekeeps client VAs - and only "
   "qualifies at all if each client's assets sit in SEPARATE WALLETS, so it turns on the wallet design. "
   "EXCHANGE IS DELIBERATELY EXCLUDED: no trading venue, no order matching. VA TRANSFER AND SETTLEMENT TOO: "
   "that definition targets moving CLIENTS' assets as a rail, not delivering your own issue to its buyer. "
   "VERIFIED AGAINST VARA'S PUBLIC REGISTER 2026-08-26, and the register settles it. Of 56 licensed VASPs, "
   "EXACTLY TWO hold Category 1 VA Issuance - Ctrl Alt Solutions DMCC and Tokinvest DMCC - and BOTH ALSO "
   "HOLD BROKER-DEALER SERVICES. Neither holds issuance alone. Issuance + Broker-Dealer is the only observed "
   "pattern in Dubai, which is why it is the Base. Both are tokenisation platforms rather than exchanges, so "
   "they are the closest comparators Aurumix has. "
   "THE SAMPLE IS TWO COMPANIES - it is the entire population of issuance licensees, not a survey, and it "
   "should be quoted as such. "
   "NEITHER HOLDS CUSTODY, and that is informative: on the register custody sits with specialists - Ceffu, "
   "BitGo, Hex Trust, Komainu, Bitpanda, Zand Bank - not with issuers. So Conservative is really the "
   "'Aurumix builds its own custody instead of partnering' case, and it has no Dubai precedent. "
   "AGGRESSIVE HAS NO DUBAI PRECEDENT EITHER: no licensee holds issuance alone, so outsourcing distribution "
   "is a strategy nobody local has taken, not merely a cheaper option. "
   "Schedule 2 also lets VARA impose additional fees at its sole discretion.",
   FMT_NUM, "vara_supervision_aed")
sp("vara_application_aed", "VARA licence application fee (one-off)", 150000, 100000, 200000, "AED",
   "VARA Schedule 2, same source. AED 100,000 for one regulated VA Activity, payable at SUBMISSION - the "
   "application is not processed until it is received, so it is spent whether or not the licence is granted. "
   "Each additional activity adds a Licence Extension Fee of 50% OF THE LOWER application fee. Base 150,000 "
   "= Issuance 100,000 + 50% of Broker-Dealer's 100,000. Conservative 200,000 adds Custody at another "
   "50,000. "
   "VARA PUBLISHES NO APPROVAL TIMELINE ANYWHERE - never give the client a date.",
   FMT_NUM, "vara_application_aed")
sp("dmcc_annual_aed", "DMCC company licence", 20265, 20265, 20265, "AED/yr",
   "DMCC published Schedule of Charges (dmcc.ae, retrieved 2026-08-26): standard trading and service licence "
   "AED 20,265/yr. THE COST OF EXISTING AS A COMPANY, separate from and additional to the VARA licence, "
   "which is the cost of doing virtual asset business. You pay both. "
   "THE ESTABLISHMENT CARD (AED 1,825/yr) WAS REMOVED FROM THIS LINE ON 2026-08-26. It exists to sponsor "
   "residence visas, so its driver is headcount, not regulation - it belongs with visas and office, where a "
   "flexi-desk runs AED 16,000-19,000 and an employment residence permit about AED 2,972 each. "
   "DMCC states charges are subject to change without notice and publishes no effective date. "
   "NO SCENARIO VARIATION: it is a published tariff, not an estimate.",
   FMT_NUM, "dmcc_annual_aed")
sp("dmcc_setup_aed", "DMCC incorporation (one-off)", 12035, 12035, 12035, "AED",
   "DMCC published Schedule of Charges: application AED 1,015 + new company registration AED 9,000 + "
   "articles of association AED 2,020. One-off, Year 1.",
   FMT_NUM, "dmcc_setup_aed")
sp("kyc_per_check", "KYC and AML - per verification", 1.85, 1.35, 1.85, "USD/check",
   "Sumsub published pricing (sumsub.com/pricing, retrieved 2026-08-26). COMPLIANCE TIER USD 1.85 IS THE "
   "RIGHT ONE, not the USD 1.35 Basic tier: Basic is sold for NON-REGULATED businesses and excludes AML "
   "screening, ongoing AML monitoring and proof-of-address, all three of which a VARA-licensed issuer needs. "
   "Aggressive carries 1.35 only to show what the cheaper tier would be worth. Sumsub charges for SUCCESSFUL "
   "verifications only.",
   FMT_USD2, "kyc_per_check")
sp("paidup_issuance_aed", "Paid-up capital - Category 1 VA Issuance", 1500000, 1500000, 1500000, "AED",
   "VARA Virtual Asset Issuance Rulebook, ARVA Rules G.1 (retrieved 2026-08-26): Paid-Up Capital of at least "
   "THE HIGHER OF (a) AED 1,500,000 and (b) 2% of the 24-month average value of Reserve Assets, WHERE "
   "APPLICABLE. "
   "UNDER OPTION A (DIRECT OWNERSHIP) LIMB (b) DOES NOT APPLY, because Rule III.C.1 attaches Reserve Assets "
   "only to ARVAs which purport to maintain a stable value in respect of a Referenced Asset - and Aurumix's "
   "customers own specific allocated grams rather than a claim on a stabilised pool. There are no Reserve "
   "Assets to take 2% of. "
   "THE FLOOR IS NOT CONDITIONAL. Choosing Option A removes the ESCALATOR, not the AED 1,500,000. The rule "
   "reads at least the higher of, so with (b) inapplicable (a) governs. Flat in every scenario for that "
   "reason. LOCKED, NOT SPENT: balance sheet only, never the P&L.",
   FMT_NUM, "paidup_issuance_aed")
sp("paidup_activity_aed", "Paid-up capital - licensed activities", 600000, 0, 1200000, "AED",
   "VARA Company Rulebook Part VI.B, per licensed VA Activity: Broker-Dealer without custody is the HIGHER "
   "of AED 600,000 or 25% of fixed annual overheads; Custody likewise. "
   "IT IS CUMULATIVE, NOT THE HIGHEST SINGLE TEST. Part VI.A requires the capital for EACH VA Activity for "
   "which the VASP is Licensed. Verified 2026-08-26. "
   "TRACKS THE ACTIVITY LADDER on the supervision-fee row: Aggressive 0 (issuance alone, distribution "
   "outsourced), Base 600,000 (Broker-Dealer), Conservative 1,200,000 (Broker-Dealer plus Custody). "
   "NOTHING TO DO WITH THE ARVA OPTION A / OPTION B CHOICE - this attaches to the LICENCE, not to the token "
   "design.",
   FMT_NUM, "paidup_activity_aed")
sp("nla_months", "Net liquid assets - months of opex", 1.2, 1.2, 1.2, "x monthly opex",
   "VARA Company Rulebook Part VI.C: Net Liquid Assets at least 1.2 x monthly operating expenses. A "
   "CONTINUOUS LIQUIDITY TEST, not a one-off subscription, and it appears NOWHERE IN THE CORPUS - added "
   "2026-08-26. NOT ADDITIVE TO PAID-UP CAPITAL: the same cash can satisfy both, so the two are shown side "
   "by side and never summed. It constrains the FORM the capital is held in, not the amount.",
   FMT_NUM2, "nla_months")
sp("insurance_usd", "Insurance - PI, D&O and crime", 45000, 20000, 90000, "USD/yr",
   "MANDATORY, and the rulebook names the policies. VARA Company Rulebook Part VI.D requires professional "
   "indemnity, directors' and officers', and COMMERCIAL CRIME COVER FOR VIRTUAL ASSETS IN HOT WALLETS, all "
   "held with a regulated insurer, in an amount 'adequate to the size and complexity of the business'. VARA "
   "sets no minimum sum and judges adequacy at licensing. Verified 2026-08-26. "
   "BENCHMARKED, NOT OBSERVED - no insurer publishes VASP rates. Base USD 45,000 builds from UAE broker rate "
   "guides: PI for a regulated financial firm at AED 10m limit runs AED 60,000-155,000/yr and D&O at the "
   "same limit AED 55,000-145,000/yr; Base takes the BOTTOM of both ranges (about USD 31,300 combined) plus "
   "about USD 13,700 for crime cover. Aggressive assumes SME limits (PI AED 22,000, D&O AED 14,000); "
   "Conservative takes the top of the regulated ranges. "
   "THE CRIME ELEMENT IS THE UNPRICED ONE and it is the hardest to place: S&P reports hot and warm wallet "
   "crime capacity at roughly USD 75m per risk against USD 900m for cold specie, and a CFTC presentation "
   "calls hot-wallet cover 'extremely limited', layered in USD 1-5m blocks. Aurumix's hot-wallet exposure is "
   "operational treasury only, so a low limit should suffice - but the market is thin and the price is not "
   "published. "
   "THE PHYSICAL GOLD IS NOT INSURED HERE. Vault storage at 0.12% already includes insurance on the metal, "
   "so covering it again would double-count. "
   "FLAT ACROSS THE HORIZON, which is a simplification: PI and D&O scale with turnover, so this should be "
   "re-quoted as the book grows. GET A BROKER QUOTE.",
   FMT_USD, "insurance_usd")
sp("audit_usd", "Audit and reserve attestation", 25000, 12000, 60000, "USD/yr",
   "TWO SEPARATE ENGAGEMENTS, both mandatory under VARA Issuance Rulebook Rule III.D (verified 2026-08-26): "
   "a SIX-MONTHLY independent audit, and an ANNUAL independent audit of financial statements. A named "
   "independent auditor must be appointed and notified to VARA, and Senior Management must submit an "
   "ATTESTATION to the accuracy of each audit. "
   "OPTION A NARROWS THE SIX-MONTHLY ONE. Rule III.D.2.a covers (a) the number and value of ARVAs in "
   "circulation and (b) the composition and value of Reserve Assets 'IF APPLICABLE'. Under direct ownership "
   "there are no Reserve Assets, so limb (b) drops out and the engagement audits token supply against gold "
   "held rather than a reserve pool. NARROWER SCOPE SHOULD PRICE LOWER than the brief assumes. "
   "THE ATTESTATION FEE IS GENUINELY UNPRICED. No issuer, auditor or regulator has ever published one. "
   "Paxos uses Withum and now KPMG, Tether uses BDO Italia, Circle uses Deloitte - all disclose scope, "
   "standard and frequency, none disclose fee. The only figure found anywhere is a secondary estimate of "
   "USD 2-12m/yr for LARGE issuers using Big Four, which is not this business. "
   "BASE USD 25,000 = about USD 5,400 for the statutory audit (UAE market guides put a regulated SME at "
   "AED 12,000-25,000/yr) plus about USD 20,000 for two six-monthly engagements at USD 10,000 each. THE "
   "SECOND HALF IS A PLACEHOLDER, NOT A QUOTE. Get one from Bureau Veritas or an equivalent.",
   FMT_USD, "audit_usd")
sp("techaudit_usd", "Technology audit and penetration testing", 15000, 6000, 35000, "USD/yr",
   "MANDATORY AND RECURRING, not a one-off. VARA Technology and Information Rulebook, E. Testing and Audit, "
   "verbatim: VASPs 'must engage a qualified and independent third-party auditor to conduct vulnerability "
   "assessments and penetration testing' at least ANNUALLY 'and prior to the introduction of any new "
   "systems, applications and products'. Verified 2026-08-26. "
   "BENCHMARKED from published penetration-testing price guides: small financial firms run USD 5,000-15,000 "
   "per engagement, broader multi-asset engagements covering web, APIs and internal network USD 10,000-35,000. "
   "Base takes the top of the small-firm range. "
   "THE PER-LAUNCH TRIGGER IS NOT MODELLED. Every new system or product needs its own test, so a year with "
   "several launches costs more than this line shows. It is a step cost tied to the product roadmap, and "
   "the roadmap does not exist yet.",
   FMT_USD, "techaudit_usd")
sp("launch_audit_usd", "Launch technology and smart contract audit (one-off)", 40000, 20000, 75000, "USD",
   "THE PRE-LAUNCH ENGAGEMENT the Technology Rulebook requires before any new system goes live, covering the "
   "token contract as well as the platform. BENCHMARKED, not quoted: published penetration-test ranges top "
   "out around USD 35,000 for a broad engagement, and smart-contract audit is a specialist skill that "
   "prices above general application testing, so Base sits at USD 40,000. "
   "THE BRIEF CARRIES USD 75,000 for a tier-1 smart contract audit. That figure is retained as the "
   "CONSERVATIVE case rather than the Base because its source was never recorded. GET A QUOTE - firms of "
   "this kind publish nothing.",
   FMT_USD, "launch_audit_usd")
sp("disc_entry", "ICS discount - entry fee", 0.25, 0.15, 0.40, "% of the fee",
   "AN AVERAGE ACROSS TIERS, deliberately, rather than a five-rung ladder. Client instruction 2026-08-26: "
   "model one blended discount on the qualifying population instead of splitting Silver / Gold / Platinum / "
   "Sovereign, because no tier DISTRIBUTION exists in this model - the ICS rows are binary, qualified or "
   "not - and inventing one would put a fabricated mix underneath every benefit number. "
   "SIP ONLY. Spot earns no ICS, so this reduces Stream 1a and never Stream 1b. "
   "WHAT AN AVERAGE HIDES: if the real ladder tops out near 75%, top-tier contributions are LOSS-MAKING on "
   "the inflow lane. At a 5% fee, Aurumix's own cost is 1.425% of the contribution (the fabrication premium "
   "on the 95% that buys metal), so the fee breaks even at a 71.5% DISCOUNT and goes negative above it. A "
   "25% average is comfortably profitable; a 75% top rung is not. That may be a deliberate cross-subsidy "
   "from the card, but it should be a decision rather than a surprise - and this row cannot show it.",
   FMT_PCT, "disc_entry")
sp("disc_card", "ICS discount - card fees", 0.20, 0.10, 0.35, "% of the fee",
   "Blended discount on Stream 4 - FX margin, ATM allowance and the card issuance fee. The corpus ladder "
   "runs FX 2.0 / 1.5 / 1.0% by tier and an ATM allowance that rises with tier; this is the average effect "
   "of all of it, on the qualifying population. WAIVED REVENUE, not a cash cost.",
   FMT_PCT, "disc_card")
sp("disc_rebate", "ICS discount - gold rebate", 0.05, 0.03, 0.10, "% of card revenue",
   "Gold Rewards, as a share of ALL CARD REVENUE - interchange (Stream 2) plus cardholder fees (Stream 4). "
   "REBASED 2026-08-26 from interchange alone, which was too narrow on three counts. Interchange is the "
   "SMALLER HALF by a long way (USD 50k against USD 833k at Y7). Stream 2 is NET OF THE PARTNER'S ~55% "
   "SHARE, and sizing a customer's reward off the residual left after a partner split uses the wrong "
   "denominator - the customer generated the whole thing. And the corpus cap was never interchange-only: it "
   "reads 'capped at the interchange AND CREDIT REVENUE that customer generated, net of custody'. "
   "THE RATE FELL FROM 15% TO 5% BECAUSE THE BASE WIDENED 17.8x. The check that matters is the rebate as a "
   "share of CARD SPEND, against the corpus ladder of 0.15 / 0.45 / 0.75% by tier: 5% lands at 0.35% of "
   "spend, mid-ladder, against a corpus blended average of about 0.18%. 15% on this base would have been "
   "1.05% of spend - ABOVE THE TOP RUNG. "
   "IT STILL CANNOT OVERRUN ITS FUNDING LINE, because it is expressed as a share of the very revenue that "
   "funds it.",
   FMT_PCT, "disc_rebate")
sp("disc_family", "ICS discount - family wallet and will", 0.20, 0.10, 0.35, "% of the fee",
   "Blended discount on Stream 3, covering the Family Portfolio and Digital Will pricing and the "
   "per-beneficiary discount. The corpus ladder runs 0 / 10 / 20 / 35 / 50% by tier plus a per-beneficiary "
   "discount starting at Platinum; this is the average effect on the qualifying population.",
   FMT_PCT, "disc_family")
sp("agent_commission", "Agent commission - share of the entry fee", 0.10, 0.05, 0.15, "% of the fee",
   "CLIENT DECISION 2026-08-26: 10% of the entry fee, ONGOING - paid for as long as the customer keeps "
   "contributing, which is how agent commission normally works and what the corpus describes (high "
   "first-year commission plus renewal commission payable only while the policy stays in force). "
   "ON A 5% ENTRY FEE THIS IS 0.50% OF EVERY CONTRIBUTION. The client's own written figure was 15%, which "
   "the brief measured at 0.75pp against 0.85pp of Y1 gross margin - 88% of it. Conservative holds 15% so "
   "that number stays reachable. "
   "APPLIED TO THE AGENT-ACQUIRED SHARE OF THE BOOK, approximated as cumulative agent-driven acquisitions "
   "over cumulative acquisitions from all channels. THE MODEL DOES NOT COHORT CUSTOMERS BY CHANNEL, so this "
   "is an approximation: it assumes agent-acquired customers behave like everyone else on ticket and "
   "persistency. If agent-sold customers persist better - which the insurance precedent suggests - this "
   "UNDERSTATES the commission. "
   "NOT MODELLED: the three-tier upline override. Whether 10% is the total across all levels or the "
   "first-level rate is unresolved, and it is the difference between one commission and three.",
   FMT_PCT, "agent_commission")
sp("referral_reward", "Referral reward - share of the referee's entry fee", 0.30, 0.20, 0.40,
   "% of six contributions",
   "F17. 30% of the entry fee the referee pays over their SIX QUALIFYING CONTRIBUTIONS, credited in grams. "
   "On a USD 33.60 ticket at a 5% fee that is 6 x 33.60 x 5% = USD 10.08 of fee, so about USD 3.02 per "
   "successful referral. "
   "THE SELF-FUNDING CLAIM WAS WITHDRAWN, NOT REPAIRED. It was tested against a 2.15% gross margin that had "
   "not yet paid for the premium or the float. Against contribution-margin LTV on the inflow lane the "
   "reward is roughly 250% of it; against all-streams LTV it is comfortably affordable. The honest frame is "
   "CAC versus LTV, and which LTV you pick decides the answer. "
   "TIMING IS SIMPLIFIED: the reward is booked when the referral is acquired, but it is really paid at the "
   "REFEREE'S GATE - six contributions later, so no earlier than about month 19 given the channel itself "
   "opens at period 13. The annual totals barely move; the monthly cash profile does.",
   FMT_PCT, "referral_reward")
sp("card_fixed_usd", "Card programme - NymCard platform and scheme fees", 30000, 18000, 72000,
   "USD/yr",
   "USD 2,500/MONTH. Client decision 2026-08-26. NymCard's standing charge for running the programme - "
   "platform, account management and the scheme assessments that come with it. "
   "BIN SPONSORSHIP WAS REMOVED FROM THIS LINE, and that is the client's correction rather than mine. The "
   "figure began at USD 120,000/yr on the assumption that a BANK would sponsor the BIN and a PROCESSOR would "
   "run the platform - two counterparties, two fees. NymCard is a CBUAE-licensed PRINCIPAL MEMBER OF BOTH "
   "VISA AND MASTERCARD, so it sponsors the BIN under its own membership and there is no bank charging "
   "separately for it. The bundle loses a component and the number falls with it. "
   "WHAT IS LEFT IS STILL REAL: no BaaS provider gives away the platform. Every one charges a standing "
   "monthly fee with a minimum, because it carries scheme and regulatory obligations on the programme "
   "manager's behalf. Taking this to zero would say the card costs nothing to run. "
   "BENCHMARKED, NOT QUOTED - NymCard publishes no pricing and neither does any comparable (Magnati, "
   "SimpliFi, Paymentology, EMX, Rain, all quote-only). "
   "ONE DEPENDENCY LEFT: BaaS providers price as a platform fee, as a SHARE OF INTERCHANGE, or both. The "
   "model already hands 60% of interchange to a programme-manager counterparty. If NymCard IS that "
   "counterparty and prices on interchange share, even this reduced line double counts and should go to "
   "zero. One question settles it: platform fee, interchange share, or both?",
   FMT_USD, "card_fixed_usd")
sp("card_setup_usd", "Card programme - setup (one-off)", 50000, 25000, 100000, "USD",
   "Programme setup, integration and certification with the issuer-processor, booked in the month the card "
   "launches. BENCHMARKED - quote-only, same as the annual fee.",
   FMT_USD, "card_setup_usd")
sp("card_per_card", "Card programme - per card issued", 4.00, 2.00, 8.00, "USD/card",
   "Card production, personalisation and delivery, on the NEWLY ISSUED flow rather than the active stock - "
   "it is a manufacturing cost, paid once per card. A blend of physical and virtual: virtual costs cents, "
   "physical with delivery runs several dollars. BENCHMARKED. "
   "IF THE PROGRAMME GOES VIRTUAL-FIRST this falls close to zero, which is a real product lever - the "
   "Aggressive case is roughly that.",
   FMT_USD2, "card_per_card")
sp("card_per_auth", "Card programme - per authorisation", 0.03, 0.02, 0.06, "USD/auth",
   "Switching and authorisation cost per transaction, on the authorisation count the revenue build already "
   "carries as a memo. SMALL PER EVENT AND LARGE IN AGGREGATE - the model runs over 320,000 authorisations "
   "at Y7. BENCHMARKED.",
   FMT_USD2, "card_per_auth")
sp("card_fraud_bps", "Card programme - fraud and chargebacks", 0.0008, 0.0004, 0.0020, "% of spend",
   "Fraud losses plus chargeback and dispute handling, as a share of card spend. 8 basis points is a "
   "well-controlled programme; poorly controlled ones run several times that, which is why Conservative is "
   "25 bps. BENCHMARKED. "
   "NOT THE SAME RISK AS THE HOT-WALLET CRIME COVER in the insurance line: that protects Aurumix's own "
   "tokens, this is loss on customer card transactions. Different exposure, different counterparty, no "
   "overlap.",
   FMT_PCT2, "card_fraud_bps")
sp("xborder_rate", "Cross-border scheme assessment", 0.0140, 0.0100, 0.0140, "% of cross-border spend",
   "CHARGED TO THE ISSUER by the card scheme whenever the merchant country differs from the issuer country. "
   "PUBLISHED RATES: Visa International Service Assessment is 1.00% single-currency (ISA Base) and 1.40% "
   "multi-currency (ISA Enhanced); Mastercard splits it as a 0.90% Issuer Cross-Border Assessment plus a "
   "0.20% Currency Conversion Assessment, so 1.10% multi-currency. Retrieved 2026-08-26. "
   "BASE TAKES 1.40% because most of Aurumix's cross-border spend IS multi-currency - Indian customers "
   "spending rupees and Gulf customers spending rials on a UAE-issued card both convert. Aggressive takes "
   "the single-currency rate. "
   "IT IS BIGGER THAN AURUMIX'S INTERCHANGE SHARE. At 1.80% interchange and a 40% retained share Aurumix "
   "earns 0.72% of spend; the scheme charges up to 1.40% on the cross-border half. What covers the gap is "
   "the FX MARGIN charged to the customer (stream 4), which means the FX line is mostly COST RECOVERY "
   "rather than margin - and the ICS ladder cuts it to 1.0% at Sovereign, where it barely covers the "
   "assessment at all.",
   FMT_PCT2, "xborder_rate")
sp("uae_spend_abroad", "UAE cardholders - share of spend abroad or online-foreign", 0.10, 0.05, 0.20,
   "% of UAE spend",
   "THE ONLY ASSUMPTION IN THE CROSS-BORDER CALCULATION. Everything else is derived: Oman, Bahrain and "
   "India customers hold a UAE-ISSUED card, so EVERY transaction they make is cross-border by definition - "
   "merchant country differs from issuer country. That alone is 59% of card spend at Y7. This parameter "
   "adds the UAE residents' own travel and foreign online spend on top, taking the total to about 63%. "
   "ASSUMPTION. It moves the cross-border share by only a few points, so it is not load-bearing - the "
   "structure is.",
   FMT_PCT, "uae_spend_abroad")
sp("prefund_days", "Card settlement prefunding - days of spend", 2.0, 1.0, 3.0, "days",
   "WORKING CAPITAL, NOT A COST. The issuer-processor requires a prefunded settlement account so card "
   "transactions can settle before customer funds arrive - the JIT funding mechanism in the credit draft's "
   "section 5. "
   "NO PUBLIC SCHEDULE EXISTS. Prefunding is contractual between programme manager, processor and sponsor, "
   "never scheme-mandated, and it is not published by anyone. The common industry pattern is 1-3 days of "
   "expected transaction volume held in a segregated account, sometimes with a rolling reserve on top for "
   "higher-risk programmes. BENCHMARKED - take it to NymCard with everything else.",
   FMT_NUM2, "prefund_days")
sp("prefund_floor", "Card settlement prefunding - minimum balance", 100000, 50000, 250000, "USD",
   "A CONTRACTUAL FLOOR, and on this book IT BINDS THROUGHOUT - two days of Y7 card spend is only about "
   "USD 43,000, well under the minimum. So the prefunding requirement is effectively a fixed sum from card "
   "launch, which makes it a FOURTH minimum-commitment structure alongside the vault, the KYC platform and "
   "the card programme fee. "
   "Industry sources put minimums for small programmes in the low-to-mid six figures; nothing is published. "
   "BENCHMARKED.",
   FMT_USD, "prefund_floor")
sp("tech_build_y1", "Technology - build, Year 1", 350000, 200000, 600000, "USD",
   "CLIENT INSTRUCTION 2026-08-26: BUILD IN Y1-Y2, MAINTENANCE FROM Y3. The heavy year. Covers the mobile "
   "app, backend ledger, wallet infrastructure, the token contract, and integration with everything the "
   "product depends on - NymCard for the card, the KYC vendor, the payment rail, the custody and trust "
   "layer. "
   "IT IS NOT THE WHOLE BUILD. The client's own app was due early September 2026, before this model's "
   "horizon opens, so some of the consumer front end is already paid for outside these figures. What sits "
   "here is the regulated-product layer built on top of it. "
   "THE SMART CONTRACT AUDIT IS NOT IN HERE - it sits in the Year 1 one-off line with the launch technology "
   "audit, so the two must not be double counted. "
   "BENCHMARKED. No comparable publishes a build budget.",
   FMT_USD, "tech_build_y1")
sp("tech_build_y2", "Technology - build, Year 2", 150000, 80000, 300000, "USD",
   "The tail of the build. Y2 is when the CARD AND CREDIT LAYER lands - both streams activate at period 13 - "
   "so this is not merely finishing Y1's work, it is a second product going live. "
   "BENCHMARKED.",
   FMT_USD, "tech_build_y2")
sp("tech_maint", "Technology - annual maintenance from Y3", 120000, 70000, 220000, "USD/yr",
   "Hosting, third-party licences, security patching and ongoing engineering once the build settles. Runs "
   "from Y3 onward at the client's instruction. "
   "FLAT, WHICH IS A SIMPLIFICATION AND UNDERSTATES THE BACK YEARS. The book goes from about 4,000 paying "
   "customers at Y3 to over 70,000 at Y7; the brief's own technology block grows roughly five-fold across a "
   "comparable span. A flat line is the simple version, not the accurate one - revisit it if the Y6-Y7 "
   "numbers start carrying weight. "
   "ON-CHAIN MINTING COST IS NOT IN HERE, AND IT IS A CHAIN DECISION WORTH SIX FIGURES. Every monthly SIP "
   "contribution is a mint - about 855,000 events a year at Y7. On an L2 that is roughly USD 8,500/yr and "
   "immaterial; on Ethereum mainnet at USD 2 a transaction it is about USD 1.7m/yr, which would be one of "
   "the largest costs in the business. NOBODY HAS RECORDED WHICH CHAIN AURUMIX ISSUES ON. Until that is "
   "settled the cost cannot be modelled, and it is left out rather than guessed - but it must not be read "
   "as zero.",
   FMT_USD, "tech_maint")
sp("cost_contingency", "Contingency on total costs", 0.15, 0.10, 0.30, "% of costs",
   "CLIENT INSTRUCTION 2026-08-26: 15% on top of everything modelled, to stand in for the cost families not "
   "yet built. "
   "IT IS A PLACEHOLDER, NOT COVERAGE, AND THE ARITHMETIC SAYS SO. What is still missing is HEADCOUNT, "
   "technology's on-chain minting cost, legal and trust, security, corporate, and tax. Headcount alone is "
   "anchored in the brief at about USD 588,000 in Year 1 - roughly 78% of the entire Year 1 cost base as "
   "currently built - against which 15% is about USD 112,000. THE CONTINGENCY COVERS A FIFTH OF ONE MISSING "
   "BLOCK, and there are five others. "
   "READ EVERY PROFIT FIGURE IN THIS MODEL AS AN UPPER BOUND until headcount lands. It is not a forecast, "
   "it is revenue less the costs that happen to have been built so far.",
   FMT_PCT, "cost_contingency")
sp("kyc_min_month", "KYC and AML - monthly minimum", 299.0, 149.0, 299.0, "USD/month",
   "Sumsub published minimum monthly commitment for the Compliance tier, same source. THE SECOND OF THREE "
   "MINIMUM-COMMITMENT STRUCTURES in this cost base - the vault minimum is the first, the card programme "
   "minimums are the third. It binds below about 162 verifications a month, which on this book is most of "
   "the first two years.",
   FMT_USD2, "kyc_min_month")
_sr[0] += 1

for key, label, default, dvl, on_val, desc in [
    ("prepaid_vs_credit", "Prepaid instead of credit", "Credit", '"Credit,Prepaid"', "Prepaid",
     "ON (Prepaid) caps interchange at 1.00% and removes the credit stream entirely. 'NOT A PRODUCT CHOICE, "
     "IT IS THE BUSINESS MODEL.' Worth ~USD 2.3m of Y10 revenue on the ten-year run."),
    ("lapsed_keeps_card", "Holders keep the card", "ON", '"ON,OFF"', "ON",
     "NOBODY HAS DECIDED THIS. It determines whether the card streams - the majority of revenue - decay with "
     "churn or are immune to it. Worth a 42% swing in terminal revenue. Default ON because nothing in the "
     "design revokes the card; report both."),
    ("premium_absorbed", "Fabrication premium borne by", "Aurumix", '"Customer,Aurumix"', "Aurumix",
     "DEFAULT AURUMIX, which REVERSES the 2026-08-21 client decision and does so deliberately. That review "
     "put the premium on the customer - grams delivered short by (1+premium), stream 1 keeping the full "
     "headline fee - which is standard for gold products where the quoted price includes the dealer spread. "
     "The model now shows procurement as a COST instead, because a cost the client can see is the point of "
     "the cost build, and because charging the customer for it is better presented as UPSIDE than baked "
     "into the base case. Switching back to Customer restores the 2026-08-21 treatment exactly: grams fall "
     "1.50%, the COGS line goes to zero, and gross profit rises by the same amount. THE TWO SETTINGS ARE "
     "MUTUALLY EXCLUSIVE BY CONSTRUCTION, which is what stops the double-count fixed at 29f98e0 returning."),
    ("premium_on_gross", "Premium charged on gross inflow", "Net new", '"Net new,Gross"', "Gross",
     "D30: the premium is paid on NET NEW grams. Gold returned by redemption is already a fabricated bar "
     "sitting in the float, and a bar that exists does not need making twice. Switching to Gross charges "
     "every gram, which is correct ONLY if redeemed metal is sold back to the dealer rather than recycled. "
     "GATED ON CORRECTION 30, WHICH IS UNDESIGNED - nobody has written down which way redeemed gold goes. "
     "Worth USD 165,755 of retained margin over the horizon, so the undesigned question has a price."),
]:
    r = _sr[0]
    ws_scen["A%d" % r] = label
    ws_scen["B%d" % r] = '=IF($C$%d="%s",1,0)' % (r, on_val)
    ws_scen["B%d" % r].font = BLACK_BOLD
    ws_scen["B%d" % r].number_format = FMT_NUM
    ws_scen["C%d" % r] = default
    ws_scen["C%d" % r].font = BLUE_BOLD
    ws_scen["D%d" % r] = on_val
    ws_scen["D%d" % r].font = SECONDARY
    ws_scen["G%d" % r] = desc
    ws_scen["G%d" % r].font = NOTE_FONT
    d = DataValidation(type="list", formula1=dvl, allowBlank=False)
    ws_scen.add_data_validation(d)
    d.add(ws_scen["C%d" % r])
    SWROW[key] = r
    SROW["sw_" + key] = r          # so sref() reaches switches too
    SORDER.append("sw_" + key)
    # Label matches the Scenario sheet EXACTLY so the same row is findable on
    # both; the 1/0 nature is carried in the unit column, not the label.
    SMETA["sw_" + key] = (label, "1/0 switch", desc, FMT_NUM)
    declare("sw_" + key, "Scenario Parameters", "$B$%d" % r)
    _sr[0] += 1
ws_scen.freeze_panes = "B8"

# ============================================================================
# ASSUMPTIONS, continued - SCENARIO-LINKED INPUTS
#
# The firm's convention (DRODE): Scenario Parameters holds the Base /
# Aggressive / Conservative columns and picks the live one; ASSUMPTIONS is the
# complete register of every input; and the MODEL reads only from Assumptions.
# So each scenario parameter is mirrored here as a GREEN link to its active
# value. Change the scenario switch and these move with it.
#
# The chain is strictly one-directional:
#     Scenario Parameters  ->  Assumptions  ->  Model  ->  Summary
# ============================================================================
a_section("SCENARIO-LINKED INPUTS (live values - move with the scenario switch)")
SMIRROR = {}
for key in SORDER:
    param, unit, why, fmt = SMETA[key]
    r = _ar[0]
    ws_assum["A%d" % r] = param
    c = ws_assum["B%d" % r]
    c.value = "='Scenario Parameters'!$B$%d" % SROW[key]
    c.font = GREEN
    if fmt:
        c.number_format = fmt
    ws_assum["C%d" % r] = unit
    ws_assum["D%d" % r] = why
    SMIRROR[key] = r
    _ar[0] += 1
note(ws_assum, _ar[0],
     "Every row above is a GREEN link to the active value on Scenario Parameters, which is itself a CHOOSE() "
     "across the Base / Aggressive / Conservative columns. Flip the scenario selector and all of them move. "
     "The Model reads these cells and never reaches into Scenario Parameters directly, so this sheet is the "
     "complete register of every input the model uses.")
_ar[0] += 2


# REBIND: from here on, a scenario reference resolves through the Assumptions
# mirror rather than the Scenario sheet. Everything below (the Model) picks this
# up; the Scenario sheet is already written and keeps its own internal refs.
def sref(k): return "Assumptions!$B$%d" % SMIRROR[k]

# ============================================================================
# MODEL
# ============================================================================
ws_model["A1"] = "=Cover!$B$2"
ws_model["A1"].font = SHEET_TITLE
widths(ws_model, {"A": 52, "B": 18})
for i in range(N_PERIODS):
    ws_model.column_dimensions[pcol(i)].width = 13
MROW, _mr = {}, [2]


def m_row(key, label, unit, fn, fmt=FMT_NUM, font=BLACK, bold=False, total=False):
    r = _mr[0]
    ws_model["A%d" % r] = label
    ws_model["A%d" % r].font = BLACK_BOLD if bold else BLACK
    ws_model["B%d" % r] = unit
    ws_model["B%d" % r].font = SECONDARY
    for i in range(N_PERIODS):
        c = ws_model[pcell(r, i)]
        c.value = fn(i)
        c.font = font
        if fmt:
            c.number_format = fmt
        c.alignment = Alignment(horizontal="center")
        if total:
            c.border = TOTALS_BORDER
    MROW[key] = r
    _mr[0] = r + 1
    return r


def mr(key, i): return pcell(MROW[key], i)
def mrng(key): return "$%s$%d:$%s$%d" % (pcol(0), MROW[key], pcol(N_PERIODS - 1), MROW[key])


# -- period grid -------------------------------------------------------------
m_row("period_idx", "Period #", "1..29", lambda i: i + 1, FMT_NUM, BLACK_BOLD, True)
m_row("period", "Period", "-", lambda i: plabel(i), None, BLACK_BOLD, True)
m_row("period_type", "Period type", "-", lambda i: "Monthly" if is_monthly(i) else "Annual", None)
m_row("year", "Model year", "1..7", lambda i: pyear(i), FMT_NUM)
m_row("cal_month", "Calendar month", "0 = annual", lambda i: cal_month(i), FMT_NUM)
m_row("n", "Months in period", "months", lambda i: n_months(i), FMT_NUM)
m_row("fy_end", "Financial-year end", "1/0", lambda i: 1 if is_fy_end(i) else 0, FMT_NUM)
_mr[0] += 1

# -- activation ---------------------------------------------------------------
# There is no band of 1/0 flag rows. Each stream carries its own activation
# INLINE, reading the month straight from the Assumptions activation calendar,
# so the month still lives in exactly one place and a reader clicking any
# stream cell sees the condition rather than having to find a flag row.
# The only exception is the region gate, which is used twelve times over (three
# regions x four region rows) and so is computed once, below.


def gate(key, i, expr):
    """Wrap a stream calculation in its activation condition."""
    return "=IF(%s>=%s,%s,0)" % (mr("period_idx", i), aref("act_" + key), expr)


def opened(key, i):
    return "IF(%s>=%s,1,0)" % (mr("period_idx", i), aref("act_" + key))

# -- seasonality -------------------------------------------------------------
banner(ws_model, _mr[0], "SEASONALITY - normalised to EXACTLY 12.000 by construction")
_mr[0] += 1
AMP = sref("seasonality_amplitude")


def sq(key, k): return "Assumptions!$%s$%d" % (get_column_letter(SEAS_COL0 + k), SEAS_ROW[key])
def sqr(key): return "Assumptions!$%s$%d:$%s$%d" % (get_column_letter(SEAS_COL0), SEAS_ROW[key],
                                                    get_column_letter(SEAS_COL0 + 11), SEAS_ROW[key])


NORM = {}
for key, label, fmt in (("acq", "Acquisition seasonality - normalised", FMT_NUM3),
                        ("spend", "Card spend seasonality - normalised", FMT_NUM3)):
    r = _mr[0]
    ws_model["A%d" % r] = label
    ws_model["B%d" % r] = "Jan..Dec"
    ws_model["B%d" % r].font = SECONDARY
    for k in range(12):
        c = ws_model.cell(row=r, column=PCOL0 + k)
        c.value = "=(1+(%s-1)*%s)*12/SUMPRODUCT(1+(%s-1)*%s)" % (sq(key, k), AMP, sqr(key), AMP)
        c.font, c.number_format = GREEN, fmt
        c.alignment = Alignment(horizontal="center")
    NORM[key] = r
    _mr[0] += 1
r = _mr[0]
ws_model["A%d" % r] = "Foreign-spend share - rescaled to the mean"
ws_model["B%d" % r] = "Jan..Dec"
ws_model["B%d" % r].font = SECONDARY
for k in range(12):
    c = ws_model.cell(row=r, column=PCOL0 + k)
    c.value = "=%s/AVERAGE(%s)*%s" % (sq("foreign", k), sqr("foreign"), sref("foreign_spend_share"))
    c.font, c.number_format = GREEN, FMT_PCT
    c.alignment = Alignment(horizontal="center")
NORM["foreign"] = r
_mr[0] += 1
r = _mr[0]
ws_model["A%d" % r] = "Seasonality sum check (must be exactly 12.000)"
ws_model["A%d" % r].font = BLACK_BOLD
ws_model["B%d" % r] = "acq | spend"
ws_model["B%d" % r].font = SECONDARY
for k, key in enumerate(("acq", "spend")):
    c = ws_model.cell(row=r, column=PCOL0 + k)
    c.value = "=SUM($%s$%d:$%s$%d)" % (pcol(0), NORM[key], pcol(11), NORM[key])
    c.font, c.number_format = BLACK_BOLD, FMT_NUM3
    declare("seas_sum_" + key, "Model", "$%s$%d" % (pcol(k), r))
MROW["seas_sum"] = r
_mr[0] += 1


def seas(key, i, fmt_annual="1"):
    rng = "$%s$%d:$%s$%d" % (pcol(0), NORM[key], pcol(11), NORM[key])
    if key == "foreign":
        return "IF(%s=1,INDEX(%s,1,%s),AVERAGE(%s))" % (mr("n", i), rng, mr("cal_month", i), rng)
    return "IF(%s=1,INDEX(%s,1,%s),%s)" % (mr("n", i), rng, mr("cal_month", i), fmt_annual)


m_row("seas_acq", "Acquisition seasonality applied", "multiplier",
      lambda i: "=" + seas("acq", i), FMT_NUM3)
m_row("seas_spend", "Card spend seasonality applied", "multiplier",
      lambda i: "=" + seas("spend", i), FMT_NUM3)
m_row("seas_foreign", "Foreign-spend share applied", "%",
      lambda i: "=" + seas("foreign", i), FMT_PCT)
_mr[0] += 1

# ============================ ACQUISITION ===================================
# Total new customers, then split across regions. The referral and saturation
# terms read the WHOLE-BOOK totals at the PREVIOUS period, which live below the
# region blocks - so these rows are reserved here and written once those row
# numbers are known. Referencing a later row at column i-1 is not circular.
banner(ws_model, _mr[0], "ACQUISITION - all channels, then split across regions")
_mr[0] += 1
# Agents are routed to the regions they actually operate in; everything else is
# allocated across open regions by market size. Saturation is applied PER
# REGION, so concentrating a channel in one market cannot push that market past
# its own ceiling while the whole-book total still looks unbreached.
# Only the agent channel is computed centrally, because agent headcount is a
# single national resource routed to the markets it operates in. Marketing and
# referral are now computed INSIDE each region block, because each market has
# its own CAC and its own referring base.
# NO CENTRAL AGENT ROW ANY MORE (2026-08-26). Agent output used to be computed
# once nationally here and then sliced by a "share of salesforce" input, with an
# open-region renormalisation row so a market still in licensing redistributed
# its people rather than idling them. Headcount now belongs to a region, so both
# the central row and the renormalisation are gone: each region block computes
# its own agent output from its own headcount, and a region with zero agents
# contributes zero by construction rather than by a routing rule.
#
# THE GATE THE OLD RENORMALISATION EXISTED TO PLUG IS NOW STRUCTURAL. It caught
# a leak on 2026-08-21 - Oman & Bahrain earning agent revenue in M12, a month
# before it opens, because the agent channel was the one channel never gated on
# region opening. The per-region row below carries the same opening gate as
# direct and referral, so the three channels are now handled identically.
# The metal price now moves. Compounded on MODEL YEAR, not period, so all twelve
# months of a year share one price - a monthly compounding would imply intra-year
# precision the annual CAGR does not carry, and would put a kink at the M24/Y3
# boundary where the grid switches from monthly to annual.
m_row("gold_px", "Gold price (this period)", "USD/g",
      lambda i: "=%s*(1+%s)^(%s-1)" % (aref("gold_price"), sref("gold_cagr"), mr("year", i)),
      FMT_NUM2, GREEN)
note(ws_model, _mr[0],
     "RISING GOLD CUTS BOTH WAYS AND THE MODEL CAPTURES BOTH. Metal already held appreciates, but a fixed "
     "USD contribution BUYS FEWER GRAMS each year - so AUM does NOT simply compound at the headline rate. "
     "Only the collateral-linked streams move with this row (card limit, and lending through it); the entry "
     "fee is a percentage of a USD contribution and is untouched by the gold price.")
_mr[0] += 1

# ======================== ONE BLOCK PER REGION ==============================
# Each region is self-contained: its own customers, contributions, AUM, cards
# and all six streams, ending in a regional subtotal. The grand total is the
# sum of those subtotals plus the non-regional B2B line.
CHURN = sref("monthly_churn")
# CLIENT DECISION 2026-08-21: the CUSTOMER bears the fabrication premium. Their
# money buys metal at spot-plus-premium, so they receive fewer grams - which is
# exactly what "Grams purchased" already does by dividing by (1+premium). The
# entry fee therefore reaches Aurumix WHOLE. An earlier cut ALSO netted the
# premium off this margin, subtracting the same USD 1.40 per USD 100 twice:
# once from the customer's grams and once from Aurumix's fee. Only one can be
# true. Standard practice for gold products is that the quoted price includes
# the dealer spread, so the customer bears it. DO NOT reintroduce a premium
# term here without removing the (1+premium) divisor in "Grams purchased".
# GROSS, and named so. It was called NET_FEE back when the premium was netted
# off it; 29f98e0 moved the premium onto the customer and stream 1 has earned
# the full headline fee ever since. The name outlived the arithmetic by five
# days and is corrected here so the next reader is not misled by it.
GROSS_FEE = aref("entry_fee")
ATM_TERMS = "+".join(
    "%s*MAX(0,%s-%s)" % (sref("atm_" + k), sref("atmm_" + k), aref("atm_allowance"))
    for k, _n, _b, _a, _c, _mb, _ma, _mc in ATM)
SUBTOTALS = []

divider(ws_model, _mr[0],
        "REVENUE - the customer engine and the six streams, by region")
_mr[0] += 2

for rg in REGIONS:
    banner(ws_model, _mr[0], "REGION: %s" % RLAB[rg].upper())
    _mr[0] += 1
    op = "1" if rg not in [a[0] for a in ACTIVATIONS] else opened(rg, 0)

    def _open(i, rg=rg):
        return "1" if rg not in [a[0] for a in ACTIVATIONS] else opened(rg, i)

    # -- customers ----------------------------------------------------------
    # Raw demand = this region's share of non-agent demand, plus whatever share
    # of agent output actually sells here. Then this region's OWN saturation.
    # direct, referral, raw, sat, new, pay, churn, hold, cum
    # CAC IS NO LONGER A CONSTANT. Straight-line from the Y1 parameter to the Y7
    # one on MODEL YEAR, so it is visible on the face of the sheet rather than
    # buried inside the acquisition formula - the client will want to see it fall.
    m_row("cacnow_" + rg, "  Marketing CAC this period (ramping Y1 to Y7)", "USD/customer",
          lambda i, rg=rg: "=%s+(%s-%s)*(%s-1)/6" % (
              sref("cac_" + rg), sref("cac7_" + rg), sref("cac_" + rg), mr("year", i)),
          FMT_USD2)
    _pay_row, _cum_row = _mr[0] + 6, _mr[0] + 9
    m_row("direct_" + rg, "  Direct-driven (this market's budget at its own CAC)", "accounts",
          lambda i, rg=rg: "=INDEX(Assumptions!$B$%d:$B$%d,%s)*%s/12*%s/%s*(1+%s)*%s" % (
              AROW["mktg_y1"], AROW["mktg_y7"], mr("year", i), sref("mktshare_" + rg),
              mr("n", i), mr("cacnow_" + rg, i), sref("organic_share"), _open(i, rg)),
          FMT_NUM2, GREEN)
    m_row("ref_" + rg, "  Referral-driven (from this market's own base)", "accounts",
          lambda i, rg=rg, pr=_pay_row: 0 if i == 0 else gate("referral", i, "%s*%s/12*%s*%s" % (
              pcell(pr, i - 1), sref("referral_rate"), sref("referral_conversion"), mr("n", i))),
          FMT_NUM2)
    # THIS REGION'S OWN AGENTS (2026-08-26). Reads this region's headcount block
    # directly - no national pool, no renormalisation. A region with a zero
    # headcount contributes zero by construction, which is why UAE and Oman &
    # Bahrain need no switch to turn agents off. Carries the same opening gate
    # as direct and referral, so all three channels are handled identically.
    m_row("agent_" + rg, "  Agent-driven (this market's own agents)", "accounts",
          lambda i, rg=rg: "=INDEX(Assumptions!$B$%d:$B$%d,%s)*%s*INDEX(Assumptions!$B$%d:$B$%d,%s)*%s*%s"
          % (AROW["agents_%s_y1" % rg], AROW["agents_%s_y7" % rg], mr("year", i),
             sref("agent_productivity"), AROW["ramp_y1"], AROW["ramp_y7"], mr("year", i),
             mr("n", i), _open(i, rg)), FMT_NUM2, GREEN)
    m_row("raw_" + rg, "  Raw demand", "accounts",
          lambda i, rg=rg: "=%s+%s+%s" % (
              mr("direct_" + rg, i), mr("ref_" + rg, i), mr("agent_" + rg, i)), FMT_NUM2)
    m_row("sat_" + rg, "  Saturation (this region's own headroom)", "x",
          lambda i, rg=rg, cr=_cum_row: 1 if i == 0 else
          "=MAX(0,1-%s/%s)" % (pcell(cr, i - 1), aref("ceil_" + rg)), FMT_NUM3)
    m_row("new_" + rg, "  New customers", "accounts",
          lambda i, rg=rg: "=%s*%s*%s" % (mr("raw_" + rg, i), mr("sat_" + rg, i),
                                          mr("seas_acq", i)), FMT_NUM2)
    _pr = _mr[0]
    m_row("pay_" + rg, "  Paying customers", "accounts",
          lambda i, rg=rg, pr=_pr: "=%s*(1-%s)^(%s/2)" % (mr("new_" + rg, i), CHURN, mr("n", i)) if i == 0
          else "=%s*(1-%s)^%s+%s*(1-%s)^(%s/2)" % (
              pcell(pr, i - 1), CHURN, mr("n", i), mr("new_" + rg, i), CHURN, mr("n", i)),
          FMT_NUM, BLACK_BOLD)
    assert MROW["pay_" + rg] == _pay_row, \
        "region row layout drifted - the referral lag would read the wrong row"
    m_row("churn_" + rg, "  Churned this period", "accounts",
          lambda i, rg=rg, pr=_pr: "=%s-%s" % (mr("new_" + rg, i), mr("pay_" + rg, i)) if i == 0
          else "=%s+%s-%s" % (pcell(pr, i - 1), mr("new_" + rg, i), mr("pay_" + rg, i)), FMT_NUM2)
    _hr = _mr[0]
    m_row("hold_" + rg, "  Holders (stopped paying, still hold gold)", "accounts",
          lambda i, rg=rg, hr=_hr: "=%s" % mr("churn_" + rg, i) if i == 0
          else "=%s+%s" % (pcell(hr, i - 1), mr("churn_" + rg, i)), FMT_NUM)
    _cr = _mr[0]
    m_row("cum_" + rg, "  Cumulative ever acquired", "accounts",
          lambda i, rg=rg, cr=_cr: "=%s" % mr("new_" + rg, i) if i == 0
          else "=%s+%s" % (pcell(cr, i - 1), mr("new_" + rg, i)), FMT_NUM)
    assert MROW["cum_" + rg] == _cum_row, \
        "region row layout drifted - the saturation lag would read the wrong row"

    # -- contributions and AUM ----------------------------------------------
    m_row("sip_" + rg, "  SIP contributions", "USD",
          lambda i, rg=rg: gate("s1a", i, "%s*%s*%s" % (mr("pay_" + rg, i), sref("ticket_" + rg),
                                                        mr("n", i))), FMT_USD)
    # paying x attach x its dial x frequency x ticket x its dial, /12 for the
    # annual frequency, x months in period. BOTH scenario dials appear exactly
    # once - dropping either silently pins that half of spot to the base case.
    m_row("spot_" + rg, "  Spot purchase volume", "USD",
          lambda i, rg=rg: gate("s1b", i, "%s*%s*%s*%s*%s*%s/12*%s" % (
              mr("pay_" + rg, i), aref("spotattach_" + rg), sref("spot_attach_mult"),
              sref("spot_frequency"), aref("spotticket_" + rg), sref("spot_ticket_mult"),
              mr("n", i))), FMT_USD)
    # THE (1+premium) DIVISOR IS THE CUSTOMER BEARING THE PREMIUM, in metal
    # delivered rather than in cash (client decision 2026-08-21). Under
    # sw_premium_absorbed the divisor collapses to 1, grams are credited at
    # spot, and the premium reappears as a real cost in the COGS band below.
    # Exactly one of the two is ever live, so the premium cannot be charged on
    # both sides of the trade again - which is the bug 29f98e0 fixed.
    m_row("grams_in_" + rg, "  Grams purchased", "grams",
          lambda i, rg=rg: "=(%s+%s)*(1-%s)/%s/(1+%s*(1-%s))" % (
              mr("sip_" + rg, i), mr("spot_" + rg, i), aref("entry_fee"),
              mr("gold_px", i), aref("fab_premium"),
              sref("sw_premium_absorbed")), FMT_NUM2)
    m_row("decay_" + rg, "  Collateral-eligible AUM decay rate (holder-weighted)", "%/period",
          lambda i, rg=rg: "=(%s+%s*IF(%s+%s=0,1,(%s+%s*%s)/(%s+%s)))*%s/12" % (
              sref("self_custody_leakage"), sref("redemption_rate"),
              mr("pay_" + rg, i), mr("hold_" + rg, i),
              mr("pay_" + rg, i), mr("hold_" + rg, i), sref("lapsed_redemption_mult"),
              mr("pay_" + rg, i), mr("hold_" + rg, i), mr("n", i)), FMT_PCT2)
    # TWO DECAY RATES, because two different things are being measured. Added
    # 2026-08-21 after the client compared this model's AUM to tokenised gold
    # platforms and found it low - correctly, because we were publishing the
    # SMALLER of two legitimate numbers.
    #   - The rate above removes BOTH redemption and gold moved out of Aurumix's
    #     control, and drives the CREDIT LIMIT. Right for that: Aurumix cannot
    #     foreclose on a token in a private wallet.
    #   - The rate below removes ONLY REDEMPTION, because self-custody moves a
    #     token, not the metal. THE GOLD IS STILL IN THE VAULT. This is
    #     "gold under custody" - the measure PAXG, Kinesis and Comtech publish,
    #     and therefore THE ONLY ONE ANY EXTERNAL COMPARISON CAN USE.
    # Publishing only the collateral figure understated the comparable headline
    # by ~18%.
    m_row("decayr_" + rg, "  Redemption-only decay rate (metal actually leaving)", "%/period",
          lambda i, rg=rg: "=%s*IF(%s+%s=0,1,(%s+%s*%s)/(%s+%s))*%s/12" % (
              sref("redemption_rate"),
              mr("pay_" + rg, i), mr("hold_" + rg, i),
              mr("pay_" + rg, i), mr("hold_" + rg, i), sref("lapsed_redemption_mult"),
              mr("pay_" + rg, i), mr("hold_" + rg, i), mr("n", i)), FMT_PCT2)
    _gr = _mr[0]
    m_row("grams_" + rg, "  Grams held (collateral-eligible)", "grams",
          lambda i, rg=rg, gr=_gr: "=%s" % mr("grams_in_" + rg, i) if i == 0
          else "=%s*(1-%s)+%s" % (pcell(gr, i - 1), mr("decay_" + rg, i), mr("grams_in_" + rg, i)),
          FMT_NUM)
    _gc = _mr[0]
    m_row("custg_" + rg, "  Grams under custody (vaulted, incl. self-custodied)", "grams",
          lambda i, rg=rg, gc=_gc: "=%s" % mr("grams_in_" + rg, i) if i == 0
          else "=%s*(1-%s)+%s" % (pcell(gc, i - 1), mr("decayr_" + rg, i), mr("grams_in_" + rg, i)),
          FMT_NUM)
    m_row("aum_" + rg, "  AUM", "USD",
          lambda i, rg=rg: "=%s*%s" % (mr("grams_" + rg, i), mr("gold_px", i)), FMT_USD, BLACK_BOLD)

    # -- cards --------------------------------------------------------------
    m_row("cardbase_" + rg, "  Card-eligible base", "accounts",
          lambda i, rg=rg: "=%s+%s*%s" % (mr("pay_" + rg, i), mr("hold_" + rg, i),
                                          sref("sw_lapsed_keeps_card")), FMT_NUM2)
    # CLIENT DECISION 2026-08-26 (CG): THE CARD AND THE CREDIT FACILITY ARE NO
    # LONGER ELIGIBILITY-GATED. Anyone on the book can take them. ICS still
    # governs BENEFITS - fee discounts, better FX, family-wallet pricing, gold
    # reward rebates - but it no longer governs ACCESS.
    #
    # This row therefore stops feeding the card chain and becomes a MEMO. It is
    # kept, not deleted, because tomorrow's cost and discount work needs an
    # ICS-entitled population to attach the benefit costs to, and rebuilding it
    # a day later would be wasted work. It feeds nothing today - confirm that by
    # searching for its key: no other formula reads it.
    m_row("qual_" + rg, "  Reaches an ICS benefit tier (memo - drives discounts, not access)", "accounts",
          lambda i, rg=rg: "=IF(%s>1,%s*%s,IF(%s>%s,INDEX(%s,1,%s-%s)*%s,0))" % (
              mr("n", i), mr("cardbase_" + rg, i), sref("ever_qualify"),
              mr("period_idx", i), sref("months_to_qualify"), mrng("cardbase_" + rg),
              mr("period_idx", i), sref("months_to_qualify"), sref("ever_qualify")), FMT_NUM,
          SECONDARY)
    # Cards now come off the CARD-ELIGIBLE BASE directly - the whole book - not
    # off the gate-cleared subset. That is a 1/0.55 = 1.82x wider population,
    # and it arrives at once rather than lagging the ~8 months the gate took.
    # Take-up HOLDS at 18% (client decision, same day). Note what that assumes:
    # 18% was Indian gold-loan penetration (<10%) uplifted BECAUSE the population
    # had been pre-selected by six proven payments. Removing the gate removes the
    # pre-selection but keeps the uplift, so this rate is now applied to a less
    # committed population than the one it was calibrated on. It is the single
    # most optimistic assumption in the card chain and the first place to look if
    # card revenue is challenged.
    m_row("cards_" + rg, "  Active cards", "cards",
          lambda i, rg=rg: gate("s2", i, "%s*%s" % (mr("cardbase_" + rg, i), sref("card_activation"))),
          FMT_NUM, BLACK_BOLD)
    # SUPERSEDED 2026-08-26 (CG). This row used to restrict credit to the PAYING
    # share of cardholders: a lapsed customer kept the plastic but lost the
    # facility, because the ICS tier lapsed. Access is no longer ICS-gated, and a
    # lapsed holder still owns gold sitting in the vault, so the collateral that
    # backs a facility is still there. EVERY CARD NOW CARRIES A CREDIT LINE.
    #
    # THE ROW IS KEPT rather than deleted, and it is worth understanding why the
    # change is aggregate-NEUTRAL. Total credit capacity was already correct:
    #     ccards x limit  =  [cards x paying/base] x [AUM/paying x LTV]
    #                     =   cards x AUM x LTV / base
    # The "paying" terms cancel, so holders' gold was ALWAYS in the pool - the
    # old pair simply spread it over fewer cards at a larger limit each. Opening
    # lines to holders changes the LABELLING, not the total. What it does fix is
    # the reported per-customer limit below, which divided AUM by paying
    # customers only and therefore read ~2.7x too high per head once holders
    # became two thirds of the base. That figure feeds the average-transaction
    # plausibility check, so the display error was real even though the
    # aggregate was not.
    m_row("ccards_" + rg, "  ...of which have a live credit line (now all of them)", "cards",
          lambda i, rg=rg: "=%s" % mr("cards_" + rg, i), FMT_NUM, BLACK_BOLD)
    # NEW CARDS THIS PERIOD - a FLOW, not a stock (added 2026-08-26). The
    # issuance fee is a one-off charged when a card is activated, so it has to
    # be levied on this row. Charging it on the card STOCK, which is what the
    # model did until today, turns a one-off into an annual fee.
    #
    # Cards are 18% of the card-eligible base, and with holders retained that
    # base is cumulative-ever-acquired, so it only grows and this flow is simply
    # the increment. The MAX(0,...) exists for the case where the "holders keep
    # the card" switch is turned OFF: the base then becomes paying customers
    # only, which CAN fall, and a shrinking card book must not book negative
    # issuance revenue.
    m_row("newcards_" + rg, "  ...newly issued this period (flow, drives the issuance fee)", "cards",
          lambda i, rg=rg: "=%s" % mr("cards_" + rg, i) if i == 0
          else "=MAX(0,%s-%s)" % (mr("cards_" + rg, i), pcell(MROW["cards_" + rg], i - 1)),
          FMT_NUM2)
    # THE CARD IS A DRAWDOWN ON THE GOLD FACILITY (client decision 2026-08-20).
    # Aurumix has no balance sheet to lend from, so credit can only be extended
    # against collateral it holds. Spending on the card IS borrowing, and the
    # customer must repay to release their gold.
    #
    # Card spend is therefore NOT income-based. It is the annual drawdown:
    #     limit (gold x LTV)  x  drawn share  x  draw events per year
    # This is the same facility stream 5 already prices, so both now read one
    # volume instead of two inconsistent ones. Under PREPAID the card is loaded
    # from salary instead, and the gold does not constrain it.
    # DENOMINATOR CORRECTED 2026-08-26: AUM is now spread over the CARD-ELIGIBLE
    # BASE, not over paying customers alone. The old divisor was a leftover from
    # when only payers could borrow; with holders at ~two thirds of the base it
    # overstated the per-head limit by ~2.7x. The AGGREGATE was unaffected (see
    # the credit-line row above - the terms cancelled), but this row is read on
    # its own as "what limit does a customer actually get", and it feeds the
    # average-transaction-size check, so the overstatement was live in both.
    #
    # NO MINIMUM BALANCE IS APPLIED, and that is deliberate. A floor - "no card
    # below USD 200 of gold" - cannot be represented in an average-based engine:
    # there is no balance DISTRIBUTION here to apply it to, only a mean, so any
    # floor would require inventing a distribution. That is precisely the
    # heterogeneity this build pushed to the Phase 5 simulation. KNOWN
    # SIMPLIFICATION: the model therefore issues notional limits to customers
    # whose real balance would be too small to bother, and the error runs one
    # way - it flatters the card streams. Size it in Phase 5, where balances are
    # distributed, and re-impose a floor there if it matters.
    m_row("limit_" + rg, "  Credit limit per customer (gold x LTV)", "USD",
          lambda i, rg=rg: "=IF(%s=0,0,%s/%s*%s)" % (
              mr("cardbase_" + rg, i), mr("aum_" + rg, i), mr("cardbase_" + rg, i),
              aref("ltv_gold")), FMT_USD)
    m_row("drawyr_" + rg, "  Annual drawdown per card (limit x drawn x draws)", "USD/yr",
          lambda i, rg=rg: "=%s*%s*%s" % (mr("limit_" + rg, i), sref("drawn_share"),
                                          sref("draw_events")), FMT_USD)
    m_row("spendcard_" + rg, "  Card spend per card per month", "USD",
          lambda i, rg=rg: "=%s/12*%s" % (mr("drawyr_" + rg, i), mr("seas_spend", i)), FMT_USD)
    m_row("spendaed_" + rg, "  Card spend", "AED",
          lambda i, rg=rg: "=%s*%s*%s*%s" % (
              mr("ccards_" + rg, i), mr("spendcard_" + rg, i), aref("aed_usd"),
              mr("n", i)), FMT_NUM)
    m_row("spendusd_" + rg, "  Card spend", "USD",
          lambda i, rg=rg: "=%s/%s" % (mr("spendaed_" + rg, i), aref("aed_usd")), FMT_USD)
    # Average ticket is DERIVED, never input: monthly spend / transactions. It
    # therefore scales with this region's SIP ticket automatically, and cannot
    # drift to a basket the customer could not fund.
    # The draw events cancel: average ticket = drawn amount / transactions per
    # draw. So it depends only on how big a borrowing is and across how many
    # purchases it is spent - not on how often the customer borrows.
    m_row("avgtxn_" + rg, "  Average transaction size (derived)", "AED",
          lambda i, rg=rg: "=IF(%s=0,0,%s*%s*%s/%s)" % (
              sref("card_txns_per_draw"), mr("limit_" + rg, i), sref("drawn_share"),
              aref("aed_usd"), sref("card_txns_per_draw")), FMT_NUM)
    # MEMO for the cost build - the processor bills per AUTHORISATION, which is
    # draws x transactions per draw, grossed up for declines. Feeds no revenue.
    m_row("auths_" + rg, "  Card authorisations (memo - drives cost, not revenue)", "auths",
          lambda i, rg=rg: "=%s*%s*%s*%s/12*(1+%s)" % (
              mr("ccards_" + rg, i), sref("draw_events"), sref("card_txns_per_draw"),
              mr("n", i), aref("decline_uplift")), FMT_NUM)

    # -- the six streams, plus the redemption cost --------------------------
    m_row("s1a_" + rg, "  Stream 1a - Entry fee, SIP", "USD",
          lambda i, rg=rg: "=%s*%s" % (mr("sip_" + rg, i), GROSS_FEE), FMT_USD)
    m_row("s1b_" + rg, "  Stream 1b - Entry fee, SPOT", "USD",
          lambda i, rg=rg: "=%s*%s" % (mr("spot_" + rg, i), GROSS_FEE), FMT_USD)
    m_row("s2_" + rg, "  Stream 2 - Card interchange", "USD",
          lambda i, rg=rg: "=%s*IF(%s=1,MIN(%s,1%%),%s)*(1-%s)" % (
              mr("spendusd_" + rg, i), sref("sw_prepaid_vs_credit"), aref("ic_gold"),
              aref("ic_gold"), sref("pm_share")), FMT_USD)
    # A SUBSCRIBER BALANCE, not a percentage of the live book. Stream 3 used to
    # be paying x attach x price, which cannot fall except when customers leave -
    # it assumed nobody ever cancels while staying a customer. Same rolling
    # shape as the customer engine: opening x survival + new x half-period
    # survival, so a subscriber acquired mid-period is not charged a full one.
    _sb = _mr[0]
    m_row("subs_" + rg, "  Family plan subscribers", "subscribers",
          lambda i, rg=rg, sb=_sb: gate("s3", i, "%s*%s*(1-%s)^(%s/2)" % (
              mr("new_" + rg, i), sref("family_attach"), sref("family_churn_m"), mr("n", i)))
          if i == 0 else gate("s3", i, "%s*(1-%s)^%s+%s*%s*(1-%s)^(%s/2)" % (
              pcell(sb, i - 1), sref("family_churn_m"), mr("n", i),
              mr("new_" + rg, i), sref("family_attach"), sref("family_churn_m"), mr("n", i))),
          FMT_NUM2)
    # Plan fee plus the beneficiary fee on every name BEYOND THE FIRST. The
    # MAX(0,...) is load-bearing: without it a conservative scenario with fewer
    # than one beneficiary per plan would generate NEGATIVE beneficiary revenue
    # and quietly net it off the plan fee.
    m_row("s3_" + rg, "  Stream 3 - Family plan and Digital Will", "USD",
          lambda i, rg=rg: "=%s*(%s+MAX(0,%s-1)*%s)/12*%s" % (
              mr("subs_" + rg, i), aref("family_price"), sref("benef_count"),
              aref("benef_fee"), mr("n", i)), FMT_USD)
    m_row("s4_" + rg, "  Stream 4 - Cardholder fees (FX, ATM, events)", "USD",
          # THREE COMPONENTS, TWO DIFFERENT BASES - the distinction is the whole
          # point of the 2026-08-26 correction:
          #   FX      -> card SPEND         (a rate on volume)
          #   ATM     -> card STOCK         (a recurring per-cardholder event)
          #   REISSUE -> card STOCK         (recurring: tier upgrades)
          #   REPLACE -> card STOCK         (recurring: loss, theft, damage)
          #   ISSUE   -> NEW cards, a FLOW  (one-off at activation)
          # Only the last one changed, and it was the largest error in the model.
          lambda i, rg=rg: gate("s4", i, "%s*%s*%s+%s*(%s)*%s/%s*%s+%s*%s/%s+%s*(%s*%s+%s*%s)/12/%s*%s" % (
              mr("spendusd_" + rg, i), mr("seas_foreign", i), aref("fx_margin"),
              mr("ccards_" + rg, i), ATM_TERMS, aref("atm_fee"), aref("aed_usd"), mr("n", i),
              mr("newcards_" + rg, i), aref("issuance_fee"), aref("aed_usd"),
              mr("cards_" + rg, i), sref("reissue_rate"), aref("issuance_fee"),
              sref("replacement_events"), aref("replacement_fee"), aref("aed_usd"), mr("n", i))),
          FMT_USD)
    # Stream 5 reads THE SAME facility and the same cardholder population as
    # streams 2 and 4. Interchange comes from the merchant, lending fees from
    # the borrower - two payers on one drawdown, which is not double-counting,
    # but the VOLUME must be a single number, and now is.
    m_row("drawn_" + rg, "  Average drawn balance outstanding", "USD",
          lambda i, rg=rg: gate("s5", i, "%s*%s*(1-%s)*%s*%s" % (
              mr("cards_" + rg, i), mr("limit_" + rg, i), sref("sw_prepaid_vs_credit"),
              sref("drawn_share"), sref("facility_turnover"))), FMT_USD)
    m_row("s5_" + rg, "  Stream 5 - Lending revenue share", "USD",
          lambda i, rg=rg: gate("s5", i, "(%s*%s*%s+%s*%s*(1-%s)*%s*%s)*%s/12" % (
              mr("drawn_" + rg, i), aref("servicing_gross"), aref("servicing_share"),
              mr("ccards_" + rg, i), mr("drawyr_" + rg, i), sref("sw_prepaid_vs_credit"),
              aref("origination_gross"), aref("origination_share"), mr("n", i)), ), FMT_USD)
    # MEMO for the cost build - redemption is a COST with no offsetting revenue
    # (VARA forbids charging for it), so the event count is carried here and the
    # cost itself arrives with the cost build. It feeds no total.
    m_row("redem_" + rg, "  Redemption events (memo - drives cost, not revenue)", "events",
          lambda i, rg=rg: gate("s0", i, "(%s+%s*%s)*%s*%s/12" % (
              mr("pay_" + rg, i), mr("hold_" + rg, i), sref("lapsed_redemption_mult"),
              sref("redemption_rate"), mr("n", i))), FMT_NUM)
    m_row("sub_" + rg, "  SUBTOTAL - %s" % RLAB[rg], "USD",
          lambda i, rg=rg: "=" + "+".join(mr("%s_%s" % (s, rg), i)
                                          for s in ("s1a", "s1b", "s2", "s3", "s4", "s5")),
          FMT_USD, BLACK_BOLD, True)
    SUBTOTALS.append("sub_" + rg)
    _mr[0] += 1

# ===================== NON-REGIONAL AND GRAND TOTAL =========================
banner(ws_model, _mr[0], "NON-REGIONAL")
_mr[0] += 1
m_row("partner_aum", "  Partner AUM", "USD",
      lambda i: gate("s6", i, "INDEX(Assumptions!$B$%d:$B$%d,%s)*%s" % (
          SMIRROR["partners_y1"], SMIRROR["partners_y7"], mr("year", i),
          sref("aum_per_partner"))), FMT_USD, GREEN)
# NOT indented: this is a whole-book stream in its own right, not a sub-item of
# a region block, and the indentation is what distinguishes the two.
m_row("s6", "Stream 6 - B2B platform fee", "USD",
      lambda i: "=%s*%s*%s/12" % (mr("partner_aum", i), sref("b2b_fee"), mr("n", i)),
      FMT_USD, BLACK_BOLD)
note(ws_model, _mr[0],
     "Stream 6 is NOT regional: partners are institutions with their own customer books, not retail "
     "customers of a region. It is reported separately rather than allocated, because any allocation would "
     "be invented.")
_mr[0] += 2

banner(ws_model, _mr[0], "WHOLE-BOOK TOTALS")
_mr[0] += 1
for key, label, parts, fmt, bold in (
    ("paying", "PAYING CUSTOMERS", ["pay_" + r for r in REGIONS], FMT_NUM, True),
    ("holders", "HOLDERS (stopped paying, still hold gold)", ["hold_" + r for r in REGIONS], FMT_NUM, True),
    ("cum_ever", "Cumulative ever acquired", ["cum_" + r for r in REGIONS], FMT_NUM, False),
    ("new_total", "NEW CUSTOMERS", ["new_" + r for r in REGIONS], FMT_NUM, True),
    ("active_cards", "ACTIVE CARDS", ["cards_" + r for r in REGIONS], FMT_NUM, True),
    ("grams_held", "GRAMS HELD", ["grams_" + r for r in REGIONS], FMT_NUM, False),
    ("grams_custody", "GRAMS UNDER CUSTODY", ["custg_" + r for r in REGIONS], FMT_NUM, True),
    ("grams_bought", "Grams purchased", ["grams_in_" + r for r in REGIONS], FMT_NUM2, False),
    ("aum", "COLLATERAL-ELIGIBLE AUM", ["aum_" + r for r in REGIONS], FMT_USD, True),
    ("qualified", "Reaches an ICS benefit tier (memo - drives discounts, not access)",
     ["qual_" + r for r in REGIONS], FMT_NUM, False),
    ("sip_inflow", "SIP contributions", ["sip_" + r for r in REGIONS], FMT_USD, False),
    ("spot_inflow", "Spot purchase volume", ["spot_" + r for r in REGIONS], FMT_USD, False),
    ("card_spend_usd", "Card spend", ["spendusd_" + r for r in REGIONS], FMT_USD, False),
):
    m_row(key, label, "-", lambda i, p=parts: "=" + "+".join(mr(x, i) for x in p), fmt, BLACK, bold)
m_row("gold_custody", "GOLD UNDER CUSTODY (comparable headline)", "USD",
      lambda i: "=%s*%s" % (mr("grams_custody", i), mr("gold_px", i)), FMT_USD, BLACK_BOLD)
note(ws_model, _mr[0],
     "THE NUMBER TO QUOTE AGAINST OTHER GOLD PLATFORMS, and the only one that compares like for like. "
     "Collateral-eligible AUM above is SMALLER because it also removes gold moved out of Aurumix's control "
     "- a token in a private wallet cannot back a credit limit, but THE METAL IS STILL IN THE VAULT, so it "
     "is still gold under custody. PAXG, Kinesis and Comtech all publish the custody measure. COMPARE IN "
     "GRAMS, NOT DOLLARS: every published USD figure carries the gold price of its own vintage, so a "
     "comparison in dollars is really a comparison of gold prices. Comtech Gold, the closest comparable - "
     "Dubai, DMCC-licensed, tokenised gold - holds ~141,000 g.")
_mr[0] += 1
m_row("gold_custody_const", "  Gold under custody at the CONSTANT M1 price (memo)", "USD",
      lambda i: "=%s*%s" % (mr("grams_custody", i), aref("gold_price")), FMT_USD)
m_row("metal_effect", "  of which is the metal moving, not the business", "USD",
      lambda i: "=%s-%s" % (mr("gold_custody", i), mr("gold_custody_const", i)), FMT_USD)
note(ws_model, _mr[0],
     "THE HONEST SERIES FOR JUDGING EXECUTION is the constant-price row: it strips the gold price out and "
     "leaves grams accumulated, which is the part Aurumix controls. The headline above will grow even if the "
     "business does not. SET GOLD APPRECIATION TO THE CONSERVATIVE SCENARIO (0.0%) TO RECOVER THE ORIGINAL "
     "FLAT-PRICE MODEL, under which these two rows are identical by construction.")
_mr[0] += 1
m_row("conservation", "CHECK: paying + holders = cumulative ever acquired", "delta",
      lambda i: "=%s+%s-%s" % (mr("paying", i), mr("holders", i), mr("cum_ever", i)), FMT_NUM3, BLACK_BOLD)
m_row("custody_check", "CHECK: custody >= collateral-eligible (must be >= 0)", "grams",
      lambda i: "=%s-%s" % (mr("grams_custody", i), mr("grams_held", i)), FMT_NUM3, BLACK_BOLD)
# Sanity memo: under the credit variant the card is secured on the customer's
# own gold, so the limit that gold supports is a hard ceiling on monthly spend.
m_row("aum_per_cust", "  Collateral-eligible AUM per paying customer (memo)", "USD",
      lambda i: "=IF(%s=0,0,%s/%s)" % (mr("paying", i), mr("aum", i), mr("paying", i)), FMT_USD)
# Total monthly card spend against TOTAL credit capacity - capacity summed as
# cards x that region's own limit. Dividing blended spend by a PAYING-weighted
# limit is the wrong aggregation: cards concentrate in the older, higher-limit
# regions, so that version reads above 1.00 even when every region is inside
# its own cap. NB the summed spend must be bracketed before dividing.
m_row("credit_capacity", "  Total credit capacity across cardholders (memo)", "USD",
      lambda i: "=" + "+".join("%s*%s" % (mr("cards_" + r, i), mr("limit_" + r, i)) for r in REGIONS),
      FMT_USD)
m_row("spend_vs_limit", "  CHECK: card spend vs credit capacity (must be <= 1.00)", "x",
      lambda i: "=IF(%s=0,0,((%s)/%s)/%s)" % (
          mr("credit_capacity", i), "+".join(mr("spendusd_" + r, i) for r in REGIONS),
          mr("n", i), mr("credit_capacity", i)), '0.00"x"', BLACK_BOLD)
_mr[0] += 1

# ---- SERVICING LOAD: REPORTED, NOT CAPPED (2026-08-26) ---------------------
# Paying customers per agent, for the regions that have agents. An agent both
# acquires and services, so there is some ceiling on the book one person can
# carry - but WHAT that ceiling is depends entirely on what servicing involves,
# and for this product it has not been designed. The comparators span an order
# of magnitude: an Indian microfinance loan officer manages 345 in practice and
# 500-550 at full stretch (CreditAccess Grameen, FY23 concall) because they
# physically attend weekly cash-collection meetings; a Bank Mitra carries ~2,335
# (SIDBI) because servicing is on-demand; a life agent's book is effectively
# unbounded because a policy needs one renewal nudge a year.
#
# AURUMIX SITS AT THE LIGHT END - the SIP is a standing auto-debit and the
# product is in an app, so nobody is collecting cash - which is why this is a
# REPORTED DIAGNOSTIC AND NOT A CONSTRAINT. It blocks nothing. It exists so the
# number is visible on the sheet and can be argued with, rather than hiding
# inside the arithmetic while a scenario quietly drives it somewhere absurd.
# Set a real ceiling only once the servicing model is defined - which is a
# mechanism-design question, not a modelling one. The 100 G Business Model S11.1
# says agents "assist investors in joining the ecosystem and maintaining their
# SIP" and pays a trail commission at each Mining Event, which implies an
# ongoing role but never defines it.
section(ws_model, _mr[0], "SERVICING LOAD - reported, not capped")
_mr[0] += 1
for rg in REGIONS:
    m_row("bookperagent_" + rg, "  Paying customers per agent - %s" % RLAB[rg], "customers",
          lambda i, rg=rg: "=IF(%s=0,0,%s/%s)" % (
              mr("agent_" + rg, i), mr("pay_" + rg, i),
              "(%s/%s/%s)" % (mr("agent_" + rg, i), sref("agent_productivity"), mr("n", i))),
          FMT_NUM2)
note(ws_model, _mr[0],
     "Blank where a region has no agents, which is UAE and Oman & Bahrain by design. Compare against "
     "345-550 (microfinance loan officer, heavy weekly servicing) and ~2,335 (Bank Mitra, on-demand "
     "servicing). Aurumix's servicing is lighter than either, so a high number here is not automatically "
     "wrong - but it should be a decision rather than an accident.")
_mr[0] += 2

for key, label, parts in (
    ("s1a", "Stream 1a - Entry fee, SIP", ["s1a_" + r for r in REGIONS]),
    ("s1b", "Stream 1b - Entry fee, SPOT", ["s1b_" + r for r in REGIONS]),
    ("s2", "Stream 2 - Card interchange", ["s2_" + r for r in REGIONS]),
    ("s3", "Stream 3 - Family plan and Digital Will", ["s3_" + r for r in REGIONS]),
    ("s4", "Stream 4 - Cardholder fees", ["s4_" + r for r in REGIONS]),
    ("s5", "Stream 5 - Lending revenue share", ["s5_" + r for r in REGIONS]),
    ("redemptions", "Redemption events (memo - drives cost, not revenue)",
     ["redem_" + r for r in REGIONS]),
    ("auths", "Card authorisations (memo - drives cost, not revenue)",
     ["auths_" + r for r in REGIONS]),
):
    _memo = key in ("redemptions", "auths")
    m_row(key, label, "events" if _memo else "USD",
          lambda i, p=parts: "=" + "+".join(mr(x, i) for x in p),
          FMT_NUM if _memo else FMT_USD)
m_row("total_rev", "TOTAL NET REVENUE", "USD",
      lambda i: "=" + "+".join(mr(s, i) for s in SUBTOTALS) + "+" + mr("s6", i),
      FMT_USD, BLACK_BOLD, True)
m_row("rev_per_cust", "Net revenue per paying customer (annualised)", "USD",
      lambda i: "=IF(%s=0,0,%s/%s/%s*12)" % (mr("paying", i), mr("total_rev", i),
                                             mr("paying", i), mr("n", i)), FMT_USD)
m_row("total_check", "CHECK: regional subtotals + B2B = sum of streams", "delta",
      lambda i: "=%s-(%s)" % (mr("total_rev", i),
                              "+".join(mr(s, i) for s in
                                       ("s1a", "s1b", "s2", "s3", "s4", "s5", "s6"))),
      FMT_NUM3, BLACK_BOLD)

# ============================================================================
# WORKING CAPITAL - THE FLOAT
#
# NOT A COST, and it is in its own band precisely so nobody reads it as one.
# The float is INVENTORY: Aurumix's own gold, bought with Aurumix's own money
# and sitting on its own shelf before any customer owns it. The money is tied
# up, not spent, and it comes back on wind-down. It belongs on the balance
# sheet and in the funding ask; it must never reach the P&L.
#
# THE SIZING RULE, verbatim from _draft_allocation-and-float.md:
#     float >= one bar denomination + a buffer of N days trailing inflow,
#     two bars being the launch setting.
# Two bars is a SETTLED RULE. N is an assumption - see float_buffer_days.
# ============================================================================
_mr[0] += 2
divider(ws_model, _mr[0],
        "WORKING CAPITAL - the float. NOT A COST: money tied up, not money spent")
_mr[0] += 2

banner(ws_model, _mr[0],
       "THE FLOAT - Aurumix's own gold, held before any customer owns it")
_mr[0] += 1

m_row("daily_grams", "  Daily demand", "grams/day",
      lambda i: "=%s/(%s*365/12)" % (mr("grams_bought", i), mr("n", i)), FMT_NUM2)

# MAX of the two halves. The floor binds until demand passes 10 g/day (around
# M6 at Base); above that the buffer term takes over and the float scales with
# the book. Writing it as MAX rather than an IF keeps the crossover implicit in
# the arithmetic instead of hard-coding the month it happens.
m_row("float_grams", "  FLOAT REQUIRED", "grams",
      lambda i: "=MAX(2*%s,%s+%s*%s)" % (
          aref("bar_grams"), aref("bar_grams"),
          sref("float_buffer_days"), mr("daily_grams", i)), FMT_NUM)
note(ws_model, _mr[0],
     "MAX(two bars, one bar + N days of demand). TWO BARS IS THE FLOOR because you need one bar to sell from "
     "and one behind it - hold only one and you drop below a full bar the moment you sell, with customers "
     "already paid for gold you cannot deliver. That half is a SETTLED RULE from the design docs. "
     "THE N-DAY BUFFER IS THE ASSUMPTION, and it sizes REPLENISHMENT - how long to get metal back into the "
     "vault, not how long demand takes to empty the shelf. The floor binds until demand reaches 10 grams a "
     "day, around month six, so THE LAUNCH FIGURE DOES NOT DEPEND ON THE BUFFER AT ALL.")
_mr[0] += 1

m_row("float_bars", "  ...in bars", "bars",
      lambda i: "=%s/%s" % (mr("float_grams", i), aref("bar_grams")), FMT_NUM2)

m_row("float_usd", "FLOAT REQUIRED (standing balance)", "USD",
      lambda i: "=%s*%s*(1+%s)" % (mr("float_grams", i), mr("gold_px", i),
                                   aref("fab_premium")), FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "AT SPOT PLUS THE PREMIUM, because the float is bought from the dealer like any other metal. This is a "
     "BALANCE, not a flow - it is what is sitting on the shelf at that date, not what was spent that period.")
_mr[0] += 1

# THE CASH CALL, and it is deliberately NOT the change in the USD balance.
# The balance also moves when gold reprices, and revaluing metal already held
# needs no new money. Only EXTRA GRAMS have to be funded.
m_row("float_topup", "  Cash needed this period (new grams only)", "USD",
      lambda i: "=%s*%s*(1+%s)" % (mr("float_grams", i), mr("gold_px", i),
                                   aref("fab_premium")) if i == 0
      else "=MAX(0,%s-%s)*%s*(1+%s)" % (
          mr("float_grams", i), mr("float_grams", i - 1), mr("gold_px", i),
          aref("fab_premium")), FMT_USD)
note(ws_model, _mr[0],
     "NOT the change in the balance above. The balance also moves when gold reprices, and metal already on "
     "the shelf needs no new money to revalue - only EXTRA GRAMS have to be funded. M1 carries the whole "
     "200 g because that is the launch purchase.")
_mr[0] += 1

_fc = _mr[0]
m_row("float_cum", "CUMULATIVE CASH INVESTED IN THE FLOAT", "USD",
      lambda i, fc=_fc: "=%s" % mr("float_topup", i) if i == 0
      else "=%s+%s" % (pcell(fc, i - 1), mr("float_topup", i)),
      FMT_USD, BLACK_BOLD, True)

m_row("card_prefund", "CARD SETTLEMENT PREFUNDING (working capital, not a cost)", "USD",
      lambda i: gate("s2", i, "MAX(%s,%s/365*%s)" % (
          sref("prefund_floor"), mr("card_spend_usd", i), sref("prefund_days"))),
      FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "CASH ON DEPOSIT WITH THE PROCESSOR so card transactions settle before customer funds arrive - the JIT "
     "funding mechanism. LIKE THE FLOAT, IT IS MONEY TIED UP RATHER THAN SPENT, and it comes back if the "
     "programme winds down. It is NOT in any cost total. "
     "THE FLOOR BINDS THROUGHOUT: two days of Y7 card spend is about USD 43,000 against a USD 100,000 "
     "minimum, so this is effectively a fixed sum from card launch - a FOURTH minimum-commitment structure. "
     "No public schedule exists for prefunding; it is contractual with the processor.")
_mr[0] += 1

m_row("float_check", "CHECK: float never falls below two bars (must be >= 0)", "grams",
      lambda i: "=%s-2*%s" % (mr("float_grams", i), aref("bar_grams")),
      FMT_NUM2, BLACK_BOLD)

# ============================================================================
# COST BASE
#
# Everything ABOVE the divider is revenue; everything BELOW it is cost. The two
# are built as separate bands on purpose - v1.0 interleaved them and it stopped
# being possible to say what any single stream actually earned.
#
# Costs arrive ONE FAMILY AT A TIME, deliberately. This band currently carries
# COGS and nothing else. Opex, acquisition, card programme, benefit costs and
# tax each land here in turn, and each appends its own subtotal key to
# COST_LINES so the totals below never need editing.
# ============================================================================
_mr[0] += 2
divider(ws_model, _mr[0],
        "COST BASE - every row below this line is money going OUT")
_mr[0] += 2

banner(ws_model, _mr[0],
       "COST OF GOODS SOLD - inside the gold trade, charged against stream 1")
_mr[0] += 1

COGS_LINES = []

# The metal that comes BACK. Derived from the custody balance rather than
# recomputed from the decay rates, so it reconciles to the revenue band by
# construction: custody(t) = custody(t-1) - returned(t) + purchased(t), so
# returned(t) is whatever that identity leaves over. Recomputing it from
# decayr would give a second, independently-driftable copy of the same number.
m_row("grams_recycled", "  Grams returned by redemption (recycled into the float)", "grams",
      lambda i: "=0" if i == 0 else "=MAX(0,%s+%s-%s)" % (
          mr("grams_custody", i - 1), mr("grams_bought", i), mr("grams_custody", i)),
      FMT_NUM2)
note(ws_model, _mr[0],
     "THE METAL NEVER PHYSICALLY LEAVES. This product has no physical redemption - exit is a cash buyback - "
     "so a redeeming customer's gold stops being theirs and becomes Aurumix's own float, already vaulted "
     "and already in bar form. That is the whole reason D30 charges the premium on net new grams.")
_mr[0] += 1

m_row("grams_fabricated", "  Net new grams - the metal that must actually be made", "grams",
      lambda i: "=MAX(0,%s-%s*(1-%s))" % (
          mr("grams_bought", i), mr("grams_recycled", i),
          sref("sw_premium_on_gross")), FMT_NUM2)

# THE COST LINE. What Aurumix hands the dealer over and above the spot value of
# the metal - the price of having a bar made, stamped and assayed.
#
# It is charged on FABRICATED grams, not on grams sold, and the gap between the
# two is the whole of D30. Under sw_premium_absorbed = 0 the customer is buying
# metal at spot+premium and this line correctly falls to zero, because Aurumix
# is then passing the cost on rather than carrying it.
m_row("cogs_premium", "COGS - Fabrication premium paid to the dealer", "USD",
      lambda i: "=%s*%s*%s*%s" % (mr("grams_fabricated", i), mr("gold_px", i),
                                  aref("fab_premium"),
                                  sref("sw_premium_absorbed")), FMT_USD, BLACK_BOLD)
COGS_LINES.append("cogs_premium")
note(ws_model, _mr[0],
     "WHAT IS NOT IN THIS LINE, ON PURPOSE. Aurumix charges every customer for the premium but pays it only "
     "on metal it actually has fabricated, so the premium on RECYCLED grams is margin it keeps - roughly "
     "USD 166k over the horizon at the modelled redemption rates. That is real and it is deliberately left "
     "out of the model: it is a consequence of owning the float, it is the client's to take or leave, and "
     "it belongs in the conversation as upside rather than in the base case as revenue.")
_mr[0] += 2

m_row("total_cogs", "TOTAL COST OF GOODS SOLD", "USD",
      lambda i: "=" + "+".join(mr(k, i) for k in COGS_LINES), FMT_USD, BLACK_BOLD, True)
_mr[0] += 2

# ---------------------------------------------------------------- OPEX ------
banner(ws_model, _mr[0],
       "OPERATING EXPENSES - the cost of running the business, not of buying the metal")
_mr[0] += 1

OPEX_LINES = []

# THE FLOAT IS IN THE VAULT TOO. Aurumix's own bars sit on the same shelf as
# the customers' and are charged storage identically. Immaterial at Y7 (0.8% of
# the metal) but 7% of it in Year 1, which is exactly when the cost base is
# tightest - so it is included rather than waved away.
m_row("storage_grams", "  Metal in the vault (customer gold + Aurumix's own float)", "grams",
      lambda i: "=%s+%s" % (mr("grams_custody", i), mr("float_grams", i)), FMT_NUM)

m_row("opex_storage", "OPEX - Vault storage", "USD",
      lambda i: "=MAX(%s*%s*365/12,%s*%s*%s*%s/12)" % (
          sref("storage_min_day"), mr("n", i),
          sref("storage_rate"), mr("storage_grams", i), mr("gold_px", i), mr("n", i)),
      FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_storage")
note(ws_model, _mr[0],
     "MAX(the vault's minimum, the percentage rate on metal held). THIS IS WHAT AURUMIX PAYS THE VAULT. It "
     "is NOT a fee charged to customers - decision 42 makes retail storage free forever, and every "
     "comparator does the same: Paxos states it 'does not charge gold storage fees to its customers at this "
     "time' and Kinesis charges 0%, funding it from a share of transaction fees. THE CLIENT'S ASSUMED "
     "0.8-1.0% CUSTODY FEE IS BOTH UNCOMPETITIVE AND ROUGHLY 7x THE REAL COST.")
_mr[0] += 1

m_row("storage_pct", "  ...as % of metal held per year (memo - watch the minimum bite)", "%/yr",
      lambda i: "=IF(%s=0,0,%s/(%s*%s)*12/%s)" % (
          mr("storage_grams", i), mr("opex_storage", i), mr("storage_grams", i),
          mr("gold_px", i), mr("n", i)), FMT_PCT3)

# ANNUAL FEES LAND ON AN ANNIVERSARY, not in even monthly slices. Both VARA
# supervision and the DMCC licence are invoices that arrive once a year, and
# VARA's is payable IN ADVANCE of conducting the activity - so they are booked
# in the first month of each model year, and in full in the annual columns.
# Smearing them monthly would flatter Year 1 cash, which is the year that
# matters most for the funding ask.
ANNIV = lambda i: "OR(%s=1,%s=0)" % (mr("cal_month", i), mr("cal_month", i))

m_row("opex_vara", "OPEX - VARA annual supervision", "USD",
      lambda i: "=IF(%s,%s/%s,0)" % (ANNIV(i), sref("vara_supervision_aed"), aref("aed_usd")),
      FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_vara")
note(ws_model, _mr[0],
     "PER LICENSED ACTIVITY, and the activity set is unconfirmed - see the Scenario note. Payable IN ADVANCE "
     "of conducting the activity, so Year 1 carries it before a single customer arrives.")
_mr[0] += 1

m_row("opex_dmcc", "OPEX - DMCC company licence", "USD",
      lambda i: "=IF(%s,%s/%s,0)" % (ANNIV(i), sref("dmcc_annual_aed"), aref("aed_usd")),
      FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_dmcc")

# KYC is the one regulatory line that MOVES WITH THE BOOK - it is charged per
# verification, so it is driven by NEW customers rather than by the stock.
m_row("opex_kyc", "OPEX - KYC and AML verification", "USD",
      lambda i: "=MAX(%s*%s,%s*%s)" % (
          sref("kyc_min_month"), mr("n", i),
          sref("kyc_per_check"), mr("new_total", i)), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_kyc")
note(ws_model, _mr[0],
     "MAX(the monthly minimum, per-verification cost x new customers). Charged on SUCCESSFUL verifications "
     "only. The minimum binds below about 162 verifications a month, which is most of the first two years.")
_mr[0] += 1

m_row("opex_oneoff", "OPEX - One-off Year 1 (licence application, incorporation, launch audit)", "USD",
      lambda i: "=IF(%s=1,(%s+%s)/%s+%s,0)" % (
          mr("period_idx", i), sref("vara_application_aed"), sref("dmcc_setup_aed"),
          aref("aed_usd"), sref("launch_audit_usd")), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_oneoff")
note(ws_model, _mr[0],
     "M1 ONLY: the VARA licence application, DMCC incorporation, and the pre-launch technology and smart "
     "contract audit. The VARA fee is spent at submission whether or not the licence is granted - the "
     "application is not processed until it is paid.")
_mr[0] += 1

m_row("opex_insurance", "OPEX - Insurance (PI, D&O, crime)", "USD",
      lambda i: "=%s*%s/12" % (sref("insurance_usd"), mr("n", i)), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_insurance")
note(ws_model, _mr[0],
     "MANDATORY under Company Rulebook VI.D, which names professional indemnity, directors' and officers', "
     "and CRIME COVER FOR VIRTUAL ASSETS IN HOT WALLETS. No minimum sum is prescribed - VARA judges adequacy "
     "at licensing. THE METAL IS NOT COVERED HERE: vault storage already includes insurance on the gold.")
_mr[0] += 1

m_row("opex_audit", "OPEX - Audit and reserve attestation", "USD",
      lambda i: "=%s*%s/12" % (sref("audit_usd"), mr("n", i)), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_audit")
note(ws_model, _mr[0],
     "A SIX-MONTHLY independent audit plus an ANNUAL financial-statement audit, both mandatory under "
     "Issuance Rulebook III.D, each followed by a Senior Management attestation to VARA. UNDER OPTION A the "
     "six-monthly engagement audits TOKEN SUPPLY, not a reserve pool - limb (b) of III.D.2.a applies only "
     "'if applicable' and direct ownership has no Reserve Assets. THE FEE IS UNPRICED BY EVERY ISSUER, "
     "AUDITOR AND REGULATOR CHECKED - this is a placeholder.")
_mr[0] += 1

m_row("opex_techaudit", "OPEX - Technology audit and penetration testing", "USD",
      lambda i: "=%s*%s/12" % (sref("techaudit_usd"), mr("n", i)), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_techaudit")
note(ws_model, _mr[0],
     "ANNUAL AND RECURRING, not a launch one-off - Technology Rulebook E requires an independent third-party "
     "vulnerability assessment and penetration test at least annually AND before any new system, "
     "application or product goes live. THE PER-LAUNCH TRIGGER IS NOT MODELLED: it is a step cost tied to a "
     "product roadmap that does not exist yet, so a year with several launches costs more than this shows.")
_mr[0] += 1

# STREAM 0. A mandatory, uncapped, zero-revenue cost that scales with the book -
# VARA Annex 2 III.E.4 forbids charging ANY fee on redemption, so nothing offsets
# it. The event count has been carried as a memo since the revenue build for
# exactly this moment.
m_row("opex_redemption", "OPEX - Redemption handling (no fee may be charged)", "USD",
      lambda i: "=%s*%s" % (mr("redemptions", i), aref("redemption_cost")),
      FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_redemption")
note(ws_model, _mr[0],
     "USD 3.20 per event: Sumsub AML RE-SCREENING at 1.85 plus operational handling at about 1.35. The "
     "outbound bank transfer fee was removed - that is the customer's to bear. NOT A DOUBLE COUNT WITH THE "
     "KYC LINE: that one charges new customers at onboarding, this one charges a different event. "
     "THE ASYMMETRY IS REGULATORY, NOT A CHOICE: Aurumix may pass through the cost of taking money IN (the "
     "rail, D31) but is forbidden from passing through the cost of paying it OUT. Immaterial at Base rates; "
     "it becomes material exactly in a redemption spike, which is when cash is scarcest.")
_mr[0] += 1

# BUILD IN Y1-Y2, MAINTENANCE FROM Y3, at the client's instruction. Two rows
# rather than one so the shape is visible: a business that spends heavily to
# build and then settles, which is what the cash profile actually looks like.
m_row("opex_techbuild", "OPEX - Technology build (Y1-Y2)", "USD",
      lambda i: "=IF(%s=1,%s/12*%s,IF(%s=2,%s/12*%s,0))" % (
          mr("year", i), sref("tech_build_y1"), mr("n", i),
          mr("year", i), sref("tech_build_y2"), mr("n", i)), FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_techbuild")
note(ws_model, _mr[0],
     "THE APP, LEDGER, WALLET, TOKEN CONTRACT AND EVERY INTEGRATION - card, KYC, payment rail, custody. Y2 "
     "carries the card and credit layer, which activates at period 13, so it is a second product going live "
     "rather than a tidy-up. THE SMART CONTRACT AUDIT IS IN THE YEAR 1 ONE-OFF LINE, not here.")
_mr[0] += 1

m_row("opex_techmaint", "OPEX - Technology maintenance (Y3 onward)", "USD",
      lambda i: "=IF(%s>=3,%s/12*%s,0)" % (mr("year", i), sref("tech_maint"), mr("n", i)),
      FMT_USD, BLACK_BOLD)
OPEX_LINES.append("opex_techmaint")
note(ws_model, _mr[0],
     "FLAT FROM Y3, WHICH UNDERSTATES THE BACK YEARS - the book grows from about 4,000 paying customers to "
     "over 70,000 across that span while this line does not move. "
     "ON-CHAIN MINTING IS NOT IN HERE AND IS NOT ZERO: about 855,000 mint events a year at Y7. On an L2 "
     "that is roughly USD 8,500/yr; on Ethereum mainnet it is about USD 1.7m/yr. NOBODY HAS RECORDED WHICH "
     "CHAIN AURUMIX ISSUES ON, so it is left out rather than guessed.")
_mr[0] += 1

m_row("total_opex", "TOTAL OPERATING EXPENSES", "USD",
      lambda i: "=" + "+".join(mr(k, i) for k in OPEX_LINES), FMT_USD, BLACK_BOLD, True)

_mr[0] += 2

# ------------------------------------------------------- ICS BENEFIT COSTS --
banner(ws_model, _mr[0],
       "ICS BENEFIT COSTS - discounts given away, out of streams that already exist")
_mr[0] += 1

BENEFIT_LINES = []

# THE QUALIFYING SHARE, and it RAMPS. Accounts take about eight months to reach
# a tier, so this is well under the 55% terminal rate in the early years - which
# is why benefits cost almost nothing at launch and become material later. It is
# derived from the existing ICS row rather than re-deriving 55%, so there is one
# source for the qualification rate and not two.
m_row("qual_share", "  Qualifying share of the book (memo - ramps to 55%)", "% of accounts",
      lambda i: "=IF(%s+%s=0,0,%s/(%s+%s))" % (
          mr("paying", i), mr("holders", i), mr("qualified", i),
          mr("paying", i), mr("holders", i)), FMT_PCT2)

# SIP ONLY - spot earns no ICS, so Stream 1b is deliberately absent here.
m_row("ben_entry", "BENEFIT - Entry fee discount (SIP only)", "USD",
      lambda i: "=%s*%s*%s" % (mr("s1a", i), mr("qual_share", i), sref("disc_entry")),
      FMT_USD, BLACK_BOLD)
BENEFIT_LINES.append("ben_entry")
note(ws_model, _mr[0],
     "STREAM 1a ONLY. Spot earns no ICS, so Stream 1b carries no discount and must never be added here. "
     "A BLENDED RATE, not a tier ladder: the model has no tier distribution, and inventing one would put a "
     "fabricated mix under every benefit figure.")
_mr[0] += 1

m_row("ben_card", "BENEFIT - Card fee discounts (FX, ATM, issuance)", "USD",
      lambda i: "=%s*%s*%s" % (mr("s4", i), mr("qual_share", i), sref("disc_card")),
      FMT_USD, BLACK_BOLD)
BENEFIT_LINES.append("ben_card")

m_row("ben_rebate", "BENEFIT - Gold rebate", "USD",
      lambda i: "=(%s+%s)*%s*%s" % (mr("s2", i), mr("s4", i), mr("qual_share", i),
                                    sref("disc_rebate")), FMT_USD, BLACK_BOLD)
BENEFIT_LINES.append("ben_rebate")
note(ws_model, _mr[0],
     "A SHARE OF ALL CARD REVENUE - interchange PLUS cardholder fees - which keeps the rebate inside its own "
     "funding line by construction. Interchange alone was the wrong base: it is the smaller half, and it is "
     "already net of the partner's ~55% share, so it measures what is LEFT rather than what the customer "
     "GENERATED. Sanity-check the rate against card spend, not against this line: 5% here is 0.35% of "
     "spend, against a corpus ladder of 0.15 / 0.45 / 0.75% by tier.")
_mr[0] += 1

m_row("rebate_pct_spend", "  ...as % of card spend (memo - check against the 0.15/0.45/0.75 ladder)", "%",
      lambda i: "=IF(%s=0,0,%s/%s)" % (mr("card_spend_usd", i), mr("ben_rebate", i),
                                       mr("card_spend_usd", i)), FMT_PCT2)

m_row("ben_family", "BENEFIT - Family wallet and will discount", "USD",
      lambda i: "=%s*%s*%s" % (mr("s3", i), mr("qual_share", i), sref("disc_family")),
      FMT_USD, BLACK_BOLD)
BENEFIT_LINES.append("ben_family")

m_row("total_benefits", "TOTAL ICS BENEFIT COSTS", "USD",
      lambda i: "=" + "+".join(mr(k, i) for k in BENEFIT_LINES), FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "CONTRA-REVENUE, NOT CASH. Every line here is revenue given away rather than money paid out, and the "
     "streams above are reported GROSS so the giveaway stays visible. Netting them into the streams would "
     "hide the single largest lever the client controls on the benefit ladder.")
_mr[0] += 1

_mr[0] += 2

# ------------------------------------------------------- ACQUISITION COSTS --
banner(ws_model, _mr[0],
       "ACQUISITION COSTS - what it costs to put a customer on the book")
_mr[0] += 1

ACQ_LINES = []

# MARKETING IS ALREADY BEING SPENT. The schedule on Assumptions drives the
# direct channel; until now it bought customers and was charged to nothing.
# It is booked ONCE, here, and must never also appear in an opex table - which
# is the error v1.0 made.
m_row("acq_marketing", "ACQ - Marketing spend", "USD",
      lambda i: "=INDEX(Assumptions!$B$%d:$B$%d,%s)/12*%s" % (
          AROW["mktg_y1"], AROW["mktg_y7"], mr("year", i), mr("n", i)),
      FMT_USD, BLACK_BOLD)
ACQ_LINES.append("acq_marketing")
note(ws_model, _mr[0],
     "A DECISION VARIABLE, not an output. It is the input that drives the direct channel, so it is booked "
     "here ONCE and never in opex. THE REVENUE-ONLY MODEL FLATTERED THIS BADLY: raising the budget 50% on "
     "2026-08-21 appeared as +28.5% customers with no offsetting entry anywhere. This row is that "
     "offsetting entry.")
_mr[0] += 1

# Agent-acquired share of the book, cumulative. The model does not cohort
# customers by channel, so this is the honest approximation of it.
m_row("agent_new", "  Agent-driven acquisitions this period", "accounts",
      lambda i: "=" + "+".join(mr("agent_" + r, i) for r in REGIONS), FMT_NUM2)
_agc = _mr[0]
m_row("agent_cum", "  Cumulative agent-driven acquisitions", "accounts",
      lambda i, ac=_agc: "=%s" % mr("agent_new", i) if i == 0
      else "=%s+%s" % (pcell(ac, i - 1), mr("agent_new", i)), FMT_NUM)
m_row("agent_share", "  Agent-acquired share of the book (memo)", "% of accounts",
      lambda i: "=IF(%s=0,0,%s/%s)" % (mr("cum_ever", i), mr("agent_cum", i),
                                       mr("cum_ever", i)), FMT_PCT2)

m_row("acq_agent", "ACQ - Agent commission", "USD",
      lambda i: "=(%s+%s)*%s*%s" % (mr("s1a", i), mr("s1b", i), mr("agent_share", i),
                                    sref("agent_commission")), FMT_USD, BLACK_BOLD)
ACQ_LINES.append("acq_agent")
note(ws_model, _mr[0],
     "10% of the entry fee, ONGOING - renewal commission for as long as the customer keeps paying, not a "
     "one-off at acquisition. Applied to the agent-acquired SHARE of entry-fee revenue, approximated from "
     "cumulative acquisitions by channel because the model does not cohort customers by channel. "
     "THE UPLINE OVERRIDE IS NOT MODELLED: the network is three tiers, and whether 10% is the total across "
     "all levels or just the first is unresolved - the difference is one commission or three.")
_mr[0] += 1

m_row("ref_new", "  Referral-driven acquisitions this period", "accounts",
      lambda i: "=" + "+".join(mr("ref_" + r, i) for r in REGIONS), FMT_NUM2)

# Priced PER REGION, because the reward is a share of the REFEREE'S OWN fee and
# tickets differ by region - a blended ticket would misprice every region.
m_row("acq_referral", "ACQ - Referral rewards", "USD",
      lambda i: "=(" + "+".join(
          "%s*%s" % (mr("ref_" + r, i), sref("ticket_" + r)) for r in REGIONS)
      + ")*6*%s*%s" % (aref("entry_fee"), sref("referral_reward")), FMT_USD, BLACK_BOLD)
ACQ_LINES.append("acq_referral")
note(ws_model, _mr[0],
     "30% of the entry fee the referee pays over SIX qualifying contributions, credited in grams. Priced per "
     "region because tickets differ and the reward is a share of the referee's own fee. "
     "BOOKED AT ACQUISITION, PAID AT THE GATE: the real payment lands six contributions later, so no earlier "
     "than about month 19. Annual totals barely move; the monthly cash profile does.")
_mr[0] += 1

m_row("total_acq", "TOTAL ACQUISITION COSTS", "USD",
      lambda i: "=" + "+".join(mr(k, i) for k in ACQ_LINES), FMT_USD, BLACK_BOLD, True)

m_row("cac_blended", "  Blended cost per new customer (memo)", "USD",
      lambda i: "=IF(%s=0,0,%s/%s)" % (mr("new_total", i), mr("total_acq", i),
                                       mr("new_total", i)), FMT_USD2)
note(ws_model, _mr[0],
     "TOTAL acquisition cost over ALL new customers, including the organic ones nobody paid for. It is "
     "therefore BELOW the paid CAC by construction, and it is the number that matters for unit economics - "
     "the marketing CAC prices only the channel it buys.")
_mr[0] += 1

_mr[0] += 2

# ----------------------------------------------------- CARD PROGRAMME -------
# ITS OWN BAND, NOT INSIDE OPEX, and that placement is the point. The card is
# the largest revenue source in the model; bury its costs in an opex table and
# nobody can answer whether it pays for itself. The memo at the foot of this
# block is the whole reason the band exists.
banner(ws_model, _mr[0],
       "CARD PROGRAMME COSTS - what the card costs to run, kept next to what it earns")
_mr[0] += 1

CARD_LINES = []

m_row("newcards_total", "  Cards newly issued this period", "cards",
      lambda i: "=" + "+".join(mr("newcards_" + r, i) for r in REGIONS), FMT_NUM2)

# FLAT FROM LAUNCH. It does not scale down because the book is small - that is
# what a minimum commitment means, and it is why the programme loses money for
# its first year or so.
m_row("card_fixed", "CARD - NymCard platform and scheme fees", "USD",
      lambda i: gate("s2", i, "%s*%s/12" % (sref("card_fixed_usd"), mr("n", i))),
      FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_fixed")
note(ws_model, _mr[0],
     "NO BIN SPONSORSHIP FEE IN HERE. NymCard is a principal member of both Visa and Mastercard, so it "
     "sponsors the BIN under its OWN membership - there is no bank to pay for it, and the line fell from "
     "USD 120,000/yr to USD 30,000 when that component came out. What remains is the platform fee and "
     "scheme assessments, which no provider waives. "
     "A BANK IS STILL NEEDED FOR THE CREDIT: lending is a different regime and a payment services licence "
     "does not reach it. BENCHMARKED, NOT QUOTED - NymCard publishes no pricing.")
_mr[0] += 1

m_row("card_setup", "CARD - Programme setup (one-off at launch)", "USD",
      lambda i: "=IF(%s=%s,%s,0)" % (mr("period_idx", i), aref("act_s2"),
                                     sref("card_setup_usd")), FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_setup")

m_row("card_production", "CARD - Card production and delivery", "USD",
      lambda i: "=%s*%s" % (mr("newcards_total", i), sref("card_per_card")),
      FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_production")

m_row("card_processing", "CARD - Authorisation and switching", "USD",
      lambda i: "=%s*%s" % (mr("auths", i), sref("card_per_auth")), FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_processing")

m_row("card_fraud", "CARD - Fraud and chargebacks", "USD",
      lambda i: "=%s*%s" % (mr("card_spend_usd", i), sref("card_fraud_bps")),
      FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_fraud")

# DERIVED, NOT ASSUMED. Oman, Bahrain and India customers carry a UAE-ISSUED
# card, so every transaction they make is cross-border by definition. Only the
# UAE residents' own travel and foreign online spend needs an assumption.
m_row("xborder_spend", "  Cross-border card spend (non-UAE in full, plus UAE spend abroad)", "USD",
      lambda i: "=%s+%s+%s*%s" % (
          mr("spendusd_gulf", i), mr("spendusd_india", i),
          mr("spendusd_uae", i), sref("uae_spend_abroad")), FMT_USD)

m_row("card_xborder", "CARD - Cross-border scheme assessment", "USD",
      lambda i: "=%s*%s" % (mr("xborder_spend", i), sref("xborder_rate")),
      FMT_USD, BLACK_BOLD)
CARD_LINES.append("card_xborder")
note(ws_model, _mr[0],
     "VISA ISA / MASTERCARD CROSS-BORDER ASSESSMENT, charged to the ISSUER whenever merchant country "
     "differs from issuer country. ABOUT 63% OF CARD SPEND QUALIFIES, and almost none of that is an "
     "assumption: a UAE-issued card spent in India is cross-border on every tap, and India is half the card "
     "book. "
     "THE RATE EXCEEDS AURUMIX'S INTERCHANGE SHARE - 1.40% against 0.72% of spend retained - so cross-border "
     "interchange does not cover its own scheme fee. The FX margin in stream 4 is what closes the gap, "
     "which means that line is largely COST RECOVERY rather than margin.")
_mr[0] += 1

m_row("total_card", "TOTAL CARD PROGRAMME COSTS", "USD",
      lambda i: "=" + "+".join(mr(k, i) for k in CARD_LINES), FMT_USD, BLACK_BOLD, True)

# THE QUESTION THE CLIENT WILL ACTUALLY ASK. Above 100% the card is consuming
# more than it earns; the crossing point is the answer to "when does the card
# start paying for itself".
m_row("card_cost_ratio", "  Card cost as % of card revenue (above 100% = the card is not paying for itself)",
      "%",
      lambda i: "=IF(%s+%s=0,0,%s/(%s+%s))" % (
          mr("s2", i), mr("s4", i), mr("total_card", i), mr("s2", i), mr("s4", i)),
      FMT_PCT, BLACK_BOLD)
note(ws_model, _mr[0],
     "CARD REVENUE IS STREAM 2 PLUS STREAM 4 - interchange and cardholder fees. It EXCLUDES stream 5, "
     "because credit revenue is a separate partner arrangement and would flatter this ratio. Expect well "
     "above 100% at launch: the fixed fee lands in full from month 13 while the card book is nearly empty, "
     "which is exactly the 12-18 months of losses the brief describes.")
_mr[0] += 1

m_row("cost_measured", "  Sub-total - costs actually modelled", "USD",
      lambda i: "=%s+%s+%s+%s+%s" % (mr("total_cogs", i), mr("total_opex", i),
                                     mr("total_benefits", i), mr("total_acq", i),
                                     mr("total_card", i)), FMT_USD, BLACK_BOLD)

m_row("cost_contingency", "  Contingency for cost families not yet built", "USD",
      lambda i: "=%s*%s" % (mr("cost_measured", i), sref("cost_contingency")),
      FMT_USD, BLACK_BOLD)
note(ws_model, _mr[0],
     "A PLACEHOLDER FOR HEADCOUNT, LEGAL AND TRUST, SECURITY, CORPORATE, TAX AND THE ON-CHAIN MINTING COST - "
     "AND IT DOES NOT COVER THEM. Headcount alone is anchored in the brief at roughly USD 588,000 in Year 1, "
     "about 78% of the whole cost base as built; 15% of that base is about USD 112,000. THE CONTINGENCY IS "
     "WORTH A FIFTH OF ONE OF THE SIX MISSING BLOCKS. Every profit figure below is an upper bound.")
_mr[0] += 1

m_row("total_cost", "TOTAL COST BASE (modelled + contingency)", "USD",
      lambda i: "=%s+%s" % (mr("cost_measured", i), mr("cost_contingency", i)),
      FMT_USD, BLACK_BOLD, True)
# NO GROSS PROFIT LINE YET, DELIBERATELY. COGS is one of six cost families and
# the only one built so far, so revenue-less-COGS would read as a margin the
# business does not have - 93% at Y7, against a cost base still missing opex,
# acquisition, the card programme, benefit costs and tax. The line arrives when
# there is a cost base to subtract, not before.

# Ties the cost band back to the REVENUE band's own inflow rows, so the two
# cannot drift apart. Non-circular: the left side is built from sip/spot
# contributions and the entry fee, the right side from grams and the premium.
m_row("cogs_check", "CHECK: cash for metal ties to grams delivered x the price paid (must be 0)", "delta",
      lambda i: "=(%s+%s)*(1-%s)-%s*%s*(1+%s*(1-%s))" % (
          mr("sip_inflow", i), mr("spot_inflow", i), aref("entry_fee"),
          mr("grams_bought", i), mr("gold_px", i), aref("fab_premium"),
          sref("sw_premium_absorbed")), FMT_NUM3, BLACK_BOLD)
note(ws_model, _mr[0],
     "THE ANTI-DOUBLE-COUNT ASSERTION. Either the customer's cash buys metal at spot+premium and the COGS "
     "line above is zero, or it buys metal at spot and the COGS line carries the premium. This row ties to "
     "zero under both settings and can only break if some future edit makes the premium bite twice - which "
     "is exactly the defect 29f98e0 was written to fix.")

# ============================================================================
# REGULATORY CAPITAL
#
# NEITHER A COST NOR INVENTORY. The float is metal Aurumix buys and can sell;
# this is equity it must SUBSCRIBE AND KEEP SUBSCRIBED to hold its licences. It
# earns nothing, it is not spent, and it never reaches the P&L - the same
# treatment the float gets, for a different reason.
#
# It sits after the cost base because the 25%-of-overheads test reads total
# operating expenses, which are computed above.
# ============================================================================
_mr[0] += 2
divider(ws_model, _mr[0],
        "REGULATORY CAPITAL - locked, not spent. Subscribed equity, not an expense")
_mr[0] += 2

banner(ws_model, _mr[0],
       "PAID-UP CAPITAL AND LIQUIDITY - what VARA requires you to keep locked up")
_mr[0] += 1

# ONE-OFFS ARE EXCLUDED. Fixed annual overheads means the RUNNING cost of the
# business; the M1 licence application and incorporation are neither fixed nor
# recurring, and leaving them in would spike the test in the one period where
# capital is tightest anyway.
m_row("cap_issuance", "  Paid-up capital - Category 1 VA Issuance", "USD",
      lambda i: "=%s/%s" % (sref("paidup_issuance_aed"), aref("aed_usd")), FMT_USD)
note(ws_model, _mr[0],
     "THE AED 1.5m FLOOR, NOT THE 2% ESCALATOR. Option A carries no Reserve Assets to take 2% of, so limb "
     "(b) of Rule G.1 is inapplicable and the floor governs. CHOOSING OPTION A REMOVES THE ESCALATOR, NOT "
     "THE FLOOR - worth stating plainly to the client, who may believe otherwise.")
_mr[0] += 1

m_row("cap_activity", "  Paid-up capital - licensed activities (Broker-Dealer)", "USD",
      lambda i: "=%s/%s" % (sref("paidup_activity_aed"), aref("aed_usd")), FMT_USD)
note(ws_model, _mr[0],
     "THE FIXED AED FLOOR ONLY. CUMULATIVE with the issuance requirement rather than an alternative to it - "
     "Part VI.A requires capital for EACH licensed activity. "
     "NOT YET APPLIED, AND DELIBERATELY: the rule is the HIGHER of the AED floor or 25% OF FIXED ANNUAL "
     "OVERHEADS (15% where custody is held). The overheads test is left out until the opex build is "
     "complete - with only storage, VARA, DMCC and KYC in the model, overheads are a fraction of the real "
     "figure and the test would read as comfortably passed when nobody has checked. REINSTATE IT ONCE "
     "HEADCOUNT AND THE REST OF OPEX LAND: at 25%, the test overtakes the AED 600,000 floor at roughly "
     "AED 2.4m of annual overheads, which a real cost base may well reach.")
_mr[0] += 1

m_row("cap_total", "TOTAL PAID-UP CAPITAL REQUIRED", "USD",
      lambda i: "=%s+%s" % (mr("cap_issuance", i), mr("cap_activity", i)),
      FMT_USD, BLACK_BOLD, True)

m_row("nla_required", "  Net liquid assets required (PROVISIONAL - opex incomplete)", "USD",
      lambda i: "=%s*(SUMIF(%s,%s,%s)-SUMIF(%s,%s,%s))/12" % (
          sref("nla_months"), mrng("year"), mr("year", i), mrng("total_opex"),
          mrng("year"), mr("year", i), mrng("opex_oneoff")), FMT_USD)
note(ws_model, _mr[0],
     "1.2 x MONTHLY operating expenses, tested continuously. NOT ADDED to paid-up capital: the same cash can "
     "satisfy both, so summing them would overstate the funding need. "
     "PROVISIONAL AND UNDERSTATED. It reads off the opex total, and opex currently carries only storage, "
     "VARA, DMCC and KYC - headcount, insurance, audit, technology and the rest are still to come. THIS "
     "FIGURE WILL RISE, probably several-fold. It is kept visible rather than removed because the "
     "requirement itself is real and appears nowhere in the corpus; the number is not yet usable.")
_mr[0] += 1

m_row("cap_liquidity_check", "CHECK: paid-up capital covers the liquidity test (must be >= 0)", "USD",
      lambda i: "=%s-%s" % (mr("cap_total", i), mr("nla_required", i)), FMT_USD, BLACK_BOLD)

# THE FUNDING PICTURE. Two different kinds of money that are both unavailable
# for spending, shown together because the client asks HOW MUCH DO I NEED, not
# how is it classified.
m_row("capital_tied_up", "TOTAL CAPITAL TIED UP (float + paid-up capital + card prefunding)", "USD",
      lambda i: "=%s+%s+%s" % (mr("float_usd", i), mr("cap_total", i),
                               mr("card_prefund", i)), FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "THE FLOAT IS METAL YOU OWN AND CAN SELL; THE PAID-UP CAPITAL IS EQUITY YOU MUST KEEP SUBSCRIBED TO "
     "HOLD A LICENCE. Both are money unavailable to spend, which is why they are added - but they are NOT "
     "the same kind of money, and the difference matters on a wind-down.")
_mr[0] += 1

# ============================================================================
# NET PROFIT
#
# Revenue less every cost that has been built, plus the contingency.
#
# THE CASH VIEW WAS REMOVED 2026-08-26 at the client's instruction. It carried
# the change in working and regulatory capital, net and cumulative cash flow,
# and a peak funding figure. The capital balances themselves are untouched and
# still sit in the WORKING CAPITAL and REGULATORY CAPITAL bands - what went is
# the translation of profit into cash. Anyone asking "how much money do I need"
# must read those bands directly; this band answers only "does it earn".
#
# ⚠ HEADCOUNT IS NOT IN THE COST BASE. Read the profit line as an upper bound.
# ============================================================================
_mr[0] += 2
divider(ws_model, _mr[0],
        "NET PROFIT - revenue less every cost modelled")
_mr[0] += 2

banner(ws_model, _mr[0], "NET PROFIT - revenue less every cost modelled, plus contingency")
_mr[0] += 1

m_row("net_profit", "NET PROFIT", "USD",
      lambda i: "=%s-%s" % (mr("total_rev", i), mr("total_cost", i)),
      FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "AN UPPER BOUND, NOT A FORECAST. Headcount, legal and trust, security, corporate, tax and the on-chain "
     "minting cost are all still absent, covered only by a 15% contingency that is worth a fraction of "
     "headcount alone. THE REAL NUMBER IS LOWER AND THE MODEL CANNOT YET SAY BY HOW MUCH.")
_mr[0] += 1

m_row("net_margin", "  Net margin", "% of revenue",
      lambda i: "=IF(%s=0,0,%s/%s)" % (mr("total_rev", i), mr("net_profit", i),
                                       mr("total_rev", i)), FMT_PCT)

_cnp = _mr[0]
m_row("cum_profit", "CUMULATIVE NET PROFIT", "USD",
      lambda i, cp=_cnp: "=%s" % mr("net_profit", i) if i == 0
      else "=%s+%s" % (pcell(cp, i - 1), mr("net_profit", i)), FMT_USD, BLACK_BOLD, True)
_mr[0] += 1

# ============================ UNIT ECONOMICS ================================
# ADDED 2026-08-31. The model computed revenue per paying customer and cost per
# NEW customer and never divided one by the other, so the first question any
# client asks - what is a customer worth against what they cost - had to be
# assembled by hand from two memo rows that are not even on the same basis.
#
# NOTHING HERE IS A NEW ASSUMPTION. Every row below is arithmetic on rows that
# already exist, which is why it belongs in this model rather than in the
# Phase 5 simulation: it needs no distribution, no cohort and no new input.
#
# NO LTV AND NO LTV:CAC HERE, DELIBERATELY. CLIENT DECISION 2026-08-31, and it
# is the right one. Both were built and then removed: they require an average
# customer LIFE, and the only life this model can produce is a single blended
# figure derived from the whole-book leak. That figure swings from 13 to 19
# months across the horizon on nothing but the AGE MIX of the base - a fast
# acquisition year floods the book with young accounts, which churn hardest,
# and the measured life falls even though no customer behaved differently.
# An LTV resting on it would look precise and would not be. THE HONEST ANSWER
# NEEDS COHORT RESOLUTION - track the January 2027 joiners as their own decaying
# group - which is a Phase 5 simulation job under the v3.0 scope decision.
#
# WHAT SURVIVES NEEDS NO LIFE ESTIMATE AT ALL. Payback is CAC over monthly
# contribution. The treadmill row is subtraction. The funding need is a running
# total. None of them ask how long anyone stays, so none of them get better
# with cohorts - which is exactly why they belong here and the LTV rows do not.
banner(ws_model, _mr[0], "UNIT ECONOMICS - payback, the acquisition treadmill and the funding need")
_mr[0] += 1

_pay_r = MROW["paying"]
m_row("attrition", "  Customers lost to holders this period", "accounts",
      lambda i, pr=_pay_r: "=MAX(0,%s-%s)" % (mr("new_total", i), mr("paying", i)) if i == 0
      else "=MAX(0,%s+%s-%s)" % (pcell(pr, i - 1), mr("new_total", i), mr("paying", i)),
      FMT_NUM)
note(ws_model, _mr[0],
     "OPENING BASE PLUS NEW LESS CLOSING BASE. Nobody counts leavers directly; they are what is left when "
     "the balance does not tie. THIS IS THE END OF A PAYING RELATIONSHIP, NOT OF THE CUSTOMER: a lapsed "
     "customer becomes a holder, keeps their gold and may still carry a card, so the tail is thin but it "
     "is not zero. Deliberately NOT converted into an average life here - see the section note above.")
_mr[0] += 1

m_row("profit_preacq", "PROFIT BEFORE ACQUISITION COST", "USD",
      lambda i: "=%s-%s+%s" % (mr("total_rev", i), mr("total_cost", i), mr("total_acq", i)),
      FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "THE RIGHT NUMERATOR TO SET AGAINST CAC, because acquisition is the thing being paid for and must not "
     "sit on both sides. DELIBERATELY CONSERVATIVE: it still carries the FIXED cost base - licences, "
     "insurance, technology - so it understates what a marginal customer contributes and improves with "
     "scale by construction. A true marginal contribution would need a fixed/variable split this model "
     "does not carry, and inventing one would put a judgement call under the payback below.")
_mr[0] += 1

m_row("contrib_per_cust", "  Contribution per paying customer (annualised)", "USD",
      lambda i: "=IF(%s=0,0,%s/%s/%s*12)" % (mr("paying", i), mr("profit_preacq", i),
                                             mr("paying", i), mr("n", i)), FMT_USD2)

m_row("payback_months", "  Payback on contribution", "months",
      lambda i: '=IF(%s<=0,"",%s/(%s/12))' % (mr("contrib_per_cust", i), mr("cac_blended", i),
                                              mr("contrib_per_cust", i)), FMT_NUM2)
note(ws_model, _mr[0],
     "CAC OVER MONTHLY CONTRIBUTION - IT NEEDS NO LIFE ESTIMATE, which is why it is here and an LTV:CAC "
     "ratio is not. AGAINST THE BLENDED COST PER NEW CUSTOMER, which includes the organic and referred "
     "arrivals nobody paid for, so it is the honest denominator for the book as a whole. Set against the "
     "UAE PAID CAC alone the payback is materially longer - the marketing CAC prices only the channel it "
     "buys. Blank where contribution is negative, which is the early years.")
_mr[0] += 1

m_row("new_per_net", "NEW CUSTOMERS ACQUIRED PER NET CUSTOMER ADDED", "accounts",
      lambda i, pr=_pay_r: '=""' if i == 0
      else '=IF(%s-%s<=0,"",%s/(%s-%s))' % (
          mr("paying", i), pcell(pr, i - 1), mr("new_total", i),
          mr("paying", i), pcell(pr, i - 1)), FMT_NUM2)
note(ws_model, _mr[0],
     "THE TREADMILL, AND THE SINGLE MOST IMPORTANT STRUCTURAL FACT IN THIS MODEL. Growth here is BOUGHT, "
     "not compounded: by Y7 the base grows by a small fraction of what is acquired and the rest leaks to "
     "holders. Blank where the base shrinks, which is a worse outcome than a large ratio, not a better one.")
_mr[0] += 1

m_row("funding_now", "  Funding required to date (P&L deficit + capital tied up)", "USD",
      lambda i: "=MAX(0,-%s)+%s" % (mr("cum_profit", i), mr("capital_tied_up", i)), FMT_USD)
_fpk = _mr[0]
m_row("funding_peak", "PEAK FUNDING NEED", "USD",
      lambda i, fp=_fpk: "=%s" % mr("funding_now", i) if i == 0
      else "=MAX(%s,%s)" % (pcell(fp, i - 1), mr("funding_now", i)),
      FMT_USD, BLACK_BOLD, True)
note(ws_model, _mr[0],
     "A RUNNING PEAK, so it never falls once reached - the business must be funded for the worst point it "
     "passes through, not for where it ends. IT INHERITS THE NET PROFIT CAVEAT: headcount, legal, security, "
     "corporate and tax are still absent from the cost base, so the deficit is an UNDERSTATEMENT and the "
     "real funding need is higher. The capital component is locked, not spent, and comes back.")
_mr[0] += 1

# ---- no reserved acquisition rows remain -----------------------------------
# The single deferred row here was the national agent pool, which had to be
# written late because the region blocks divided into it. Agent output is now
# computed inside each region block from that region's own headcount, so
# nothing is left to back-fill.
CEIL = "+".join(aref("ceil_" + k) for k in REGIONS)

ws_model.freeze_panes = "C8"


# ============================================================================
# SUMMARY
# ============================================================================
ws_summ["A1"] = "=Cover!$B$2"
ws_summ["A1"].font = SHEET_TITLE
ws_summ["A2"] = "Summary"
ws_summ["A2"].font = SHEET_SUB
widths(ws_summ, {"A": 46, "B": 4, **{ycol(y): 16 for y in range(1, N_YEARS + 1)}})
for y in range(1, N_YEARS + 1):
    c = ws_summ["%s4" % ycol(y)]
    c.value, c.fill, c.font, c.border = "Y%d" % y, HEADER_FILL, HEADER_FONT, BORDER
    c.alignment = Alignment(horizontal="center")


def yr_expr(key, y, how="sum"):
    """Y1/Y2 aggregate the twelve monthly columns; Y3-Y7 are single columns."""
    if y <= 2:
        lo, hi = (y - 1) * 12, y * 12 - 1
        if how == "sum":
            return "=SUM(Model!%s:%s)" % (mr(key, lo), mr(key, hi))
        return "=Model!%s" % mr(key, hi)
    return "=Model!%s" % mr(key, N_MONTHLY + y - 3)


_sy = [6]


def s_block(title, rows):
    """Emit a Summary block and RETURN {key: row}, so anything that needs to
    reference these rows reads the real row number instead of a hardcoded one.
    Hardcoding is what broke the revenue mix when a block was inserted above it."""
    r = _sy[0]
    section(ws_summ, r, title, span=11)
    _sy[0] = r + 1
    placed = {}
    for key, label, how, fmt, bold in rows:
        rr = _sy[0]
        ws_summ["A%d" % rr] = label
        ws_summ["A%d" % rr].font = BLACK_BOLD if bold else BLACK
        for y in range(1, N_YEARS + 1):
            c = ws_summ["%s%d" % (ycol(y), rr)]
            c.value = yr_expr(key, y, how)
            c.font, c.number_format = GREEN, fmt
            if bold:
                c.border = TOTALS_BORDER
        placed[key] = rr
        _sy[0] += 1
    _sy[0] += 1
    return placed


REG_ROWS = s_block("REVENUE BY REGION", [
    (s, "  %s" % RLAB[s[4:]], "sum", FMT_USD, False) for s in SUBTOTALS
] + [
    ("s6", "  Non-regional: B2B platform fee", "sum", FMT_USD, False),
    ("total_rev", "TOTAL REVENUE", "sum", FMT_USD, True),
])

STREAM_ROWS = s_block("REVENUE BY STREAM", [
    ("s1a", "Stream 1a: Entry fee - SIP", "sum", FMT_USD, False),
    ("s1b", "Stream 1b: Entry fee - SPOT", "sum", FMT_USD, False),
    ("s2", "Stream 2: Card interchange", "sum", FMT_USD, False),
    ("s3", "Stream 3: Family plan and Digital Will", "sum", FMT_USD, False),
    ("s4", "Stream 4: Cardholder fees", "sum", FMT_USD, False),
    ("s5", "Stream 5: Lending revenue share", "sum", FMT_USD, False),
    ("s6", "Stream 6: B2B platform fee", "sum", FMT_USD, False),
    # Same number as the regional total above - shown again as a reconciliation,
    # labelled distinctly so the two are never read as separate figures.
    ("total_rev", "Total by stream (reconciles to the regional total above)", "sum", FMT_USD, True),
])

# Mix rows reference the ACTUAL row numbers returned above. These were once
# hardcoded, and inserting a block above them silently pointed every percentage
# at the wrong numerator and the wrong denominator.
TOT_ROW = REG_ROWS["total_rev"]
r = _sy[0]
section(ws_summ, r, "REVENUE MIX (% of total)", span=11)
_sy[0] = r + 1
# Labels are suffixed so they are DISTINCT from the dollar rows above. Two rows
# on one sheet sharing a label is how a lookup silently reads dollars as a
# percentage - which is exactly what happened when this block was first read.
for key, label in [("s1a", "Stream 1a: SIP entry fee - share"),
                   ("s1b", "Stream 1b: Spot entry fee - share"),
                   ("s2", "Stream 2: Card interchange - share"),
                   ("s3", "Stream 3: Family and Will - share"),
                   ("s4", "Stream 4: Cardholder fees - share"),
                   ("s5", "Stream 5: Lending - share"),
                   ("s6", "Stream 6: B2B platform - share")]:
    rr = _sy[0]
    ws_summ["A%d" % rr] = label
    for y in range(1, N_YEARS + 1):
        c = ws_summ["%s%d" % (ycol(y), rr)]
        c.value = "=IF(%s$%d=0,0,%s%d/%s$%d)" % (ycol(y), TOT_ROW, ycol(y),
                                                 STREAM_ROWS[key], ycol(y), TOT_ROW)
        c.font, c.number_format = BLACK, FMT_PCT
    _sy[0] += 1
rr = _sy[0]
ws_summ["A%d" % rr] = "CHECK: mix sums to 100%"
ws_summ["A%d" % rr].font = BLACK_BOLD
for y in range(1, N_YEARS + 1):
    c = ws_summ["%s%d" % (ycol(y), rr)]
    c.value = "=SUM(%s%d:%s%d)" % (ycol(y), _sy[0] - 7, ycol(y), _sy[0] - 1)
    c.font, c.number_format = BLACK_BOLD, FMT_PCT
MIX_CHECK_ROW = rr
_sy[0] += 2

s_block("COST BASE", [
    ("grams_fabricated", "  Net new grams fabricated", "sum", FMT_NUM, False),
    ("cogs_premium", "  Fabrication premium paid to the dealer", "sum", FMT_USD, False),
    ("total_cogs", "TOTAL COST OF GOODS SOLD", "sum", FMT_USD, True),
    ("opex_storage", "  Vault storage", "sum", FMT_USD, False),
    ("storage_pct", "    ...effective rate on metal held", "last", FMT_PCT3, False),
    ("opex_vara", "  VARA annual supervision", "sum", FMT_USD, False),
    ("opex_dmcc", "  DMCC company licence", "sum", FMT_USD, False),
    ("opex_kyc", "  KYC and AML verification", "sum", FMT_USD, False),
    ("opex_insurance", "  Insurance (PI, D&O, crime)", "sum", FMT_USD, False),
    ("opex_audit", "  Audit and reserve attestation", "sum", FMT_USD, False),
    ("opex_techaudit", "  Technology audit and penetration testing", "sum", FMT_USD, False),
    ("opex_redemption", "  Redemption handling", "sum", FMT_USD, False),
    ("opex_techbuild", "  Technology build (Y1-Y2)", "sum", FMT_USD, False),
    ("opex_techmaint", "  Technology maintenance (Y3+)", "sum", FMT_USD, False),
    ("opex_oneoff", "  One-off Year 1 (licence, incorporation, launch audit)", "sum", FMT_USD, False),
    ("total_opex", "TOTAL OPERATING EXPENSES", "sum", FMT_USD, True),
    ("ben_entry", "  ICS - entry fee discount (SIP only)", "sum", FMT_USD, False),
    ("ben_card", "  ICS - card fee discounts", "sum", FMT_USD, False),
    ("ben_rebate", "  ICS - gold rebate", "sum", FMT_USD, False),
    ("ben_family", "  ICS - family wallet and will", "sum", FMT_USD, False),
    ("total_benefits", "TOTAL ICS BENEFIT COSTS", "sum", FMT_USD, True),
    ("acq_marketing", "  Marketing spend", "sum", FMT_USD, False),
    ("acq_agent", "  Agent commission", "sum", FMT_USD, False),
    ("acq_referral", "  Referral rewards", "sum", FMT_USD, False),
    ("total_acq", "TOTAL ACQUISITION COSTS", "sum", FMT_USD, True),
    ("card_fixed", "  Card - NymCard platform and scheme fees", "sum", FMT_USD, False),
    ("card_setup", "  Card - programme setup", "sum", FMT_USD, False),
    ("card_production", "  Card - production and delivery", "sum", FMT_USD, False),
    ("card_processing", "  Card - authorisation and switching", "sum", FMT_USD, False),
    ("card_fraud", "  Card - fraud and chargebacks", "sum", FMT_USD, False),
    ("card_xborder", "  Card - cross-border scheme assessment", "sum", FMT_USD, False),
    ("total_card", "TOTAL CARD PROGRAMME COSTS", "sum", FMT_USD, True),
    ("card_cost_ratio", "    Card cost as % of card revenue", "last", FMT_PCT, False),
    ("cac_blended", "    Blended cost per new customer", "last", FMT_USD2, False),
    ("cost_measured", "  Sub-total - costs actually modelled", "sum", FMT_USD, False),
    ("cost_contingency", "  Contingency for families not yet built", "sum", FMT_USD, False),
    ("total_cost", "TOTAL COST BASE", "sum", FMT_USD, True),
])

s_block("NET PROFIT (upper bound - headcount not yet in the cost base)", [
    ("net_profit", "NET PROFIT", "sum", FMT_USD, True),
    ("net_margin", "  Net margin", "last", FMT_PCT, False),
    ("cum_profit", "CUMULATIVE NET PROFIT", "last", FMT_USD, True),
])

s_block("UNIT ECONOMICS (derived - no new assumptions)", [
    ("profit_preacq", "PROFIT BEFORE ACQUISITION COST", "sum", FMT_USD, True),
    ("contrib_per_cust", "  Contribution per paying customer (annualised)", "last", FMT_USD2, False),
    ("attrition", "  Customers lost to holders", "sum", FMT_NUM, False),
    ("cac_blended", "  Blended CAC (repeated here so payback can be read)", "last", FMT_USD2, False),
    ("payback_months", "PAYBACK ON CONTRIBUTION (months)", "last", FMT_NUM2, True),
    ("new_per_net", "NEW CUSTOMERS ACQUIRED PER NET CUSTOMER ADDED", "last", FMT_NUM2, True),
    ("funding_peak", "PEAK FUNDING NEED (P&L deficit + capital tied up)", "last", FMT_USD, True),
])

s_block("REGULATORY CAPITAL (locked, not spent)", [
    ("cap_issuance", "  Paid-up capital - VA Issuance", "last", FMT_USD, False),
    ("cap_activity", "  Paid-up capital - licensed activities", "last", FMT_USD, False),
    ("cap_total", "TOTAL PAID-UP CAPITAL REQUIRED", "last", FMT_USD, True),
    ("nla_required", "  Net liquid assets required (separate test)", "last", FMT_USD, False),
    ("card_prefund", "  Card settlement prefunding", "last", FMT_USD, False),
    ("capital_tied_up", "TOTAL CAPITAL TIED UP (float + capital + prefunding)", "last", FMT_USD, True),
])

s_block("WORKING CAPITAL - THE FLOAT (not a cost)", [
    ("float_grams", "  Float required, grams", "last", FMT_NUM, False),
    ("float_usd", "  FLOAT REQUIRED (year-end balance)", "last", FMT_USD, True),
    ("float_topup", "  Cash needed that year (new grams only)", "sum", FMT_USD, False),
    ("float_cum", "CUMULATIVE CASH INVESTED IN THE FLOAT", "last", FMT_USD, True),
])

s_block("CUSTOMERS AND AUM (year end)", [
    ("new_total", "New customers in year", "sum", FMT_NUM, False),
    ("paying", "Paying customers", "close", FMT_NUM, True),
    ("holders", "Holders (stopped paying, still hold gold)", "close", FMT_NUM, False),
    ("cum_ever", "Cumulative ever acquired", "close", FMT_NUM, False),
    ("active_cards", "Active cards", "close", FMT_NUM, False),
    ("grams_custody", "Grams under custody", "close", FMT_NUM, False),
    ("gold_custody", "Gold under custody (comparable headline)", "close", FMT_USD, True),
    ("grams_held", "Grams held (collateral-eligible)", "close", FMT_NUM, False),
    ("aum", "Collateral-eligible AUM", "close", FMT_USD, True),
    ("rev_per_cust", "Net revenue per paying customer (annualised)", "close", FMT_USD, False),
])
note(ws_summ, _sy[0],
     "REVENUE ONLY. Operating costs, tax, working capital, cash and funding are added in a later build, so "
     "there is no profit or break-even line here. Stream 1 IS reported net of the fabrication premium, "
     "because that is cost OF REVENUE rather than an operating cost - reporting it gross would read ~43% high.")

# ============================================================================
registered, rejected = 0, []
for nm, ref in sorted(NAMED):
    try:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))
        registered += 1
    except Exception as exc:                              # noqa: BLE001
        rejected.append((nm, ref, str(exc)))
wb.save(OUTPUT)

print("=" * 78)
print("AURUMIX REVENUE MODEL  -  simplified single build")
print("=" * 78)
print("Output       : %s" % OUTPUT)
print("Sheets       : %s" % " | ".join(wb.sheetnames))
print("Periods      : %d (%d monthly + %d annual), columns %s..%s, M1 = Jan %d"
      % (N_PERIODS, N_MONTHLY, N_ANNUAL, pcol(0), pcol(N_PERIODS - 1), START_YEAR))
print("Rows         : Assumptions %d | Scenario %d | Model %d | Summary %d"
      % (_ar[0], _sr[0], _mr[0], _sy[0]))
print("Scenario     : %d parameters + %d structural switches" % (len(SROW) - len(SWROW), len(SWROW)))
print("Named ranges : %d registered, %d rejected" % (registered, len(rejected)))
for nm, ref, exc in rejected:
    print("   REJECTED %-28s %-36s %s" % (nm, ref, exc))
print("-" * 78)
print("Engine       : rolling balance, opening + new - churned = closing, by region")
print("Kept as inputs, not engines:")
print("  - card eligibility (%% who ever qualify, months to qualify)")
print("  - the HOLDERS balance (stopped paying, still hold gold)")
print("  - the fabrication premium, now borne by AURUMIX as a real COGS line")
print("    (switchable: sw_premium_absorbed=0 restores the 2026-08-21 treatment,")
print("     where the customer bears it in grams and this cost falls to zero)")
print("  - the ATM draw distribution (a mean returns exactly zero)")
print("Deleted to the Phase 5 simulation: archetypes, run-of-6 chain, lifecycle")
print("  curves, the convolution, the six-state machine, withdrawal buckets.")
print("Cost base    : COGS + opex + ICS benefits + acquisition + card programme.")
print("  Acquisition, card programme, benefit costs and tax still to come;")
print("  each appends to COGS_LINES or OPEX_LINES and the totals follow.")
print("Out of scope for now: tax, working capital, cash, funding.")
print("=" * 78)
